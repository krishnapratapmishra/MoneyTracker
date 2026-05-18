from flask import Flask, render_template, request, jsonify, send_from_directory, send_file, session, redirect, url_for
from functools import wraps
import sqlite3, os, tempfile, urllib.request, urllib.parse, json as _json
from datetime import datetime
import pandas as pd
from werkzeug.utils import secure_filename

try:
    import yfinance as yf
    YF_AVAILABLE = True
except ImportError:
    YF_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'um-local-secret-2026')

# Single-user credentials — username fixed, password persisted in DB
APP_USERNAME     = os.environ.get('APP_USERNAME', 'admin')
_DEFAULT_PASSWORD = os.environ.get('APP_PASSWORD', 'Universe')

def _get_password():
    """Read current password from DB; fall back to default if not set."""
    try:
        conn = get_db()
        row  = conn.execute("SELECT value FROM app_settings WHERE key='app_password'").fetchone()
        conn.close()
        return row[0] if row else _DEFAULT_PASSWORD
    except Exception:
        return _DEFAULT_PASSWORD

def _set_password(new_pwd, conn=None):
    """Persist new password to DB. Pass an existing conn to avoid a second connection."""
    if conn is not None:
        conn.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('app_password', ?)", (new_pwd,))
        return
    c = get_db()
    c.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('app_password', ?)", (new_pwd,))
    c.commit()
    c.close()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated

@app.before_request
def require_login():
    public = {'/login', '/logout', '/api/profile/avatar'}
    if request.path in public or request.path.startswith('/static/'):
        return
    if not session.get('logged_in'):
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Unauthorized'}), 401
        return redirect(url_for('login', next=request.path))

DB_PATH = os.path.join(os.path.dirname(__file__), 'money_tracker.db')

def r2(v):
    """Round a float to 2 decimal places. Safe for None/falsy values."""
    try: return round(float(v or 0), 2)
    except (TypeError, ValueError): return 0.0

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn

def init_schema():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL CHECK(type IN ('income','expense','investment')),
            category TEXT NOT NULL, sub_category TEXT, amount REAL NOT NULL,
            date TEXT NOT NULL, note TEXT, created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS loans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month TEXT NOT NULL, loan_type TEXT NOT NULL, amount REAL NOT NULL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS portfolio (
            AssetID       TEXT PRIMARY KEY REFERENCES InvestMapping(AssetID),
            InvestedValue REAL DEFAULT 0,
            CurrentValue  REAL DEFAULT 0,
            ReturnValue   REAL DEFAULT 0,
            Purpose       TEXT,
            ReturnPCT     REAL DEFAULT 0,
            UpdateAt      TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS invest_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT, entry_date TEXT, stock_name TEXT,
            asset_type TEXT, quantity REAL, action TEXT, price REAL,
            invested_value REAL, current_value REAL, profit REAL,
            profit_pct REAL, rationale TEXT, month TEXT
        );
        CREATE TABLE IF NOT EXISTS alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, type TEXT NOT NULL,
            condition TEXT NOT NULL, threshold REAL NOT NULL, category TEXT,
            period TEXT DEFAULT 'monthly', is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS nse_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            symbol TEXT NOT NULL UNIQUE,
            company_name TEXT DEFAULT '',
            ltp REAL DEFAULT 0,
            prev_close REAL DEFAULT 0,
            change_pct REAL DEFAULT 0,
            high_52w REAL DEFAULT 0,
            low_52w REAL DEFAULT 0,
            from_52w_high_pct REAL DEFAULT 0,
            volume INTEGER DEFAULT 0,
            sector TEXT DEFAULT '',
            category TEXT DEFAULT 'Shares',
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS loan_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            loan_name TEXT NOT NULL,
            loan_type TEXT NOT NULL,
            loan_amount REAL NOT NULL DEFAULT 0,
            total_repayment REAL NOT NULL DEFAULT 0,
            start_date TEXT NOT NULL,
            target_close_date TEXT NOT NULL,
            status TEXT DEFAULT 'active',
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS monthly_investment_calc (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            month TEXT NOT NULL,
            symbol TEXT NOT NULL,
            asset_type TEXT DEFAULT '',
            qty_bought REAL DEFAULT 0,
            qty_sold REAL DEFAULT 0,
            net_qty REAL DEFAULT 0,
            avg_buy_price REAL DEFAULT 0,
            total_invested REAL DEFAULT 0,
            current_price REAL DEFAULT 0,
            current_value REAL DEFAULT 0,
            unrealized_pnl REAL DEFAULT 0,
            unrealized_pnl_pct REAL DEFAULT 0,
            updated_at TEXT DEFAULT (datetime('now')),
            UNIQUE(month, symbol)
        );
        CREATE TABLE IF NOT EXISTS magnet_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            magnet TEXT NOT NULL,
            metric_name TEXT NOT NULL,
            emoji TEXT DEFAULT '📌',
            current_value TEXT,
            target_value TEXT,
            note TEXT,
            recorded_date TEXT DEFAULT (date('now')),
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS um_vision_cards (
            id TEXT PRIMARY KEY,
            magnet TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            photo_data TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS um_vision_cards_default (
            id TEXT PRIMARY KEY,
            magnet TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            photo_data TEXT,
            vision_type TEXT DEFAULT 'general',
            created_at TEXT,
            updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS assets (
            AssetEntryID  INTEGER PRIMARY KEY AUTOINCREMENT,
            MappingID     INTEGER NOT NULL REFERENCES AssetMapping(MappingID),
            purpose       TEXT,
            qty           REAL DEFAULT 0,
            avgprice      REAL DEFAULT 0,
            ltp           REAL DEFAULT 0,
            investedvalue REAL DEFAULT 0,
            currentvalue  REAL DEFAULT 0,
            pnl           REAL DEFAULT 0,
            pnlpct        REAL DEFAULT 0,
            lastsynced    TEXT,
            updatedat     TEXT DEFAULT (datetime('now')),
            targetpct     REAL DEFAULT 25,
            sip_level_pct REAL DEFAULT 10,
            max_allocation REAL DEFAULT 50000
        );
        CREATE TABLE IF NOT EXISTS wealth (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purpose TEXT UNIQUE NOT NULL,
            target REAL NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS raw_upload_meta (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            source_type TEXT    NOT NULL,
            file_name   TEXT,
            row_count   INTEGER DEFAULT 0,
            uploaded_at TEXT    DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS raw_upload_data (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id   INTEGER NOT NULL REFERENCES raw_upload_meta(id) ON DELETE CASCADE,
            source_type TEXT    NOT NULL,
            broker      TEXT    NOT NULL,
            instrument  TEXT    NOT NULL,
            data_type   TEXT    NOT NULL,
            name        TEXT,
            symbol      TEXT,
            isin        TEXT,
            trade_type  TEXT,
            trade_date  TEXT,
            quantity    REAL,
            price       REAL,
            amount      REAL,
            exchange    TEXT,
            order_id    TEXT,
            status      TEXT,
            raw_data    TEXT,
            uploaded_at TEXT    DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS InvestMapping (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            AssetClass   TEXT NOT NULL,
            AssetCategory TEXT NOT NULL,
            AssetType    TEXT NOT NULL,
            UNIQUE(AssetClass, AssetCategory, AssetType)
        );
    """)
    conn.commit()

    # Seed InvestMapping reference data (insert if not already present)
    _INVEST_MAPPING_SEED = [
        ('Commodities', 'Gold',           'Bonds'),
        ('Commodities', 'Gold',           'Physical Gold'),
        ('Commodities', 'Gold',           'Digital Gold'),
        ('Commodities', 'Gold',           'Mutual Fund'),
        ('Commodities', 'Gold',           'ETF'),
        ('Commodities', 'Silver',         'ETF'),
        ('Growth',      'Equity',         'Stocks'),
        ('Growth',      'Equity',         'Mutual Fund'),
        ('Growth',      'Equity',         'ETF'),
        ('Liquidity',   'Emergency Fund', 'Savings Account'),
        ('Liquidity',   'Emergency Fund', 'Liquid Mutual Fund'),
        ('Liquidity',   'Emergency Fund', 'Short-term FD'),
        ('Real State',  'Real Estate',    'Plot'),
        ('Real State',  'Real Estate',    'Flat'),
        ('Retirement',  'Fixed Income',   'Employer PF'),
        ('Retirement',  'Hybrid',         'Pension Scheme'),
        ('Stability',   'Fixed Income',   'PPF'),
        ('Stability',   'Fixed Income',   'Employee PF'),
        ('Stability',   'Fixed Income',   'EPF'),
        ('Stability',   'Fixed Income',   'Government Bond/Yojana'),
    ]
    conn.executemany(
        "INSERT OR IGNORE INTO InvestMapping (AssetClass, AssetCategory, AssetType) VALUES (?,?,?)",
        _INVEST_MAPPING_SEED
    )
    conn.commit()

    # Add AssetID column to InvestMapping (formatted primary key: Asset01, Asset02, …)
    try:
        conn.execute("ALTER TABLE InvestMapping ADD COLUMN AssetID TEXT")
        conn.commit()
    except Exception:
        pass  # Column already exists

    # Backfill AssetID for any row where it is NULL, using next sequential number
    # based on MAX of existing numeric suffixes (not the raw id, which may have gaps)
    conn.execute("""
        UPDATE InvestMapping
        SET AssetID = 'Asset' || printf('%02d',
            (SELECT COALESCE(MAX(CAST(SUBSTR(AssetID, 6) AS INTEGER)), 0) + 1
             FROM InvestMapping WHERE AssetID GLOB 'Asset[0-9]*')
            + (SELECT COUNT(*) FROM InvestMapping im2
               WHERE im2.id < InvestMapping.id AND (im2.AssetID IS NULL OR im2.AssetID = ''))
        )
        WHERE AssetID IS NULL OR AssetID = ''
    """)
    conn.commit()

    # Trigger: auto-set AssetID on every new INSERT using MAX(numeric suffix)+1
    conn.execute("DROP TRIGGER IF EXISTS trg_investmapping_assetid")
    conn.execute("""
        CREATE TRIGGER trg_investmapping_assetid
        AFTER INSERT ON InvestMapping
        FOR EACH ROW
        WHEN NEW.AssetID IS NULL OR NEW.AssetID = ''
        BEGIN
            UPDATE InvestMapping
            SET AssetID = 'Asset' || printf('%02d',
                COALESCE((SELECT MAX(CAST(SUBSTR(AssetID, 6) AS INTEGER))
                          FROM InvestMapping WHERE AssetID GLOB 'Asset[0-9]*'), 0) + 1
            )
            WHERE id = NEW.id;
        END
    """)
    conn.commit()

    # Rebuild AssetMapping with new structure (MappingID PK) if old schema exists
    am_cols = {row[1] for row in conn.execute("PRAGMA table_info(AssetMapping)").fetchall()}
    if 'MappingID' not in am_cols:
        conn.execute("DROP TABLE IF EXISTS AssetMapping")
        conn.execute("""
            CREATE TABLE AssetMapping (
                MappingID   INTEGER PRIMARY KEY AUTOINCREMENT,
                AssetName   TEXT NOT NULL,
                AssetSymbol TEXT DEFAULT '',
                AssetId     TEXT NOT NULL REFERENCES InvestMapping(AssetID)
            )
        """)
        conn.commit()

    # Drop old assets table if it uses old schema (id column instead of AssetEntryID)
    assets_cols = {row[1] for row in conn.execute("PRAGMA table_info(assets)").fetchall()}
    if assets_cols and 'AssetEntryID' not in assets_cols:
        conn.execute("DROP TABLE IF EXISTS assets")
        conn.execute("""
            CREATE TABLE assets (
                AssetEntryID  INTEGER PRIMARY KEY AUTOINCREMENT,
                MappingID     INTEGER NOT NULL REFERENCES AssetMapping(MappingID),
                purpose       TEXT,
                qty           REAL DEFAULT 0,
                avgprice      REAL DEFAULT 0,
                ltp           REAL DEFAULT 0,
                investedvalue REAL DEFAULT 0,
                currentvalue  REAL DEFAULT 0,
                pnl           REAL DEFAULT 0,
                pnlpct        REAL DEFAULT 0,
                lastsynced    TEXT,
                updatedat     TEXT DEFAULT (datetime('now')),
                targetpct     REAL DEFAULT 25,
                sip_level_pct REAL DEFAULT 10,
                max_allocation REAL DEFAULT 50000
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS gtt_params (
                symbol        TEXT PRIMARY KEY,
                target_pct    REAL DEFAULT 25,
                sip_level_pct REAL DEFAULT 10,
                max_allocation REAL DEFAULT 50000,
                updated_at    TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.commit()

    # Drop old portfolio table if it has old schema (integer PK / asset column)
    port_cols = {row[1] for row in conn.execute("PRAGMA table_info(portfolio)").fetchall()}
    if port_cols and 'AssetID' not in port_cols:
        conn.execute("DROP TABLE IF EXISTS portfolio")
        conn.execute("""
            CREATE TABLE portfolio (
                AssetID       TEXT PRIMARY KEY REFERENCES InvestMapping(AssetID),
                InvestedValue REAL DEFAULT 0,
                CurrentValue  REAL DEFAULT 0,
                ReturnValue   REAL DEFAULT 0,
                Purpose       TEXT,
                ReturnPCT     REAL DEFAULT 0,
                UpdateAt      TEXT DEFAULT (datetime('now'))
            )
        """)
        conn.commit()

    # Migrate: add columns to other non-portfolio tables if missing
    for migration in [
        "ALTER TABLE nse_master ADD COLUMN category TEXT DEFAULT 'Shares'",
        "ALTER TABLE wealth ADD COLUMN target_date TEXT",
        "ALTER TABLE wealth ADD COLUMN linked_asset_ids TEXT DEFAULT ''",  # JSON array of AssetIDs
        # InvestMapping: price-fetch configuration columns
        "ALTER TABLE InvestMapping ADD COLUMN PriceFetchMode TEXT DEFAULT 'MANUAL'",
        "ALTER TABLE InvestMapping ADD COLUMN Symbol        TEXT DEFAULT ''",
        "ALTER TABLE InvestMapping ADD COLUMN WeightGrams   REAL DEFAULT 1",
        "ALTER TABLE InvestMapping ADD COLUMN Purity        TEXT DEFAULT '24K'",
        "ALTER TABLE InvestMapping ADD COLUMN InterestRate  REAL DEFAULT 0",
        # assets: denormalised name + symbol for quick reference
        "ALTER TABLE assets ADD COLUMN assetname TEXT DEFAULT ''",
        "ALTER TABLE assets ADD COLUMN symbol     TEXT DEFAULT ''",
        # assets: per-asset GTT strategy parameters
        "ALTER TABLE assets ADD COLUMN sip_level_pct  REAL DEFAULT 10",
        "ALTER TABLE assets ADD COLUMN max_allocation  REAL DEFAULT 50000",
        # Vision board: type-aware vision cards
        "ALTER TABLE um_vision_cards ADD COLUMN vision_type TEXT DEFAULT 'general'",
        # Vision board: type-aware status entries
        "ALTER TABLE magnet_status ADD COLUMN vision_id   TEXT",
        "ALTER TABLE magnet_status ADD COLUMN vision_type TEXT DEFAULT 'general'",
        "ALTER TABLE magnet_status ADD COLUMN feel        TEXT",
        "ALTER TABLE magnet_status ADD COLUMN cost        REAL",
        "ALTER TABLE magnet_status ADD COLUMN trip_name   TEXT",
        "ALTER TABLE magnet_status ADD COLUMN start_date  TEXT",
        "ALTER TABLE magnet_status ADD COLUMN end_date    TEXT",
    ]:
        try:
            conn.execute(migration)
            conn.commit()
        except Exception:
            pass

    # Ensure gtt_params table exists (created unconditionally for existing DBs)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS gtt_params (
            symbol        TEXT PRIMARY KEY,
            target_pct    REAL DEFAULT 25,
            sip_level_pct REAL DEFAULT 10,
            max_allocation REAL DEFAULT 50000,
            updated_at    TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.commit()

    # Ensure InvestMapping.AssetID has a UNIQUE index so foreign-key references
    # from AssetMapping and portfolio are valid in strict FK mode (DB Browser, etc.)
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_investmapping_assetid
        ON InvestMapping(AssetID)
    """)
    conn.commit()

    # Backfill PriceFetchMode + defaults for the 15 seed InvestMapping rows
    # Uses (AssetClass, AssetCategory, AssetType) as natural key
    # Only updates rows where PriceFetchMode is still NULL / empty / 'MANUAL'
    # so user-edited rows are not overwritten.
    _IM_DEFAULTS = [
        # (AssetClass, AssetCategory, AssetType,        Mode,              Symbol, WeightGrams, Purity,  Rate)
        ('Commodities','Gold',        'Bonds',           'EQUITY',          '',     1,           '24K',   0),
        ('Commodities','Gold',        'Physical Gold',   'COMMODITY_GOLD',  '',     1,           '24K',   0),
        ('Commodities','Gold',        'Digital Gold',    'COMMODITY_GOLD',  '',     1,           '24K',   0),
        ('Commodities','Gold',        'Mutual Fund',     'MF',              '',     1,           '24K',   0),
        ('Commodities','Gold',        'ETF',             'EQUITY',          '',     1,           '24K',   0),
        ('Commodities','Silver',      'ETF',             'EQUITY',          '',     1,           '24K',   0),
        ('Growth',     'Equity',      'Stocks',          'EQUITY',          '',     1,           '24K',   0),
        ('Growth',     'Equity',      'Mutual Fund',     'MF',              '',     1,           '24K',   0),
        ('Growth',     'Equity',      'ETF',             'EQUITY',          '',     1,           '24K',   0),
        ('Liquidity',  'Emergency Fund','Savings Account', 'RATE_BASED',      '',     1,           '24K',   3.5),
        ('Liquidity',  'Emergency Fund','Liquid Mutual Fund','MF',           '',     1,           '24K',   0),
        ('Liquidity',  'Emergency Fund','Short-term FD',  'RATE_BASED',      '',     1,           '24K',   7.0),
        ('Real State', 'Real Estate', 'Plot',            'MANUAL',          '',     1,           '24K',   0),
        ('Real State', 'Real Estate', 'Flat',            'MANUAL',          '',     1,           '24K',   0),
        ('Retirement', 'Fixed Income','Employer PF',     'RATE_BASED',      '',     1,           '24K',   8.25),
        ('Retirement', 'Hybrid',      'Pension Scheme',  'RATE_BASED',      '',     1,           '24K',   10.0),
        ('Stability',  'Fixed Income','PPF',             'RATE_BASED',      '',     1,           '24K',   7.1),
        ('Stability',  'Fixed Income','Employee PF',     'RATE_BASED',      '',     1,           '24K',   8.25),
        ('Stability',  'Fixed Income','EPF',             'RATE_BASED',      '',     1,           '24K',   8.25),
        ('Stability',  'Fixed Income','Government Bond/Yojana', 'RATE_BASED','',   1,           '24K',   8.2),
    ]
    for row in _IM_DEFAULTS:
        cls, cat, typ, mode, sym, wt, purity, rate = row
        conn.execute("""
            UPDATE InvestMapping
            SET PriceFetchMode = ?,
                Symbol         = CASE WHEN (Symbol IS NULL OR Symbol = '') THEN ? ELSE Symbol END,
                WeightGrams    = CASE WHEN (WeightGrams IS NULL OR WeightGrams = 0) THEN ? ELSE WeightGrams END,
                Purity         = CASE WHEN (Purity IS NULL OR Purity = '') THEN ? ELSE Purity END,
                InterestRate   = CASE WHEN (InterestRate IS NULL OR InterestRate = 0) THEN ? ELSE InterestRate END
            WHERE AssetClass=? AND AssetCategory=? AND AssetType=?
              AND (PriceFetchMode IS NULL OR PriceFetchMode = '' OR PriceFetchMode = 'MANUAL')
        """, (mode, sym, wt, purity, rate, cls, cat, typ))
    # Force-set rates for RATE_BASED rows even if already RATE_BASED (so corrections are applied)
    _RATE_UPDATES = [
        ('Retirement', 'Fixed Income','Employer PF',      8.25),
        ('Retirement', 'Hybrid',      'Pension Scheme',  10.0),
        ('Stability',  'Fixed Income','PPF',              7.1),
        ('Stability',  'Fixed Income','Employee PF',      8.25),
        ('Stability',  'Fixed Income','EPF',              8.25),
        ('Stability',  'Fixed Income','Government Bond/Yojana', 8.2),
    ]
    for cls, cat, typ, rate in _RATE_UPDATES:
        conn.execute("""
            UPDATE InvestMapping SET InterestRate=?
            WHERE AssetClass=? AND AssetCategory=? AND AssetType=?
              AND (InterestRate IS NULL OR InterestRate = 0)
        """, (rate, cls, cat, typ))
    conn.commit()

    # ── DB Cleanup: drop unused tables ────────────────────────────────────────
    conn.executescript("""
        DROP TABLE IF EXISTS budgets;
        DROP TABLE IF EXISTS raw_transactions;
    """)
    conn.commit()

    # ── Schema additions: new cross-link columns ───────────────────────────────
    for migration in [
        "ALTER TABLE transactions ADD COLUMN source_ref TEXT",
        "ALTER TABLE transactions ADD COLUMN auto_created INTEGER DEFAULT 0",
        "ALTER TABLE loans ADD COLUMN txn_id INTEGER",
        "ALTER TABLE assets ADD COLUMN onboarding_seed INTEGER DEFAULT 0",
        # alerts: columns added after initial release
        "ALTER TABLE alerts ADD COLUMN period TEXT DEFAULT 'monthly'",
        "ALTER TABLE alerts ADD COLUMN is_active INTEGER DEFAULT 1",
        "ALTER TABLE alerts ADD COLUMN created_at TEXT DEFAULT (datetime('now'))",
        # loan_master: columns added after initial release
        "ALTER TABLE loan_master ADD COLUMN loan_amount REAL DEFAULT 0",
        "ALTER TABLE loan_master ADD COLUMN total_repayment REAL DEFAULT 0",
        "ALTER TABLE loan_master ADD COLUMN start_date TEXT DEFAULT ''",
        "ALTER TABLE loan_master ADD COLUMN target_close_date TEXT DEFAULT ''",
        "ALTER TABLE loan_master ADD COLUMN status TEXT DEFAULT 'active'",
        # invest_transactions: month column added later
        "ALTER TABLE invest_transactions ADD COLUMN month TEXT",
    ]:
        try:
            conn.execute(migration)
            conn.commit()
        except Exception:
            pass  # Column already exists — safe to ignore

    # ── Ensure app_settings table exists ──────────────────────────────────────
    conn.execute("""
        CREATE TABLE IF NOT EXISTS app_settings (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    """)
    conn.commit()

    # ── Fix AUTOBEES: remap to Asset09 (ETF) and ensure it's in nse_master ────
    conn.execute("""
        UPDATE AssetMapping SET AssetId='Asset09'
        WHERE AssetName='AUTOBEES' AND AssetId='Asset07'
    """)
    conn.execute("""
        INSERT OR IGNORE INTO nse_master (symbol, category)
        VALUES ('AUTOBEES', 'ETF')
    """)
    # Also fix GOLDBEES which was wrongly in Asset07 (Stocks); it belongs in
    # a commodity-ETF mapping — Asset09 is Growth/ETF so keep it there for LTP
    # sync purposes (the frontend commodity check already routes it correctly)
    conn.execute("""
        INSERT OR IGNORE INTO nse_master (symbol, category)
        VALUES ('GOLDBEES', 'ETF')
    """)
    conn.execute("""
        UPDATE nse_master SET category='ETF'
        WHERE symbol='GOLDBEES' AND (category IS NULL OR category='Shares')
    """)
    conn.commit()

    # ── Demat / Stock Trading Tables ──────────────────────────────────────────
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS demat_wallet (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            txn_date   TEXT NOT NULL,
            type       TEXT NOT NULL,
            amount     REAL NOT NULL,
            note       TEXT DEFAULT '',
            ref_txn_id INTEGER DEFAULT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS stock_holdings (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            symbol        TEXT NOT NULL,
            isin          TEXT DEFAULT '',
            trade_type    TEXT NOT NULL,
            buy_date      TEXT NOT NULL,
            qty_original  REAL NOT NULL,
            qty_remaining REAL NOT NULL,
            buy_price     REAL NOT NULL,
            lot_source    TEXT DEFAULT 'TRADEBOOK',
            created_at    TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS stock_transactions (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            trade_date       TEXT NOT NULL,
            symbol           TEXT NOT NULL,
            isin             TEXT DEFAULT '',
            action           TEXT NOT NULL,
            qty              REAL NOT NULL,
            price            REAL DEFAULT 0,
            trade_type       TEXT NOT NULL,
            exchange         TEXT DEFAULT 'NSE',
            stt              REAL DEFAULT 0,
            other_charges    REAL DEFAULT 0,
            total_charges    REAL DEFAULT 0,
            net_amount       REAL DEFAULT 0,
            zerodha_order_id TEXT DEFAULT '',
            lot_ids_affected TEXT DEFAULT '[]',
            lot_source       TEXT DEFAULT 'TRADEBOOK',
            created_at       TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS stock_dividends (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            symbol       TEXT NOT NULL,
            date         TEXT NOT NULL,
            per_share    REAL DEFAULT 0,
            total_amount REAL NOT NULL,
            created_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS stock_pnl (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            sell_txn_id   INTEGER,
            symbol        TEXT NOT NULL,
            sell_date     TEXT NOT NULL,
            trade_type    TEXT NOT NULL,
            qty_sold      REAL NOT NULL,
            avg_buy_price REAL NOT NULL,
            sell_price    REAL NOT NULL,
            gross_pnl     REAL NOT NULL,
            charges       REAL DEFAULT 0,
            net_pnl       REAL NOT NULL,
            holding_days  INTEGER DEFAULT 0,
            tax_category  TEXT DEFAULT 'STCG',
            created_at    TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit()

    # Ensure app_settings table exists (may not exist on first run if assets table was new)
    conn.execute("""CREATE TABLE IF NOT EXISTS app_settings (
        key TEXT PRIMARY KEY, value TEXT NOT NULL)""")
    # Ensure onboarding_complete flag exists in app_settings
    conn.execute("INSERT OR IGNORE INTO app_settings (key, value) VALUES ('demat_onboarded','0')")
    # Migrate: normalize legacy 'Long Term' trade_type to 'LONG'
    try:
        conn.execute("UPDATE stock_holdings SET trade_type='LONG' WHERE trade_type='Long Term'")
        conn.execute("UPDATE stock_transactions SET trade_type='LONG' WHERE trade_type='Long Term'")
        conn.execute("UPDATE stock_pnl SET trade_type='LONG' WHERE trade_type='Long Term'")
    except Exception:
        pass

    conn.commit()

    conn.close()

# ── Seed data: restored on every factory / fresh reset ────────────────────────
_IM_SEED = [
    (1,   'Commodities', 'Gold',           'Bond',                  'Asset01',   'COMMODITY_GOLD', '',                            1.0, '24K', 0.0),
    (2,   'Commodities', 'Gold',           'Physical Gold',         'Asset02',   'COMMODITY_GOLD', '',                            1.0, '24K', 0.0),
    (3,   'Commodities', 'Gold',           'Digital Gold',          'Asset03',   'COMMODITY_GOLD', '',                            1.0, '24K', 0.0),
    (4,   'Commodities', 'Gold',           'Mutual Fund',           'Asset04',   'MF',             'SBI Gold Direct Plan Growth', 1.0, '24K', 0.0),
    (5,   'Commodities', 'Gold',           'ETF',                   'Asset05',   'EQUITY',         'GOLDBEES',                    1.0, '24K', 0.0),
    (6,   'Commodities', 'Silver',         'ETF',                   'Asset06',   'EQUITY',         'SILVERBEES',                  1.0, '24K', 0.0),
    (7,   'Growth',      'Equity',         'Stocks',                'Asset07',   'EQUITY',         '',                            1.0, '24K', 0.0),
    (8,   'Growth',      'Equity',         'Mutual Fund',           'Asset08',   'MF',             '',                            1.0, '24K', 0.0),
    (9,   'Growth',      'Equity',         'ETF',                   'Asset09',   'EQUITY',         '',                            1.0, '24K', 0.0),
    (10,  'Real State',  'Real Estate',    'Plot',                  'Asset10',   'MANUAL',         '',                            1.0, '24K', 0.0),
    (11,  'Real State',  'Real Estate',    'Flat',                  'Asset11',   'MANUAL',         '',                            1.0, '24K', 0.0),
    (13,  'Stability',   'Fixed Income',   'PPF',                   'Asset13',   'RATE_BASED',     '',                            1.0, '24K', 7.1),
    (15,  'Stability',   'Fixed Income',   'Sukanya Samriddhi',     'Asset15',   'RATE_BASED',     '',                            1.0, '24K', 8.2),
    (837, 'Liquidity',   'Emergency Fund', 'Savings Account',       'Asset837',  'RATE_BASED',     '',                            1.0, '24K', 3.5),
    (838, 'Liquidity',   'Emergency Fund', 'Liquid Mutual Fund',    'Asset838',  'MF',             '',                            1.0, '24K', 0.0),
    (839, 'Liquidity',   'Emergency Fund', 'Short-term FD',         'Asset839',  'RATE_BASED',     '',                            1.0, '24K', 6.0),
    (914, 'Retirement',  'Hybrid',         'Pension Scheme',        'Asset914',  'RATE_BASED',     '',                            1.0, '24K', 10.0),
    (917, 'Stability',   'Fixed Income',   'Government Bond/Yojana','Asset917',  'RATE_BASED',     '',                            1.0, '24K', 8.2),
    (1800,'Commodities', 'Gold',           'Gold Bond',             'Asset1800', 'EQUITY',         '',                            1.0, '24K', 0.0),
    (1962,'Commodities', 'Gold',           'Bonds',                 'Asset1962', 'EQUITY',         '',                            1.0, '24K', 0.0),
    (2574,'Retirement',  'Fixed Income',   'Employer PF',           'Asset2574', 'RATE_BASED',     '',                            1.0, '24K', 8.7),
    (2575,'Stability',   'Fixed Income',   'Employee PF',           'Asset2575', 'RATE_BASED',     '',                            1.0, '24K', 8.7),
    (2592,'Stability',   'Fixed Income',   'EPF',                   'Asset2592', 'RATE_BASED',     '',                            1.0, '24K', 8.25),
    (2692,'Retirement',  'Real Estate',    'Flat',                  'Asset2692', 'MANUAL',         'Flat',                        1.0, '24K', 0.0),
]

_VC_SEED = [
    ('12159342-e898-4b96-bf16-755311156013', 'health',       'Body Vitals in Balance',       'My VITALs are inline and Healthy.',                                                                                                        'milestone'),
    ('12e3031a-7767-4b79-874f-539c43db4b3d', 'health',       '7 hrs Quality Sleep',          'Taking 7 hours of Quality Sleep. 10 PM to 5 AM.',                                                                                          'goal'),
    ('1ce57b83-6975-4fa4-b5be-4f77ee6a7efd', 'health',       'Fasting Freedom',              'Maintaining Fasting Glucose always between 80-70',                                                                                         'goal'),
    ('2cecc5a6-d17d-486b-9eae-1d01cba63a08', 'health',       'Protective Strength (HDL)',    'Maintaining HDL between 50-70',                                                                                                            'goal'),
    ('3ca81e0a-36d0-42a6-9bc9-6b7680136c21', 'health',       'HbA1c Mastery',                'Always Between 5.4 and 5.',                                                                                                                'goal'),
    ('40454e22-4c8b-4b14-8e7d-66ad1d256fe8', 'vision',       'My Villa',                     '“I live happily with my family in our own villa, where friends and relatives are always warmly welcomed.”',              'milestone'),
    ('4cd6e70c-29c4-44b5-a756-cfefa360a860', 'relationship', 'Family Trip',                  '✨ “My family and I enjoy meaningful trips every 6 months, filled with love, bonding, peace, and happy memories.” ✨', 'experience'),
    ('912b4d44-873e-49cc-963b-ccc121a9b641', 'health',       'Lipid Reset',                  'Marinating my Triglycerides level between 130 -110',                                                                                       'goal'),
    ('9866f51b-3a8f-4025-88ae-7d7732c6a952', 'vision',       '5 Source of Income',           '✨“I attract multiple streams of income that consistently generate ₹3 lakh or more in-hand every month.”✨',       'goal'),
    ('a323bb54-5474-4acc-b776-375e227d6401', 'career',       'DevOps Tech Lead',             'Azure DevOps Certified.',                                                                                                                  'milestone'),
    ('b5cfd4b4-b677-41b1-88a7-5309a530b0da', 'health',       'Heart Cleanse (LDL)',          'Maintaining Below 100.',                                                                                                                   'goal'),
    ('bfecd2ec-798b-4c43-a660-2fd2bc2fdac8', 'health',       'EveryDay Gym',                 'Minimum of 30 Mins of exercise every Day',                                                                                                 'habit'),
    ('c3381318-0c2b-401d-adf1-ec42e879005a', 'relationship', 'Trip with Friends',            '✨ “I attract joyful trips filled with laughter, adventure, deep friendships, and unforgettable memories.”✨',           'experience'),
    ('ddc46d29-7eb0-4f91-aaaf-e18c91c11ffb', 'health',       'Body Balance',                 'Always between 65-70 kg',                                                                                                                  'general'),
    ('f462422a-d4b5-438a-9c81-4464c7cb9c56', 'relationship', 'Quality Time With My Partner', '✨ “30 minutes daily. Just us. No mobiles. No distractions. Only love, connection, and togetherness.” ✨',             'habit'),
    ('fcb41ad6-6d0d-4749-9870-5ceefc41dfaf', 'vision',       'My Car',                       'Upgrading My EV to High Range and Best Model',                                                                                             'milestone'),
]

def _restore_seed_data(conn):
    """Re-seed InvestMapping and um_vision_cards after any full/factory reset."""
    conn.execute("DELETE FROM InvestMapping")
    conn.executemany(
        "INSERT INTO InvestMapping (id,AssetClass,AssetCategory,AssetType,AssetID,PriceFetchMode,Symbol,WeightGrams,Purity,InterestRate) VALUES (?,?,?,?,?,?,?,?,?,?)",
        _IM_SEED
    )
    # Restore vision cards from frozen default table (full images intact)
    # Falls back to text-only _VC_SEED only if default table is empty
    conn.execute("DELETE FROM um_vision_cards")
    conn.execute("""
        INSERT OR IGNORE INTO um_vision_cards
            (id, magnet, title, description, photo_data, vision_type, created_at, updated_at)
        SELECT id, magnet, title, description, photo_data,
               COALESCE(vision_type, 'general'),
               COALESCE(created_at, datetime('now')),
               COALESCE(updated_at, datetime('now'))
        FROM um_vision_cards_default
    """)
    if conn.execute("SELECT COUNT(*) FROM um_vision_cards").fetchone()[0] == 0:
        conn.executemany(
            "INSERT OR IGNORE INTO um_vision_cards (id,magnet,title,description,vision_type,created_at) VALUES (?,?,?,?,?,datetime('now'))",
            _VC_SEED
        )


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        stored_pwd  = _get_password()
        valid_names = {APP_USERNAME.lower()}
        try:
            conn = get_db()
            row  = conn.execute("SELECT value FROM app_settings WHERE key='user_name'").fetchone()
            conn.close()
            if row and row['value']:
                valid_names.add(row['value'].strip().lower())
        except Exception:
            pass
        if username.lower() in valid_names and password == stored_pwd:
            session['logged_in'] = True
            next_url = request.args.get('next') or '/'
            return redirect(next_url)
        error = 'Invalid username or password.'
    # Read display name from app_settings (safe — no sensitive data exposed)
    user_name = ''
    try:
        conn = get_db()
        row = conn.execute("SELECT value FROM app_settings WHERE key='user_name'").fetchone()
        conn.close()
        if row and row['value']:
            user_name = row['value'].strip()
    except Exception:
        pass
    return render_template('login.html', error=error, user_name=user_name)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/change_password', methods=['POST'])
def api_change_password():
    body        = request.get_json(force=True)
    current_pwd = body.get('current_password', '')
    new_pwd     = body.get('new_password', '').strip()
    if not current_pwd or not new_pwd:
        return jsonify({'error': 'Both current and new password are required'}), 400
    if current_pwd != _get_password():
        return jsonify({'error': 'Current password is incorrect'}), 403
    if len(new_pwd) < 6:
        return jsonify({'error': 'New password must be at least 6 characters'}), 400
    _set_password(new_pwd)
    return jsonify({'success': True})

@app.route('/api/change_username', methods=['POST'])
@login_required
def api_change_username():
    body     = request.get_json(force=True)
    new_name = (body.get('new_username') or '').strip()
    password = body.get('password', '')
    if not new_name:
        return jsonify({'error': 'Username cannot be empty'}), 400
    if len(new_name) < 3:
        return jsonify({'error': 'Username must be at least 3 characters'}), 400
    if password != _get_password():
        return jsonify({'error': 'Password is incorrect'}), 403
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('user_name',?)", (new_name,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/')
@login_required
def dashboard():
    return render_template('index.html')

@app.route('/palette')
def palette():
    return render_template('palette.html')

@app.route('/docs/<path:filename>')
def serve_docs(filename):
    """Serve project markdown documents (SOP.md, PROJECT_DOCUMENT.md)."""
    allowed = {'SOP.md', 'PROJECT_DOCUMENT.md'}
    if filename not in allowed:
        return jsonify({'error': 'Not found'}), 404
    return send_from_directory(os.path.dirname(os.path.abspath(__file__)), filename)

@app.route('/api/docs/<docname>')
def api_docs(docname):
    """Return markdown file contents as plain text for in-app rendering."""
    allowed = {'sop': 'SOP.md', 'project': 'PROJECT_DOCUMENT.md'}
    fname = allowed.get(docname)
    if not fname:
        return jsonify({'error': 'Unknown document'}), 404
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return f.read(), 200, {'Content-Type': 'text/plain; charset=utf-8'}
    except FileNotFoundError:
        return jsonify({'error': f'{fname} not found'}), 404

@app.route('/api/summary')
def api_summary():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    conn = get_db(); c = conn.cursor()
    def q(sql, *p): return c.execute(sql, p).fetchone()[0]
    income     = q("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='income' AND date LIKE ?",     f'{month}%')
    loan_emi   = q("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='expense' AND category='Loan EMI' AND date LIKE ?", f'{month}%')
    expense    = q("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='expense' AND category!='Loan EMI' AND date LIKE ?", f'{month}%')
    investment = q("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='investment' AND date LIKE ?", f'{month}%')
    savings    = income - expense - loan_emi - investment
    savings_rate = round(savings / income * 100, 2) if income > 0 else 0
    y, m = map(int, month.split('-'))
    pm = m - 1 if m > 1 else 12; py = y if m > 1 else y - 1
    prev = f'{py}-{pm:02d}'
    pi  = q("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='income' AND date LIKE ?",     f'{prev}%')
    pl  = q("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='expense' AND category='Loan EMI' AND date LIKE ?", f'{prev}%')
    pe  = q("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='expense' AND category!='Loan EMI' AND date LIKE ?", f'{prev}%')
    pv  = q("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='investment' AND date LIKE ?", f'{prev}%')
    ps  = pi - pe - pl - pv
    def delta(cur, prv): return round(((cur - prv) / prv) * 100, 2) if prv != 0 else 0
    conn.close()
    return jsonify({
        'income': r2(income), 'expense': r2(expense), 'loan_emi': r2(loan_emi), 'investment': r2(investment),
        'savings': r2(savings), 'savings_rate': savings_rate,
        'delta_income': delta(income, pi), 'delta_expense': delta(expense, pe),
        'delta_loan': delta(loan_emi, pl), 'delta_investment': delta(investment, pv),
        'delta_savings': delta(savings, ps),
    })

@app.route('/api/monthly_trend')
def api_monthly_trend():
    conn = get_db(); c = conn.cursor()
    # All expense rows (to split loan vs non-loan)
    rows = c.execute("""SELECT strftime('%Y-%m',date) as month, type, category,
                               COALESCE(SUM(amount),0) as total
                        FROM transactions GROUP BY month,type,category ORDER BY month""").fetchall()
    conn.close()
    months = sorted(set(r['month'] for r in rows if r['month']))
    data = {m: {'income':0,'expense':0,'loan_emi':0,'investment':0} for m in months}
    for r in rows:
        mo = r['month']
        if not mo: continue
        if r['type'] == 'income':      data[mo]['income']     += r['total']
        elif r['type'] == 'investment': data[mo]['investment'] += r['total']
        elif r['type'] == 'expense':
            if r['category'] == 'Loan EMI': data[mo]['loan_emi'] += r['total']
            else:                           data[mo]['expense']  += r['total']
    for m in months:
        d = data[m]
        d['income']     = r2(d['income'])
        d['expense']    = r2(d['expense'])
        d['loan_emi']   = r2(d['loan_emi'])
        d['investment'] = r2(d['investment'])
        d['savings']    = r2(d['income'] - d['expense'] - d['loan_emi'] - d['investment'])
    return jsonify([{'month': m, **data[m]} for m in months])

@app.route('/api/category_breakdown')
def api_category_breakdown():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""SELECT type, category, COALESCE(SUM(amount),0) as total
                        FROM transactions WHERE date LIKE ?
                        GROUP BY type,category ORDER BY type, total DESC""", (f'{month}%',)).fetchall()
    # Loan EMI breakdown by sub_category (individual loan types)
    loan_rows = c.execute("""SELECT COALESCE(sub_category,'Unknown') as category,
                                    COALESCE(SUM(amount),0) as total
                             FROM transactions
                             WHERE type='expense' AND category='Loan EMI' AND date LIKE ?
                             GROUP BY sub_category ORDER BY total DESC""", (f'{month}%',)).fetchall()
    conn.close()
    result = {}
    for r in rows: result.setdefault(r['type'], []).append({'category': r['category'], 'total': r['total']})
    result['loan_emi'] = [dict(r) for r in loan_rows]
    return jsonify(result)

@app.route('/api/transactions')
def api_transactions():
    month = request.args.get('month', ''); ttype = request.args.get('type', '')
    conn = get_db(); c = conn.cursor()
    q = "SELECT * FROM transactions WHERE 1=1"; p = []
    if month: q += " AND date LIKE ?"; p.append(f'{month}%')
    if ttype: q += " AND type=?"; p.append(ttype)
    q += " ORDER BY date DESC"
    rows = c.execute(q, p).fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

def _mf_best_match(query, matches):
    """Score MFAPI search results and return the best match for the query.
    Prefers Direct plans over Regular, Growth over IDCW/Dividend,
    and maximises word-overlap with the query."""
    q_words = set(query.lower().split())
    q_lower = query.lower()
    has_direct  = 'direct'  in q_lower
    has_growth  = 'growth'  in q_lower
    has_regular = 'regular' in q_lower

    def score(m):
        name = m['schemeName'].lower()
        words = set(name.split())
        overlap = len(q_words & words)
        if has_direct  and 'regular' in name: overlap -= 3
        if has_growth  and ('idcw' in name or 'dividend' in name): overlap -= 2
        if has_regular and 'direct' in name: overlap -= 3
        if has_direct  and 'direct' in name: overlap += 2
        if has_growth  and 'growth' in name: overlap += 1
        return overlap

    return max(matches, key=score)


def _mf_search(query):
    """Search MFAPI.in with progressive fallback for long fund names.
    Returns list of match dicts, or empty list."""
    words = query.split()
    # Try progressively shorter prefixes until we get results
    for n in range(len(words), 2, -1):
        q = ' '.join(words[:n])
        url = 'https://api.mfapi.in/mf/search?q=' + urllib.parse.quote(q)
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        try:
            with urllib.request.urlopen(req, timeout=8) as r:
                results = _json.loads(r.read())
            if results:
                return results
        except Exception:
            pass
    return []


@app.route('/api/mf_nav')
def get_mf_nav():
    """Search MFAPI.in for a mutual fund's latest NAV by name.
    Returns top 5 matches plus the best match's current NAV."""
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify({'error': 'No query'}), 400
    try:
        # 1. Search for matching schemes (with fallback for long names)
        results = _mf_search(query)
        if not results:
            return jsonify({'error': f'No fund found for: {query}'}), 404
        # 2. Fetch NAV for the best match
        best = _mf_best_match(query, results)
        nav_url = f'https://api.mfapi.in/mf/{best["schemeCode"]}'
        req2 = urllib.request.Request(nav_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req2, timeout=8) as r:
            nav_data = _json.loads(r.read())
        latest = nav_data['data'][0]
        return jsonify({
            'scheme_code': best['schemeCode'],
            'scheme_name': best['schemeName'],
            'nav':  float(latest['nav']),
            'date': latest['date'],
            'matches': [{'code': x['schemeCode'], 'name': x['schemeName']} for x in results[:6]],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/portfolio_units')
def get_portfolio_units():
    """Return net units held per fund/stock from invest_transactions (BUY − SELL)."""
    conn = get_db(); c = conn.cursor()
    rows = c.execute("""
        SELECT stock_name, asset_type,
               SUM(CASE WHEN UPPER(action)='BUY' THEN quantity ELSE -quantity END) AS net_units,
               SUM(CASE WHEN UPPER(action)='BUY' THEN invested_value ELSE -invested_value END) AS net_invested
        FROM invest_transactions
        WHERE stock_name IS NOT NULL AND stock_name != ''
        GROUP BY stock_name, asset_type
        HAVING net_units > 0
        ORDER BY LOWER(asset_type), stock_name
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/gold_price')
def get_gold_price():
    """Fetch live 24K gold price in INR per gram.
    Combines XAU/USD from gold-api.com with USD/INR from frankfurter.app.
    1 troy oz = 31.1035 grams."""
    try:
        # XAU spot price in USD per troy oz
        req1 = urllib.request.Request(
            'https://api.gold-api.com/price/XAU',
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req1, timeout=6) as r:
            xau = _json.loads(r.read())
        usd_per_oz = float(xau['price'])

        # USD → INR exchange rate
        req2 = urllib.request.Request(
            'https://api.frankfurter.app/latest?from=USD&to=INR',
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req2, timeout=6) as r:
            fx = _json.loads(r.read())
        inr_per_usd = float(fx['rates']['INR'])

        inr_per_gram = round((usd_per_oz * inr_per_usd) / 31.1035)
        return jsonify({
            'price_inr_per_gram': inr_per_gram,
            'usd_per_oz': round(usd_per_oz, 2),
            'inr_per_usd': round(inr_per_usd, 2),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/transactions/last_amounts')
def api_last_amounts():
    """Return the most-recently-entered amount per category for a given type.
    Excludes auto-posted deduction entries (sub_category='Choice Pay Deduction').
    Used by the Add Transaction modal to pre-fill fields on first open."""
    ttype = request.args.get('type', '')
    if not ttype:
        return jsonify({})
    conn = get_db(); c = conn.cursor()
    rows = c.execute(
        """SELECT category, amount FROM transactions
           WHERE type=? AND (sub_category IS NULL OR sub_category != 'Choice Pay Deduction')
           ORDER BY date DESC, id DESC""",
        (ttype,)
    ).fetchall()
    conn.close()
    seen = {}
    for r in rows:
        if r['category'] not in seen:
            seen[r['category']] = r['amount']
    return jsonify(seen)

@app.route('/api/transactions', methods=['POST'])
def add_transaction():
    d = request.json; conn = get_db()
    conn.execute("INSERT INTO transactions (type,category,sub_category,amount,date,note) VALUES (?,?,?,?,?,?)",
                 (d['type'], d['category'], d.get('sub_category',''), float(d['amount']), d['date'], d.get('note','')))
    conn.commit()
    nid     = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    sub_cat = (d.get('sub_category') or '').strip()
    # Sync to demat_wallet for demat transfers
    if sub_cat == 'Demat Deposit':
        conn.execute(
            "INSERT INTO demat_wallet (txn_date, type, amount, note, ref_txn_id) VALUES (?,?,?,?,?)",
            (d['date'], 'DEPOSIT', float(d['amount']), d.get('note','Zerodha Pay-in'), nid)
        )
        conn.commit()
    elif sub_cat == 'Demat Withdrawal':
        conn.execute(
            "INSERT INTO demat_wallet (txn_date, type, amount, note, ref_txn_id) VALUES (?,?,?,?,?)",
            (d['date'], 'WITHDRAW', float(d['amount']), d.get('note','Zerodha Pay-out'), nid)
        )
        conn.commit()
    # ── Cross-link: investment → invest_transactions + monthly_investment_calc ──
    if d.get('type') == 'investment' and not d.get('auto_created'):
        stock_name = (d.get('sub_category') or d.get('category') or '').strip()
        asset_type = d.get('asset_type', '')
        qty        = float(d.get('quantity', 1))
        price      = float(d.get('price', float(d['amount'])))
        action     = d.get('action', 'BUY')
        _crosslink_investment(conn, nid, stock_name, asset_type, qty, price, action, d.get('note',''), d['date'])
        conn.commit()

    # ── Cross-link: Loan EMI expense → loans table ─────────────────────────────
    if d.get('type') == 'expense' and d.get('category', '').lower() in ('loan emi', 'loan_emi', 'emi') and not d.get('auto_created'):
        month = d['date'][:7]
        loan_type = d.get('sub_category') or d.get('category')
        conn.execute("INSERT INTO loans (month, loan_type, amount, txn_id) VALUES (?,?,?,?)",
                     (month, loan_type, float(d['amount']), nid))
        conn.commit()

    # ── Cross-link: monthly savings → Emergency Fund asset (current month only) ─
    if d.get('type') in ('income', 'expense') and not d.get('auto_created'):
        _sync_month_savings_to_ef(conn, d['date'][:7])
        conn.commit()

    conn.close()
    return jsonify({'success': True, 'id': nid})

@app.route('/api/transactions/<int:tid>', methods=['DELETE'])
def delete_transaction(tid):
    conn = get_db(); conn.execute("DELETE FROM transactions WHERE id=?", (tid,)); conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/portfolio')
def api_portfolio():
    """All portfolio rows joined to InvestMapping for AssetClass/Category/Type tree."""
    conn = get_db()
    rows = conn.execute("""
        SELECT p.AssetID        AS asset_id,
               im.AssetClass    AS asset_class,
               im.AssetCategory AS asset_category,
               im.AssetType     AS asset_type,
               p.InvestedValue  AS invested_value,
               p.CurrentValue   AS current_value,
               p.ReturnValue    AS return_value,
               p.ReturnPCT      AS return_pct,
               p.Purpose        AS purpose,
               p.UpdateAt       AS updated_at
        FROM portfolio p
        JOIN InvestMapping im ON p.AssetID = im.AssetID
        WHERE (p.InvestedValue > 0 OR p.CurrentValue > 0)
        ORDER BY im.AssetClass, im.AssetCategory, im.AssetType
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/portfolio/asset_types')
def api_portfolio_asset_types():
    """Distinct asset_type values from portfolio joined to InvestMapping."""
    conn = get_db()
    rows = conn.execute("""
        SELECT DISTINCT im.AssetType
        FROM portfolio p
        JOIN InvestMapping im ON p.AssetID = im.AssetID
        WHERE im.AssetType IS NOT NULL ORDER BY im.AssetType
    """).fetchall()
    conn.close()
    return jsonify([r['AssetType'] for r in rows])

@app.route('/api/portfolio/summary')
def api_portfolio_summary():
    """Portfolio summary grouped by AssetClass, plus per-type rows and grand totals."""
    conn = get_db(); c = conn.cursor()

    # Group by AssetClass (tree level 1)
    by_class = c.execute("""
        SELECT im.AssetClass    AS asset_class,
               COALESCE(SUM(p.InvestedValue), 0) AS invested_value,
               COALESCE(SUM(p.CurrentValue),  0) AS current_value,
               COALESCE(SUM(p.ReturnValue),   0) AS return_value
        FROM portfolio p
        JOIN InvestMapping im ON p.AssetID = im.AssetID
        GROUP BY im.AssetClass ORDER BY current_value DESC
    """).fetchall()
    by_class = [dict(r) for r in by_class]
    for r in by_class:
        r['return_pct'] = round(r['return_value'] / r['invested_value'] * 100, 2) if r['invested_value'] > 0 else 0

    # Individual rows (AssetType level) for drilldown / charts
    by_type = c.execute("""
        SELECT p.AssetID        AS asset_id,
               im.AssetClass    AS asset_class,
               im.AssetCategory AS asset_category,
               im.AssetType     AS asset_type,
               p.InvestedValue  AS invested_value,
               p.CurrentValue   AS current_value,
               p.ReturnValue    AS return_value,
               p.ReturnPCT      AS return_pct,
               p.Purpose        AS purpose
        FROM portfolio p
        JOIN InvestMapping im ON p.AssetID = im.AssetID
        ORDER BY im.AssetClass, im.AssetCategory, im.AssetType
    """).fetchall()
    by_type = [dict(r) for r in by_type]

    total_inv = sum(r['invested_value'] for r in by_class)
    total_val = sum(r['current_value']  for r in by_class)
    conn.close()
    return jsonify({
        'by_class':     by_class,
        'by_type':      by_type,
        'total_invested': total_inv,
        'total_value':    total_val,
        'total_current':  total_val,   # alias for backward compat
        'total_return': r2(total_val - total_inv),
        'return_pct': round((total_val - total_inv) / total_inv * 100, 2) if total_inv else 0,
    })

@app.route('/api/portfolio/<asset_id>', methods=['PATCH'])
def patch_portfolio(asset_id):
    """Update CurrentValue for a portfolio row identified by InvestMapping.AssetID."""
    d = request.json; conn = get_db()
    cv = float(d.get('current_value', 0))
    row = conn.execute("SELECT InvestedValue FROM portfolio WHERE AssetID=?", (asset_id,)).fetchone()
    iv  = float(row['InvestedValue']) if row else 0
    ret = r2(cv - iv)
    ret_pct = r2(ret / iv * 100) if iv > 0 else 0
    conn.execute("""UPDATE portfolio SET CurrentValue=?, ReturnValue=?, ReturnPCT=?, UpdateAt=datetime('now')
                    WHERE AssetID=?""", (cv, ret, ret_pct, asset_id))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/portfolio/update', methods=['POST'])
def update_portfolio():
    """Update CurrentValue by asset_id (InvestMapping.AssetID)."""
    d = request.json; conn = get_db()
    asset_id = (d.get('asset_id') or d.get('asset') or '').strip()
    cv = float(d['current_value'])
    row = conn.execute("SELECT InvestedValue FROM portfolio WHERE AssetID=?", (asset_id,)).fetchone()
    iv  = float(row['InvestedValue']) if row else 0
    ret = r2(cv - iv)
    ret_pct = r2(ret / iv * 100) if iv > 0 else 0
    conn.execute("""UPDATE portfolio SET CurrentValue=?, ReturnValue=?, ReturnPCT=?, UpdateAt=datetime('now')
                    WHERE AssetID=?""", (cv, ret, ret_pct, asset_id))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/invest_transactions')
def api_invest_tx():
    asset_type = request.args.get('type',  '')
    stock      = request.args.get('stock', '')
    month      = request.args.get('month', '')
    conn = get_db(); c = conn.cursor()
    q = "SELECT * FROM invest_transactions WHERE 1=1"; p = []
    if asset_type: q += " AND asset_type=?"; p.append(asset_type)
    if stock:      q += " AND stock_name=?"; p.append(stock)
    if month:      q += " AND (month LIKE ? OR entry_date LIKE ?)"; p += [f'{month}%', f'{month}%']
    q += " ORDER BY entry_date DESC LIMIT 500"
    rows = c.execute(q, p).fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/invest_transactions/summary')
def api_invest_tx_summary():
    """Aggregate investment transactions by asset_type.
    Returns total invested per type, monthly breakdown, and distinct months list."""
    month = request.args.get('month', '')
    conn = get_db(); c = conn.cursor()
    where = "WHERE action IN ('BUY','Buy','buy')"
    params = []
    if month:
        where += " AND (month LIKE ? OR entry_date LIKE ?)"
        params += [f'{month}%', f'{month}%']
    # Total distinct active months across all transactions (for avg/month on cards)
    total_months_row = c.execute("""
        SELECT COUNT(DISTINCT COALESCE(month, substr(entry_date,1,7))) AS n
        FROM invest_transactions
        WHERE action IN ('BUY','Buy','buy')
          AND COALESCE(month, substr(entry_date,1,7)) IS NOT NULL
          AND COALESCE(month, substr(entry_date,1,7)) != ''
    """).fetchone()
    total_active_months = int(total_months_row['n'] or 1)

    # Per asset type totals
    by_type = c.execute(f"""
        SELECT asset_type,
               COUNT(*) AS tx_count,
               SUM(invested_value) AS total_invested,
               SUM(CASE WHEN action IN ('BUY','Buy','buy') THEN invested_value ELSE 0 END) AS inflow,
               SUM(CASE WHEN action IN ('SELL','Sell','sell') THEN invested_value ELSE 0 END) AS outflow,
               COUNT(DISTINCT COALESCE(month, substr(entry_date,1,7))) AS active_months
        FROM invest_transactions {where}
        GROUP BY asset_type ORDER BY total_invested DESC
    """, params).fetchall()
    # Monthly breakdown per type
    monthly = c.execute(f"""
        SELECT COALESCE(month, substr(entry_date,1,7)) AS mo,
               asset_type,
               SUM(invested_value) AS invested
        FROM invest_transactions {where}
        GROUP BY mo, asset_type ORDER BY mo
    """, params).fetchall()
    # All distinct months
    months = c.execute("""
        SELECT DISTINCT COALESCE(month, substr(entry_date,1,7)) AS mo
        FROM invest_transactions WHERE mo IS NOT NULL AND mo != ''
        ORDER BY mo DESC
    """).fetchall()
    conn.close()
    return jsonify({
        'by_type':         [dict(r) for r in by_type],
        'monthly':         [dict(r) for r in monthly],
        'months':          [r['mo'] for r in months if r['mo']],
        'total_months':    total_active_months,
    })

@app.route('/api/loans/summary')
def api_loans_summary():
    conn = get_db(); c = conn.cursor()
    latest_month = c.execute("SELECT MAX(month) FROM loans").fetchone()[0] or ''
    rows = c.execute("""
        SELECT l.loan_type,
               SUM(l.amount)                     AS total_paid,
               COALESCE(lc.amount, 0)            AS current_emi
        FROM loans l
        LEFT JOIN loans lc
               ON lc.loan_type = l.loan_type AND lc.month = ?
        GROUP BY l.loan_type
        ORDER BY COALESCE(lc.amount, 0) DESC, SUM(l.amount) DESC
    """, (latest_month,)).fetchall()
    conn.close()
    result = [dict(r) for r in rows]
    for r in result:
        r['is_active'] = r['current_emi'] > 0
    return jsonify(result)

@app.route('/api/expense_categories')
def api_expense_categories():
    """Return distinct expense categories from transactions table."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT category FROM transactions WHERE type='expense' AND category IS NOT NULL AND category!='' ORDER BY category"
    ).fetchall()
    conn.close()
    return jsonify([r['category'] for r in rows])

@app.route('/api/loans/emi_month')
def api_loans_emi_month():
    """Return expected vs paid EMI totals for a given month."""
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    conn = get_db(); c = conn.cursor()
    # Expected: sum expected_emi of active loans (calculated same as api_loan_master_list)
    loans = [dict(r) for r in c.execute(
        "SELECT * FROM loan_master WHERE status='active'"
    ).fetchall()]
    today_dt = datetime.now()
    expected_total = 0.0
    next_to_close  = None
    min_balance    = float('inf')
    for loan in loans:
        total_paid = c.execute(
            "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='expense' AND category='Loan EMI' AND sub_category=?",
            (loan['loan_name'],)
        ).fetchone()[0]
        balance = max(0.0, loan['total_repayment'] - total_paid)
        try:
            target_dt = datetime.strptime(loan['target_close_date'], '%Y-%m-%d')
        except Exception:
            target_dt = today_dt
        months_remaining = max(1, (target_dt.year - today_dt.year) * 12 + (target_dt.month - today_dt.month))
        expected_emi = round(balance / months_remaining, 0) if balance > 0 else 0.0
        expected_total += expected_emi
        if balance > 0 and balance < min_balance:
            min_balance   = balance
            next_to_close = {'name': loan['loan_name'], 'balance': round(balance, 2), 'expected_emi': expected_emi}
    # Paid: sum of Loan EMI expense transactions this month
    paid_total = c.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='expense' AND category='Loan EMI' AND date LIKE ?",
        (f'{month}%',)
    ).fetchone()[0]
    conn.close()
    deficit = expected_total - paid_total
    deficit_pct = round(deficit / expected_total * 100, 1) if expected_total > 0 else 0.0
    return jsonify({
        'expected_total': round(expected_total, 2),
        'paid_total':     round(paid_total, 2),
        'deficit':        round(deficit, 2),
        'deficit_pct':    deficit_pct,
        'next_to_close':  next_to_close,
        'month':          month,
    })

@app.route('/api/loan_master')
def api_loan_master_list():
    conn = get_db(); c = conn.cursor()
    loans = [dict(r) for r in c.execute(
        "SELECT * FROM loan_master ORDER BY loan_type, start_date"
    ).fetchall()]
    today_dt = datetime.now()
    updated = False
    for loan in loans:
        total_paid = c.execute("""
            SELECT COALESCE(SUM(amount), 0) FROM transactions
            WHERE type='expense' AND category='Loan EMI' AND sub_category=?
        """, (loan['loan_name'],)).fetchone()[0]
        loan['total_paid'] = round(total_paid, 2)
        loan['balance'] = round(max(0.0, loan['total_repayment'] - total_paid), 2)
        loan['repayment_pct'] = round(total_paid / loan['total_repayment'] * 100, 2) if loan['total_repayment'] > 0 else 0.0
        loan['remaining_pct'] = round(100.0 - loan['repayment_pct'], 2)
        try:
            target_dt = datetime.strptime(loan['target_close_date'], '%Y-%m-%d')
        except Exception:
            target_dt = today_dt
        months_remaining = max(1, (target_dt.year - today_dt.year) * 12 + (target_dt.month - today_dt.month))
        loan['months_remaining'] = months_remaining
        loan['expected_emi'] = round(loan['balance'] / months_remaining, 0) if loan['balance'] > 0 else 0.0
        if loan['balance'] <= 0 and loan['status'] == 'active':
            c.execute("UPDATE loan_master SET status='closed' WHERE id=?", (loan['id'],))
            loan['status'] = 'closed'
            updated = True
    if updated:
        conn.commit()
    conn.close()
    return jsonify(loans)

@app.route('/api/loan_master', methods=['POST'])
def api_loan_master_add():
    d = request.json
    if not d or not d.get('loan_name'):
        return jsonify({'error': 'loan_name required'}), 400
    conn = get_db()
    conn.execute("""INSERT INTO loan_master
        (loan_name, loan_type, loan_amount, total_repayment, start_date, target_close_date, status)
        VALUES (?,?,?,?,?,?,?)""",
        (d['loan_name'].strip(), d['loan_type'], float(d.get('loan_amount', 0)),
         float(d.get('total_repayment', 0)), d['start_date'], d['target_close_date'], 'active'))
    conn.commit()
    nid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({'success': True, 'id': nid})

@app.route('/api/loan_master/<int:lid>', methods=['DELETE'])
def api_loan_master_delete(lid):
    conn = get_db()
    conn.execute("DELETE FROM loan_master WHERE id=?", (lid,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/loan_master/<int:lid>/close', methods=['POST'])
def api_loan_master_close(lid):
    conn = get_db()
    conn.execute("UPDATE loan_master SET status='closed' WHERE id=?", (lid,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

# ── Universe Magnet: Status entries (DB-backed, date-tracked for trend charts) ──

@app.route('/api/magnet_status/<magnet>', methods=['GET'])
def get_magnet_status(magnet):
    conn = get_db(); c = conn.cursor()
    # Latest entry per metric_name — use MAX(id) as tiebreaker so same-date dupes collapse
    latest = c.execute("""
        SELECT ms.* FROM magnet_status ms
        INNER JOIN (
            SELECT metric_name, MAX(id) as max_id
            FROM magnet_status WHERE magnet=? GROUP BY metric_name
        ) mx ON ms.id=mx.max_id
        WHERE ms.magnet=? ORDER BY ms.metric_name
    """, (magnet, magnet)).fetchall()
    # History for all metrics (for future trend charts)
    history = c.execute("""
        SELECT * FROM magnet_status WHERE magnet=?
        ORDER BY metric_name, recorded_date
    """, (magnet,)).fetchall()
    conn.close()
    return jsonify({
        'latest':  [dict(r) for r in latest],
        'history': [dict(r) for r in history],
    })

@app.route('/api/magnet_status', methods=['POST'])
def save_magnet_status():
    d = request.json
    magnet      = d.get('magnet','').strip()
    metric_name = d.get('metric_name','').strip()
    if not magnet or not metric_name:
        return jsonify({'error': 'magnet and metric_name required'}), 400
    conn = get_db()
    conn.execute("""
        INSERT INTO magnet_status
            (magnet, metric_name, emoji, current_value, target_value, note, recorded_date,
             vision_id, vision_type, feel, cost, trip_name, start_date, end_date)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        magnet, metric_name,
        d.get('emoji','📌'),
        d.get('current_value',''),
        d.get('target_value',''),
        d.get('note',''),
        d.get('recorded_date', datetime.now().strftime('%Y-%m-%d')),
        d.get('vision_id'),
        d.get('vision_type','general'),
        d.get('feel'),
        d.get('cost'),
        d.get('trip_name'),
        d.get('start_date'),
        d.get('end_date'),
    ))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    return jsonify({'success': True, 'id': new_id})

@app.route('/api/magnet_status/<int:sid>', methods=['DELETE'])
def delete_magnet_status(sid):
    conn = get_db()
    conn.execute("DELETE FROM magnet_status WHERE id=?", (sid,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/record_daily_scores', methods=['POST'])
@login_required
def record_daily_scores():
    """Upsert one magnet_status row per metric per day (called automatically on page load)."""
    entries = request.get_json(force=True) or []
    today   = datetime.now().strftime('%Y-%m-%d')
    conn    = get_db()
    for e in entries:
        magnet      = (e.get('magnet') or '').strip()
        metric_name = (e.get('metric_name') or '').strip()
        if not magnet or not metric_name:
            continue
        existing = conn.execute(
            "SELECT id FROM magnet_status WHERE magnet=? AND metric_name=? AND recorded_date=?",
            (magnet, metric_name, today)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE magnet_status SET current_value=?, target_value=?, emoji=?, note=? WHERE id=?",
                (str(e.get('current_value', '')), str(e.get('target_value', '')),
                 e.get('emoji', '📊'), e.get('note', ''), existing['id'])
            )
        else:
            conn.execute(
                """INSERT INTO magnet_status
                   (magnet, metric_name, emoji, current_value, target_value, note, recorded_date)
                   VALUES (?,?,?,?,?,?,?)""",
                (magnet, metric_name,
                 e.get('emoji', '📊'),
                 str(e.get('current_value', '')),
                 str(e.get('target_value', '')),
                 e.get('note', ''),
                 today)
            )
    conn.commit(); conn.close()
    return jsonify({'success': True, 'date': today, 'count': len(entries)})

@app.route('/api/magnet_status/update_type', methods=['POST'])
@login_required
def update_magnet_status_type():
    """Update vision_type on all status rows linked to a vision card (called when user changes type in edit)."""
    d = request.get_json(force=True) or {}
    vision_id   = d.get('vision_id', '').strip()
    metric_name = d.get('metric_name', '').strip()
    magnet      = d.get('magnet', '').strip()
    vision_type = d.get('vision_type', 'habit').strip()
    if not (vision_id or metric_name) or not magnet:
        return jsonify({'error': 'vision_id or metric_name + magnet required'}), 400
    conn = get_db()
    if vision_id:
        conn.execute(
            "UPDATE magnet_status SET vision_type=? WHERE magnet=? AND (vision_id=? OR metric_name=?)",
            (vision_type, magnet, vision_id, metric_name)
        )
    else:
        conn.execute(
            "UPDATE magnet_status SET vision_type=? WHERE magnet=? AND metric_name=?",
            (vision_type, magnet, metric_name)
        )
    conn.commit(); conn.close()
    return jsonify({'success': True})

# ── Universe Magnet: Vision cards (DB-backed) ────────────────────────────────

@app.route('/api/um_vision/<magnet>', methods=['GET'])
def get_um_vision(magnet):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM um_vision_cards WHERE magnet=? ORDER BY created_at", (magnet,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/um_vision', methods=['POST'])
def save_um_vision():
    d = request.json
    vid    = d.get('id')
    magnet = d.get('magnet','').strip()
    title  = d.get('title','').strip()
    if not magnet or not title:
        return jsonify({'error': 'magnet and title required'}), 400
    conn = get_db()
    if vid:
        conn.execute("""
            UPDATE um_vision_cards SET title=?, description=?, photo_data=?,
            vision_type=COALESCE(?,vision_type), updated_at=datetime('now')
            WHERE id=?
        """, (title, d.get('description',''), d.get('photo_data',''),
              d.get('vision_type'), vid))
        conn.commit(); conn.close()
        return jsonify({'success': True, 'id': vid})
    else:
        import uuid
        new_id = str(uuid.uuid4())
        conn.execute("""
            INSERT INTO um_vision_cards (id, magnet, title, description, photo_data, vision_type)
            VALUES (?,?,?,?,?,?)
        """, (new_id, magnet, title, d.get('description',''), d.get('photo_data',''),
              d.get('vision_type','general')))
        conn.commit(); conn.close()
        return jsonify({'success': True, 'id': new_id})

@app.route('/api/um_vision/<vid>', methods=['DELETE'])
def delete_um_vision(vid):
    conn = get_db()
    # Cascade-delete linked status entries for this vision
    conn.execute("DELETE FROM magnet_status WHERE vision_id=?", (vid,))
    conn.execute("DELETE FROM um_vision_cards WHERE id=?", (vid,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/alerts')
def api_alerts():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    conn = get_db(); c = conn.cursor()
    alerts = [dict(a) for a in c.execute("SELECT * FROM alerts WHERE is_active=1").fetchall()]
    for a in alerts:
        actual = 0
        if a['condition'] == 'category_exceeds':
            actual = c.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type=? AND category=? AND date LIKE ?",
                               (a['type'], a['category'], f'{month}%')).fetchone()[0]
            a['triggered'] = actual > a['threshold']
        elif a['condition'] == 'savings_below':
            inc = c.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='income' AND date LIKE ?", (f'{month}%',)).fetchone()[0]
            exp = c.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='expense' AND date LIKE ?", (f'{month}%',)).fetchone()[0]
            inv = c.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='investment' AND date LIKE ?", (f'{month}%',)).fetchone()[0]
            actual = inc - exp - inv; a['triggered'] = actual < a['threshold']
        elif a['condition'] == 'total_below':
            actual = c.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type=? AND date LIKE ?",
                               (a['type'], f'{month}%')).fetchone()[0]
            a['triggered'] = actual < a['threshold']
        a['actual'] = actual
    conn.close()
    return jsonify(alerts)

@app.route('/api/alerts', methods=['POST'])
def add_alert():
    d = request.json; conn = get_db()
    conn.execute("INSERT INTO alerts (name,type,condition,threshold,category,period) VALUES (?,?,?,?,?,?)",
                 (d['name'], d['type'], d['condition'], float(d['threshold']), d.get('category'), d.get('period','monthly')))
    conn.commit(); aid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]; conn.close()
    return jsonify({'success': True, 'id': aid})

@app.route('/api/alerts/<int:aid>', methods=['PUT'])
def update_alert(aid):
    d = request.json; conn = get_db()
    fields, vals = [], []
    if 'threshold' in d: fields.append('threshold=?'); vals.append(float(d['threshold']))
    if 'name' in d:      fields.append('name=?');      vals.append(d['name'])
    if fields:
        conn.execute(f"UPDATE alerts SET {','.join(fields)} WHERE id=?", vals + [aid])
        conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/alerts/<int:aid>', methods=['DELETE'])
def delete_alert(aid):
    conn = get_db(); conn.execute("DELETE FROM alerts WHERE id=?", (aid,)); conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/alloc_targets', methods=['GET'])
@login_required
def api_alloc_targets_get():
    """Return investment allocation targets stored in app_settings."""
    import json as _json
    conn = get_db()
    row  = conn.execute("SELECT value FROM app_settings WHERE key='alloc_targets'").fetchone()
    conn.close()
    targets = _json.loads(row['value']) if row else {}
    return jsonify(targets)

@app.route('/api/alloc_targets', methods=['POST'])
@login_required
def api_alloc_targets_save():
    """Persist allocation targets to app_settings as JSON."""
    import json as _json
    data = request.get_json(force=True) or {}
    # Validate: values must be numbers 0-100, sum ≤ 100
    total = sum(float(v) for v in data.values() if isinstance(v, (int, float)))
    if total > 100.5:
        return jsonify({'ok': False, 'error': f'Sum {total:.1f}% exceeds 100%'}), 400
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('alloc_targets', ?)",
                 (_json.dumps(data),))
    conn.commit(); conn.close()
    return jsonify({'ok': True})

@app.route('/api/alerts/alloc_drift')
@login_required
def api_alerts_alloc_drift():
    """Return allocation drift alerts: asset classes where actual % deviates >5% from target."""
    import json as _json
    conn = get_db(); c = conn.cursor()
    try:
        # Load targets from DB
        row = c.execute("SELECT value FROM app_settings WHERE key='alloc_targets'").fetchone()
        if not row:
            conn.close()
            return jsonify([])
        targets = _json.loads(row['value'])
        if not any(v > 0 for v in targets.values()):
            conn.close()
            return jsonify([])

        # Compute actual allocation from assets table
        # AssetClass lives in InvestMapping, linked via AssetMapping.AssetId
        by_class = {}
        rows = c.execute("""
            SELECT COALESCE(im.AssetClass, 'Other') cls,
                   COALESCE(SUM(a.currentvalue), 0) cur
            FROM assets a
            JOIN AssetMapping am  ON am.MappingID = a.MappingID
            JOIN InvestMapping im ON im.AssetID   = am.AssetId
            GROUP BY cls
        """).fetchall()
        total = sum(r['cur'] for r in rows)
        if total <= 0:
            conn.close()
            return jsonify([])
        for r in rows:
            by_class[r['cls']] = r['cur']

        CLASS_ICONS = {
            'Growth':'📈','Commodities':'🪙','Stability':'🛡️',
            'Liquidity':'💧','Retirement':'👴','Real State':'🏠'
        }
        alerts = []
        for cls, tgt in targets.items():
            tgt = float(tgt or 0)
            if tgt == 0:
                continue
            # Match case-insensitively
            actual_val = next(
                (v for k, v in by_class.items() if k.lower() == cls.lower()), 0
            )
            actual_pct = (actual_val / total) * 100
            drift = actual_pct - tgt
            if abs(drift) > 5:
                over = drift > 0
                alerts.append({
                    'cls':        cls,
                    'icon':       CLASS_ICONS.get(cls, '📊'),
                    'target_pct': round(tgt, 1),
                    'actual_pct': round(actual_pct, 2),
                    'drift':      round(drift, 1),
                    'over':       over,
                    'msg': (
                        f"Over-allocated by {abs(drift):.1f}% — consider rebalancing"
                        if over else
                        f"Under-allocated by {abs(drift):.1f}% — consider adding to {cls}"
                    )
                })
        conn.close()
        return jsonify(alerts)
    except Exception as e:
        conn.close()
        return jsonify([])


@app.route('/api/alerts/seed_defaults', methods=['POST'])
@login_required
def api_alerts_seed_defaults():
    """Seed a set of sensible default alert rules when none exist."""
    conn = get_db(); c = conn.cursor()
    existing = c.execute("SELECT COUNT(*) FROM alerts").fetchone()[0]
    if existing > 0:
        conn.close()
        return jsonify({'ok': True, 'msg': 'Alerts already exist', 'seeded': 0})
    # Estimate income from recent transactions
    income = c.execute(
        "SELECT COALESCE(AVG(monthly),0) FROM ("
        "  SELECT strftime('%Y-%m',date) m, SUM(amount) monthly"
        "  FROM transactions WHERE type='income' GROUP BY m ORDER BY m DESC LIMIT 3"
        ")"
    ).fetchone()[0] or 50000
    defaults = [
        ('💸 Monthly Expense Limit',    'expense', 'expense_limit',     round(income * 0.60), None,       'monthly'),
        ('🍽️ Food & Dining Limit',      'expense', 'category_exceeds',  round(income * 0.15), 'Food',     'monthly'),
        ('🛒 Shopping Limit',           'expense', 'category_exceeds',  round(income * 0.10), 'Shopping', 'monthly'),
        ('💰 Min Monthly Savings',      'income',  'savings_below',     round(income * 0.20), None,       'monthly'),
        ('🆘 Emergency Fund Minimum',   'expense', 'emergency_fund_below', round(income * 6),  None,       'monthly'),
    ]
    seeded = 0
    for name, typ, cond, thresh, cat, period in defaults:
        c.execute(
            "INSERT INTO alerts (name,type,condition,threshold,category,period,is_active) VALUES (?,?,?,?,?,?,1)",
            (name, typ, cond, thresh, cat, period)
        )
        seeded += 1
    conn.commit(); conn.close()
    return jsonify({'ok': True, 'seeded': seeded})


@app.route('/api/alerts/smart')
@login_required
def smart_alerts():
    """Investment-aware smart alerts for login popup."""
    try:
        conn = get_db(); c = conn.cursor()
        month = datetime.now().strftime('%Y-%m')
        triggered = []

        # ── User-defined rule alerts ─────────────────────────────────────────
        rules = [dict(r) for r in c.execute("SELECT * FROM alerts WHERE is_active=1").fetchall()]
        for a in rules:
            try:
                cond, actual = a['condition'], 0
                if cond == 'loan_emi_exceeds':
                    actual = c.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='loan' AND date LIKE ?", (f'{month}%',)).fetchone()[0]
                    if actual > a['threshold']:
                        triggered.append({'id':a['id'],'name':a['name'],'icon':'🏦','type':'loan_emi','actual':actual,'threshold':a['threshold'],'msg':f"Loan EMI ₹{actual:,.0f} exceeded limit ₹{a['threshold']:,.0f}"})
                elif cond == 'expense_limit':
                    actual = c.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='expense' AND date LIKE ?", (f'{month}%',)).fetchone()[0]
                    if actual > a['threshold']:
                        triggered.append({'id':a['id'],'name':a['name'],'icon':'💸','type':'expense_limit','actual':actual,'threshold':a['threshold'],'msg':f"Monthly expenses ₹{actual:,.0f} exceeded limit ₹{a['threshold']:,.0f}"})
                elif cond == 'emergency_fund_below':
                    actual = c.execute("SELECT COALESCE(SUM(currentvalue),0) FROM assets WHERE LOWER(assetname) LIKE '%emergency%' OR LOWER(purpose) LIKE '%emergency%'").fetchone()[0]
                    if actual < a['threshold']:
                        triggered.append({'id':a['id'],'name':a['name'],'icon':'🆘','type':'emergency_fund','actual':actual,'threshold':a['threshold'],'msg':f"Emergency fund ₹{actual:,.0f} below target ₹{a['threshold']:,.0f}"})
                elif cond == 'category_exceeds':
                    actual = c.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type=? AND category=? AND date LIKE ?",
                                       (a['type'], a['category'], f'{month}%')).fetchone()[0]
                    if actual > a['threshold']:
                        triggered.append({'id':a['id'],'name':a['name'],'icon':'⚠️','type':'budget','actual':actual,'threshold':a['threshold'],'msg':f"{a['category']} spend ₹{actual:,.0f} exceeded ₹{a['threshold']:,.0f}"})
            except Exception:
                continue

        # ── Investment smart alerts (assets + AssetMapping + nse_master) ─────
        try:
            stocks = c.execute("""
                SELECT m.AssetName name, m.AssetSymbol symbol,
                       COALESCE(n.ltp, a.ltp, 0)       ltp,
                       COALESCE(a.investedvalue, 0)     inv,
                       COALESCE(a.currentvalue,  0)     cur,
                       COALESCE(n.high_52w, 0)          h52,
                       COALESCE(n.low_52w,  0)          l52
                FROM assets a
                JOIN AssetMapping m ON m.MappingID = a.MappingID
                LEFT JOIN nse_master n ON n.symbol = m.AssetSymbol
                                       OR n.symbol = (m.AssetSymbol || '.NS')
                WHERE m.AssetSymbol IS NOT NULL AND m.AssetSymbol != ''
            """).fetchall()
        except Exception:
            stocks = []

        for s in stocks:
            try:
                ltp, h52, l52 = float(s['ltp']), float(s['h52']), float(s['l52'])
                inv, cur      = float(s['inv']),  float(s['cur'])
                name          = s['name'] or s['symbol']
                if ltp <= 0: continue
                if h52 > 0 and (h52 - ltp) / h52 <= 0.02:
                    triggered.append({'icon':'📈','type':'near_52w_high','name':f"{name} near 52W High",'actual':ltp,'threshold':h52,'msg':f"LTP ₹{ltp:,.1f} is within 2% of 52W High ₹{h52:,.1f} — consider booking profits"})
                if l52 > 0 and ltp > 0 and (ltp - l52) / max(l52, 0.01) <= 0.02:
                    triggered.append({'icon':'📉','type':'near_52w_low','name':f"{name} near 52W Low",'actual':ltp,'threshold':l52,'msg':f"LTP ₹{ltp:,.1f} is near 52W Low ₹{l52:,.1f} — review stop-loss"})
                if inv > 0 and (cur - inv) / inv >= 1.0:
                    pct = round((cur - inv) / inv * 100)
                    triggered.append({'icon':'🚀','type':'profit_100','name':f"{name} — {pct}% Profit",'actual':cur,'threshold':inv,'msg':f"Investment up {pct}% (₹{inv:,.0f} → ₹{cur:,.0f}) — consider booking partial profits"})
            except Exception:
                continue

        conn.close()
        return jsonify(triggered)
    except Exception:
        return jsonify([])

@app.route('/api/alerts/vision_goals')
@login_required
def vision_goal_alerts():
    """Global alerts for Goal-type visions — compares last status value to target."""
    try:
        conn = get_db(); c = conn.cursor()
        magnets = ['health', 'relationship', 'career', 'vision']
        alerts = []
        for magnet in magnets:
            try:
                visions = c.execute(
                    "SELECT * FROM um_vision_cards WHERE magnet=? AND vision_type='goal'", (magnet,)
                ).fetchall()
            except Exception:
                continue
            for v in visions:
                try:
                    vid   = str(v['id'])
                    title = str(v['title'])
                    last  = c.execute("""
                        SELECT current_value, target_value FROM magnet_status
                        WHERE (vision_id=? OR metric_name=?) AND magnet=?
                          AND current_value IS NOT NULL AND current_value != ''
                        ORDER BY id DESC LIMIT 1
                    """, (vid, title, magnet)).fetchone()
                    if not last:
                        continue
                    current = float(last['current_value'])
                    raw_target = last['target_value']
                    if not raw_target:
                        continue
                    target = float(raw_target)
                    if target == 0:
                        continue
                    pct_to_goal  = abs(current - target) / abs(target) * 100
                    pct_progress = current / target * 100
                    if pct_to_goal <= 5:
                        alerts.append({
                            'vision_id': vid, 'title': title, 'magnet': magnet,
                            'current': current, 'target': target,
                            'pct_to_goal': round(pct_to_goal, 1),
                            'severity': 'green', 'icon': '🎯',
                            'msg': f"Almost there! {pct_progress:.1f}% of goal reached — only {pct_to_goal:.1f}% away"
                        })
                    elif pct_to_goal > 25:
                        alerts.append({
                            'vision_id': vid, 'title': title, 'magnet': magnet,
                            'current': current, 'target': target,
                            'pct_to_goal': round(pct_to_goal, 1),
                            'severity': 'red', 'icon': '⚠️',
                            'msg': f"Off track — {pct_to_goal:.1f}% away from goal (current: {current:g}, target: {target:g})"
                        })
                except Exception:
                    continue
        conn.close()
        return jsonify(alerts)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── EXCEL IMPORT HELPERS ──────────────────────────────────────────────────────
def _xl_safe(v):
    try: return float(v) if pd.notna(v) else 0.0
    except: return 0.0

def _xl_income(conn, xlsx):
    df = pd.read_excel(xlsx, sheet_name='Income', header=2)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df['Month'].apply(lambda x: hasattr(x, 'year'))].copy()
    df = df[df['Salary Credit'].notna()].copy()
    rows = []
    cats = [
        ('Salary Credit','Salary','Salary Credit'),
        ('Incentive Bonus','Salary','Incentive / Bonus'),
        ('Rati','Income','Rati'),
        ('Family','Income','Family'),
        ('Investment Return','Investment Return','Returns'),
        ('Loan Return','Income','Loan Return'),
    ]
    for _, row in df.iterrows():
        dt = str(row['Month'])[:10]
        for col, cat, sub in cats:
            val = row.get(col, 0)
            if pd.notna(val) and float(val) > 0:
                rows.append(('income', cat, sub, float(val), dt, 'Excel Upload'))
    conn.executemany("INSERT INTO transactions (type,category,sub_category,amount,date,note) VALUES (?,?,?,?,?,?)", rows)
    conn.commit()
    return len(rows)

def _xl_expenses(conn, xlsx):
    df = pd.read_excel(xlsx, sheet_name='Expenses', header=3)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df['Month'].apply(lambda x: hasattr(x, 'year'))].copy()
    df = df[df['Actual  Expense'].notna() & (df['Actual  Expense'] > 0)].copy()
    cats = [
        ('Kid','Family',"Kids' Expenses"),
        ('Payments','Bills','Subscriptions & Payments'),
        ('HouseHold','Household','Household Expenses'),
        ('Transport','Transport','Transport'),
        ('Family Care','Family','Family Care'),
        ('Shopping','Shopping','Shopping'),
        ('Luxary','Lifestyle','Luxury & Entertainment'),
        ('Support','Family','Family Support'),
    ]
    rows = []
    for _, row in df.iterrows():
        dt = str(row['Month'])[:10]
        for col, cat, sub in cats:
            val = row.get(col, 0)
            if pd.notna(val) and float(val) > 0:
                rows.append(('expense', cat, sub, float(val), dt, 'Excel Upload'))
    conn.executemany("INSERT INTO transactions (type,category,sub_category,amount,date,note) VALUES (?,?,?,?,?,?)", rows)
    conn.commit()
    return len(rows)

def _xl_loans(conn, xlsx):
    df = pd.read_excel(xlsx, sheet_name='Loan', header=3)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df['Month'].apply(lambda x: hasattr(x, 'year'))].copy()
    df = df[df['Total Loan'].notna() & (df['Total Loan'] > 0)].copy()
    loan_types = ['Home Loan','Personal Loan','Car Loan','NBFC Loan (LIC)',
                  'EPF Loan','Education Loan','Gadaget EMI','Credit Card EMI',
                  'Family loan','Credit Purchase','Secret Partner']
    rows, tx_rows = [], []
    for _, row in df.iterrows():
        dt = str(row['Month'])[:10]
        for lt in loan_types:
            val = row.get(lt, 0)
            if pd.notna(val) and float(val) > 0:
                rows.append((dt[:7], lt, float(val)))
                tx_rows.append(('expense','Loan EMI',lt,float(val),dt,'Excel Upload'))
    conn.executemany("INSERT INTO loans (month,loan_type,amount) VALUES (?,?,?)", rows)
    conn.executemany("INSERT INTO transactions (type,category,sub_category,amount,date,note) VALUES (?,?,?,?,?,?)", tx_rows)
    conn.commit()
    return len(rows)

def _xl_investments(conn, xlsx):
    df = pd.read_excel(xlsx, sheet_name='Investment', header=5)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df['Entry Month'].apply(lambda x: hasattr(x, 'year'))].copy()
    df = df[df['Total Investment'].notna()].copy()
    cats = [
        ('Stocks','Stocks','Direct Equity'),
        ('Mutual Fund','Mutual Fund','MF SIP'),
        ('Gold MF','Gold','Gold Mutual Fund'),
        ('Gold ETF','Gold','Gold ETF'),
        ('Gold','Gold','Physical Gold'),
        ('SGB','Gold','Sovereign Gold Bond'),
        ('EPF','Fixed Return','EPF Contribution'),
        ('NPS','Retirement','NPS Contribution'),
        ('Sukanya Samriddhi','Fixed Return','Sukanya Samriddhi'),
        ('Land','Real Estate','Land'),
        ('Flat','Real Estate','Flat'),
    ]
    rows = []
    for _, row in df.iterrows():
        dt = str(row['Entry Month'])[:10]
        for col, cat, sub in cats:
            val = row.get(col, 0)
            if pd.notna(val) and float(val) > 0:
                rows.append(('investment', cat, sub, float(val), dt, 'Excel Upload'))
    conn.executemany("INSERT INTO transactions (type,category,sub_category,amount,date,note) VALUES (?,?,?,?,?,?)", rows)
    conn.commit()
    return len(rows)

def _xl_portfolio(conn, xlsx):
    """Portfolio is now derived from the assets table via InvestMapping JOINs.
    Excel Corpus sheet import is skipped — add assets via the Add Asset modal or CSV upload."""
    return 0

def _xl_invest_tx(conn, xlsx):
    df = pd.read_excel(xlsx, sheet_name='Investment Transactions', header=1)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df['Entry Date'].apply(lambda x: hasattr(x, 'year'))].copy()
    rows = []
    for _, row in df.iterrows():
        # Normalise month to YYYY-MM
        raw_month = row.get('Month', '')
        if hasattr(raw_month, 'year'):          # pandas Timestamp
            month_str = raw_month.strftime('%Y-%m')
        else:
            s = str(raw_month).strip()
            month_str = s[:7] if len(s) >= 7 else s   # take first 7 chars → YYYY-MM
        rows.append((
            str(row['Entry Date'])[:10],
            str(row.get('Stock Name','')).strip(),
            str(row.get('Type','')).strip(),
            _xl_safe(row.get('Quantity')),
            str(row.get('Buy/Sell','')).strip(),
            _xl_safe(row.get('Price')),
            _xl_safe(row.get('Invested Value')),
            _xl_safe(row.get('Current Value')),
            _xl_safe(row.get('Profit')),
            _xl_safe(row.get('Profit %')),
            str(row.get('Rationale','')).strip(),
            month_str,
        ))
    conn.executemany("""INSERT INTO invest_transactions
       (entry_date,stock_name,asset_type,quantity,action,price,invested_value,
        current_value,profit,profit_pct,rationale,month)
       VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", rows)
    conn.commit()
    return len(rows)

def _refresh_monthly_calc(conn):
    """Recompute monthly_investment_calc from invest_transactions.
    Normalises the month field to YYYY-MM regardless of how it was stored
    (handles 'YYYY-MM-DD', 'YYYY-MM-DD HH:MM:SS', plain 'YYYY-MM', etc.)."""
    conn.execute("DELETE FROM monthly_investment_calc")
    conn.execute("""
        INSERT INTO monthly_investment_calc
            (month, symbol, asset_type, qty_bought, qty_sold, net_qty,
             avg_buy_price, total_invested, updated_at)
        SELECT
            -- Normalise to YYYY-MM: use month col if it looks like a date, else fall back to entry_date
            substr(
                COALESCE(
                    CASE WHEN month GLOB '????-??*' THEN month ELSE NULL END,
                    CASE WHEN entry_date GLOB '????-??*' THEN entry_date ELSE NULL END,
                    month
                ), 1, 7
            ) AS norm_month,
            stock_name,
            MAX(asset_type),
            SUM(CASE WHEN UPPER(action) = 'BUY'  THEN quantity ELSE 0 END),
            SUM(CASE WHEN UPPER(action) = 'SELL' THEN quantity ELSE 0 END),
            SUM(CASE WHEN UPPER(action) = 'BUY'  THEN quantity ELSE -quantity END),
            CASE WHEN SUM(CASE WHEN UPPER(action)='BUY' THEN quantity ELSE 0 END) > 0
                 THEN SUM(CASE WHEN UPPER(action)='BUY' THEN invested_value ELSE 0 END) /
                      SUM(CASE WHEN UPPER(action)='BUY' THEN quantity ELSE 0 END)
                 ELSE 0 END,
            SUM(CASE WHEN UPPER(action) = 'BUY' THEN invested_value ELSE 0 END),
            datetime('now')
        FROM invest_transactions
        WHERE stock_name IS NOT NULL AND stock_name != ''
          AND (month IS NOT NULL OR entry_date IS NOT NULL)
        GROUP BY norm_month, stock_name
        HAVING norm_month IS NOT NULL AND norm_month != ''
    """)
    conn.commit()

# ── ASSETS TABLE HELPERS ─────────────────────────────────────────────────────

def _update_portfolio_from_assets(conn):
    """Recompute all portfolio rows by aggregating assets per InvestMapping.AssetID.
    Always uses the full 3-table JOIN: assets → AssetMapping → InvestMapping.
    """
    conn.execute("""
        INSERT OR REPLACE INTO portfolio (AssetID, InvestedValue, CurrentValue, ReturnValue, ReturnPCT, UpdateAt)
        SELECT
            im.AssetID,
            COALESCE(SUM(a.investedvalue), 0),
            COALESCE(SUM(a.currentvalue),  0),
            COALESCE(SUM(a.currentvalue),  0) - COALESCE(SUM(a.investedvalue), 0),
            CASE WHEN COALESCE(SUM(a.investedvalue), 0) > 0
                 THEN (COALESCE(SUM(a.currentvalue), 0) - COALESCE(SUM(a.investedvalue), 0))
                      / SUM(a.investedvalue) * 100
                 ELSE 0 END,
            datetime('now')
        FROM assets a
        JOIN AssetMapping am ON a.MappingID = am.MappingID
        JOIN InvestMapping im ON am.AssetId = im.AssetID
        GROUP BY im.AssetID
    """)
    conn.commit()

# ── ASSETS ENDPOINTS ──────────────────────────────────────────────────────────
# Base SELECT used by all read routes — returns old-style aliases for frontend compatibility
_ASSETS_SELECT = """
    SELECT
        a.AssetEntryID          AS id,
        a.AssetEntryID,
        a.MappingID,
        am.AssetName            AS asset,
        am.AssetSymbol          AS symbol,
        im.AssetType            AS asset_type,
        im.AssetClass           AS asset_class,
        im.AssetCategory        AS asset_category,
        im.AssetID              AS invest_id,
        a.purpose,
        a.qty,
        a.avgprice              AS avg_price,
        a.ltp,
        a.investedvalue         AS invested_value,
        a.currentvalue          AS current_value,
        a.pnl,
        a.pnlpct                AS pnl_pct,
        a.lastsynced            AS last_synced,
        a.updatedat             AS updated_at,
        a.targetpct             AS target_pct,
        COALESCE(a.sip_level_pct, 10)   AS sip_level_pct,
        COALESCE(a.max_allocation, 50000) AS max_allocation
    FROM assets a
    JOIN AssetMapping am ON a.MappingID = am.MappingID
    JOIN InvestMapping im ON am.AssetId = im.AssetID
"""

@app.route('/api/assets')
def api_assets_list():
    asset_type = request.args.get('type', '')
    conn = get_db()
    q = _ASSETS_SELECT + " WHERE 1=1"; p = []
    if asset_type:
        q += " AND LOWER(im.AssetType) LIKE ?"; p.append(f'%{asset_type.lower()}%')
    q += " ORDER BY im.AssetType, am.AssetName"
    rows = conn.execute(q, p).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/assets/rebuild', methods=['POST'])
def api_assets_rebuild():
    """Not supported — assets now require an AssetMapping entry (MappingID FK)."""
    return jsonify({'error': 'Rebuild not supported with new asset schema. Add assets via the Add Asset modal.'}), 400

@app.route('/api/assets/<int:aid>', methods=['PATCH'])
def api_assets_patch(aid):
    d = request.json; conn = get_db()
    row = conn.execute("SELECT * FROM assets WHERE AssetEntryID=?", (aid,)).fetchone()
    if not row:
        conn.close(); return jsonify({'error': 'Not found'}), 404
    ltp        = float(d.get('ltp',        row['ltp']       or 0))
    qty        = float(d.get('qty',        row['qty']       or 0))
    avg_price  = float(d.get('avg_price',  row['avgprice']  or 0))
    target_pct    = float(d.get('target_pct', row['targetpct'] if row['targetpct'] is not None else 25))
    sip_level_pct = float(d.get('sip_level_pct', row['sip_level_pct'] if row['sip_level_pct'] is not None else 10))
    max_allocation = float(d.get('max_allocation', row['max_allocation'] if row['max_allocation'] is not None else 50000))
    purpose    = d.get('purpose', row['purpose'] or '') or None
    invested = r2(qty * avg_price)
    current = r2(qty * ltp)
    pnl = r2(current - invested)
    pnl_pct = r2(pnl / invested * 100) if invested > 0 else 0
    conn.execute("""
        UPDATE assets SET ltp=?, qty=?, avgprice=?, targetpct=?, sip_level_pct=?, max_allocation=?, purpose=?,
            investedvalue=?, currentvalue=?, pnl=?, pnlpct=?,
            lastsynced=datetime('now'), updatedat=datetime('now')
        WHERE AssetEntryID=?
    """, (ltp, qty, avg_price, target_pct, sip_level_pct, max_allocation, purpose, invested, current, pnl, pnl_pct, aid))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/gtt_params')
def api_gtt_params_list():
    """Return all GTT params keyed by symbol, merged with assets table values."""
    conn = get_db()
    # Base: gtt_params table
    rows = conn.execute("SELECT symbol, target_pct, sip_level_pct, max_allocation FROM gtt_params").fetchall()
    result = {r['symbol'].upper(): {'target_pct': r['target_pct'], 'sip_level_pct': r['sip_level_pct'], 'max_allocation': r['max_allocation'], 'asset_id': None} for r in rows}
    # Merge assets table — prefer assets values for linked symbols
    assets = conn.execute("""
        SELECT a.AssetEntryID AS id, UPPER(TRIM(am.AssetSymbol)) AS sym,
               COALESCE(a.targetpct,25) AS target_pct,
               COALESCE(a.sip_level_pct,10) AS sip_level_pct,
               COALESCE(a.max_allocation,50000) AS max_allocation
        FROM assets a
        JOIN AssetMapping am ON a.MappingID = am.MappingID
        JOIN InvestMapping im ON am.AssetId = im.AssetID
        WHERE am.AssetSymbol IS NOT NULL AND am.AssetSymbol != ''
    """).fetchall()
    for a in assets:
        if a['sym']:
            result[a['sym']] = {'target_pct': a['target_pct'], 'sip_level_pct': a['sip_level_pct'], 'max_allocation': a['max_allocation'], 'asset_id': a['id']}
    conn.close()
    return jsonify(result)

@app.route('/api/gtt_params/<symbol>', methods=['PATCH'])
def api_gtt_params_patch(symbol):
    """Upsert GTT params for a symbol. Also syncs to assets table if linked."""
    d = request.json or {}
    sym = symbol.upper().strip()
    conn = get_db()
    target_pct    = float(d.get('target_pct',    25))
    sip_level_pct = float(d.get('sip_level_pct', 10))
    max_allocation = float(d.get('max_allocation', 50000))
    # Upsert gtt_params
    conn.execute("""
        INSERT INTO gtt_params (symbol, target_pct, sip_level_pct, max_allocation, updated_at)
        VALUES (?,?,?,?,datetime('now'))
        ON CONFLICT(symbol) DO UPDATE SET
            target_pct=excluded.target_pct,
            sip_level_pct=excluded.sip_level_pct,
            max_allocation=excluded.max_allocation,
            updated_at=excluded.updated_at
    """, (sym, target_pct, sip_level_pct, max_allocation))
    # Also sync to assets table if a matching asset exists
    conn.execute("""
        UPDATE assets SET targetpct=?, sip_level_pct=?, max_allocation=?, updatedat=datetime('now')
        WHERE AssetEntryID IN (
            SELECT a.AssetEntryID FROM assets a
            JOIN AssetMapping am ON a.MappingID=am.MappingID
            WHERE UPPER(TRIM(am.AssetSymbol))=?
        )
    """, (target_pct, sip_level_pct, max_allocation, sym))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/assets/sync_stocks', methods=['POST'])
def api_assets_sync_stocks():
    """Sync LTP for Stocks and ETF assets via yfinance, then update portfolio."""
    if not YF_AVAILABLE:
        return jsonify({'error': 'yfinance not installed — run: pip install yfinance'}), 503
    conn = get_db()
    rows = conn.execute(
        _ASSETS_SELECT + """
        WHERE UPPER(im.AssetType) IN ('STOCKS','EQUITY','DIRECT EQUITY','ETF')
           OR UPPER(im.AssetType) LIKE '%ETF%'
           OR UPPER(im.AssetType) LIKE '%STOCK%'
        """
    ).fetchall()
    if not rows:
        conn.close()
        return jsonify({'success': True, 'synced': 0, 'message': 'No stock/ETF assets found'})
    results = []
    for r in rows:
        raw = (r['symbol'] or '').strip()
        asset_name = (r['asset'] or '').strip()
        # Resolve ticker: handle full names (spaces), AMFI codes (digits), empty
        sym = raw.upper()
        if not sym or ' ' in sym or sym.isdigit() or len(sym) > 12:
            # Auto-search by asset name
            try:
                hits = yf.Search(asset_name or raw, max_results=5).quotes
                resolved = next((h['symbol'] for h in hits if h.get('symbol','').endswith('.NS')), None) or \
                           next((h['symbol'] for h in hits if h.get('symbol','').endswith('.BO')), None)
                if resolved:
                    sym = resolved
                    conn.execute("UPDATE AssetMapping SET AssetSymbol=? WHERE AssetName=? AND AssetSymbol=?",
                                 (sym.replace('.NS','').replace('.BO',''), asset_name, raw))
            except Exception:
                pass
        if not sym:
            results.append({'asset': asset_name, 'status': 'no_symbol'})
            continue
        if '.' not in sym:
            sym += '.NS'
        try:
            # NS→BO fallback: some stocks (e.g. NSDL) are only on BSE
            def _get_price(s):
                info = yf.Ticker(s).info
                p = float(info.get('currentPrice') or info.get('regularMarketPrice') or 0)
                if p == 0:
                    try: p = float(yf.Ticker(s).fast_info.last_price or 0)
                    except Exception: pass
                return p
            ltp = _get_price(sym)
            used_sym = sym
            if ltp == 0 and sym.upper().endswith('.NS'):
                bo = sym[:-3] + '.BO'
                ltp = _get_price(bo)
                if ltp > 0: used_sym = bo
            if ltp > 0:
                qty      = float(r['qty'])
                invested = float(r['invested_value'])
                current = r2(qty * ltp)
                pnl = r2(current - invested)
                pnl_pct = r2(pnl / invested * 100) if invested > 0 else 0
                conn.execute("""
                    UPDATE assets SET ltp=?, currentvalue=?, pnl=?, pnlpct=?,
                        lastsynced=datetime('now'), updatedat=datetime('now')
                    WHERE AssetEntryID=?
                """, (ltp, current, pnl, pnl_pct, r['id']))
                results.append({'asset': asset_name, 'symbol': used_sym, 'ltp': ltp, 'status': 'ok'})
            else:
                results.append({'asset': asset_name, 'symbol': sym, 'status': 'no_price'})
        except Exception as e:
            results.append({'asset': asset_name, 'symbol': sym, 'error': str(e), 'status': 'error'})
    conn.commit()
    _update_portfolio_from_assets(conn)
    conn.close()
    ok = sum(1 for r in results if r['status'] == 'ok')
    return jsonify({'success': True, 'synced': ok, 'failed': len(results) - ok, 'results': results})

@app.route('/api/assets/sync_mf', methods=['POST'])
def api_assets_sync_mf():
    """Sync NAV for Mutual Fund assets via MFAPI.in, then update portfolio."""
    import time
    conn = get_db()
    rows = conn.execute(
        _ASSETS_SELECT + " WHERE LOWER(im.AssetType) LIKE '%mutual%' OR LOWER(im.AssetType) = 'mf'"
    ).fetchall()
    if not rows:
        conn.close()
        return jsonify({'success': True, 'synced': 0, 'message': 'No mutual fund assets found'})
    results = []
    for r in rows:
        try:
            scheme_code = (r['symbol'] or '').strip() or None
            nav = None
            if not scheme_code:
                matches = _mf_search(r['asset'])
                if matches:
                    best_match = _mf_best_match(r['asset'], matches)
                    scheme_code = str(best_match['schemeCode'])
                    conn.execute(
                        "UPDATE AssetMapping SET AssetSymbol=? WHERE MappingID=?",
                        (scheme_code, r['MappingID'])
                    )
            if scheme_code:
                req2 = urllib.request.Request(
                    f'https://api.mfapi.in/mf/{scheme_code}',
                    headers={'User-Agent': 'Mozilla/5.0'}
                )
                with urllib.request.urlopen(req2, timeout=8) as resp2:
                    nav_data = _json.loads(resp2.read())
                nav = float(nav_data['data'][0]['nav'])
            if nav and nav > 0:
                qty      = float(r['qty'])
                invested = float(r['invested_value'])
                current  = qty * nav
                pnl = r2(current - invested)
                pnl_pct = r2(pnl / invested * 100) if invested > 0 else 0
                conn.execute("""
                    UPDATE assets SET ltp=?, currentvalue=?, pnl=?, pnlpct=?,
                        lastsynced=datetime('now'), updatedat=datetime('now')
                    WHERE AssetEntryID=?
                """, (nav, current, pnl, pnl_pct, r['id']))
                results.append({'asset': r['asset'], 'nav': nav, 'status': 'ok'})
            else:
                results.append({'asset': r['asset'], 'status': 'no_nav'})
        except Exception as e:
            results.append({'asset': r['asset'], 'error': str(e), 'status': 'error'})
        time.sleep(0.3)
    conn.commit()
    _update_portfolio_from_assets(conn)
    conn.close()
    ok = sum(1 for r in results if r['status'] == 'ok')
    return jsonify({'success': True, 'synced': ok, 'failed': len(results) - ok, 'results': results})

@app.route('/api/assets/sync_gold', methods=['POST'])
def api_assets_sync_gold():
    """Sync gold price for Physical Gold assets, then update portfolio."""
    try:
        req1 = urllib.request.Request(
            'https://api.gold-api.com/price/XAU', headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req1, timeout=6) as r:
            xau = _json.loads(r.read())
        usd_per_oz = float(xau['price'])
        req2 = urllib.request.Request(
            'https://api.frankfurter.app/latest?from=USD&to=INR', headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req2, timeout=6) as r:
            fx = _json.loads(r.read())
        inr_per_gram = round((usd_per_oz * float(fx['rates']['INR'])) / 31.1035)
    except Exception as e:
        return jsonify({'error': f'Failed to fetch gold price: {e}'}), 500

    conn = get_db()
    rows = conn.execute(
        _ASSETS_SELECT + """
        WHERE LOWER(im.AssetType) LIKE '%gold%'
          AND LOWER(im.AssetType) NOT LIKE '%etf%'
          AND LOWER(im.AssetType) NOT LIKE '%mf%'
          AND LOWER(im.AssetType) NOT LIKE '%mutual%'
        """
    ).fetchall()
    count = 0
    for r in rows:
        qty      = float(r['qty'])
        invested = float(r['invested_value'])
        current  = qty * inr_per_gram
        pnl = r2(current - invested)
        pnl_pct = r2(pnl / invested * 100) if invested > 0 else 0
        conn.execute("""
            UPDATE assets SET ltp=?, currentvalue=?, pnl=?, pnlpct=?,
                lastsynced=datetime('now'), updatedat=datetime('now')
            WHERE AssetEntryID=?
        """, (inr_per_gram, current, pnl, pnl_pct, r['id']))
        count += 1
    conn.commit()
    _update_portfolio_from_assets(conn)
    conn.close()
    return jsonify({'success': True, 'synced': count, 'price_inr_per_gram': inr_per_gram})

# ── EXCEL UPLOAD ──────────────────────────────────────────────────────────────
@app.route('/api/upload_excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only .xlsx / .xls files accepted'}), 400
    mode = request.form.get('mode', 'replace')
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        f.save(tmp.name); tmp.close()
        conn = get_db()
        if mode == 'replace':
            for tbl in ('transactions', 'loans', 'invest_transactions'):
                conn.execute(f'DELETE FROM {tbl}')
            conn.commit()
        results, errors = {}, {}
        for key, fn in [
            ('income',               _xl_income),
            ('expenses',             _xl_expenses),
            ('loans',                _xl_loans),
            ('investments',          _xl_investments),
            ('portfolio',            _xl_portfolio),
            ('invest_transactions',  _xl_invest_tx),
        ]:
            try:
                results[key] = fn(conn, tmp.name)
            except Exception as e:
                errors[key] = str(e)
        _refresh_monthly_calc(conn)
        # Portfolio is now derived from the assets table via _update_portfolio_from_assets.
        # invest_transactions → assets auto-rebuild is not supported in new schema
        # (assets require an AssetMapping entry). Skipping auto-rebuild.
        results['assets_rebuilt'] = 0
        conn.close()
        return jsonify({'success': True, 'mode': mode, 'results': results, 'errors': errors})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try: os.unlink(tmp.name)
        except: pass

# ── NSE MASTER ────────────────────────────────────────────────────────────────
@app.route('/api/nse')
def api_nse_list():
    conn = get_db()
    rows = conn.execute("SELECT * FROM nse_master ORDER BY symbol").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/nse', methods=['POST'])
def api_nse_add():
    d = request.json
    if not d or not d.get('symbol'):
        return jsonify({'error': 'symbol required'}), 400
    sym = d['symbol'].strip().upper()
    cat = d.get('category', 'Shares')
    conn = get_db()
    conn.execute(
        "INSERT OR IGNORE INTO nse_master (symbol, company_name, sector, category) VALUES (?,?,?,?)",
        (sym, d.get('company_name', ''), d.get('sector', ''), cat)
    )
    conn.commit(); conn.close()
    return jsonify({'success': True, 'symbol': sym})

@app.route('/api/nse/<symbol>', methods=['DELETE'])
def api_nse_delete(symbol):
    conn = get_db()
    conn.execute("DELETE FROM nse_master WHERE symbol=?", (symbol.upper(),))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/nse/auto_add', methods=['POST'])
def api_nse_auto_add():
    """Add distinct symbols from assets table WHERE asset LIKE %tab% → nse_master.
    Body: {tab: "Shares"}  (defaults to "Shares" if omitted)."""
    d   = request.get_json(silent=True) or {}
    tab = (d.get('tab') or 'Shares').strip()
    conn = get_db()
    rows = conn.execute("""
        SELECT DISTINCT am.AssetSymbol AS symbol
        FROM assets a
        JOIN AssetMapping am ON a.MappingID = am.MappingID
        JOIN InvestMapping im ON am.AssetId = im.AssetID
        WHERE LOWER(im.AssetType) LIKE LOWER(?)
          AND am.AssetSymbol IS NOT NULL AND am.AssetSymbol != ''
    """, (f'%{tab}%',)).fetchall()
    added = 0
    for r in rows:
        sym = r['symbol'].strip().upper()
        res = conn.execute(
            "INSERT OR IGNORE INTO nse_master (symbol, category) VALUES (?, ?)", (sym, tab)
        )
        if res.rowcount:
            added += 1
    conn.commit(); conn.close()
    return jsonify({'success': True, 'added': added, 'total': len(rows), 'tab': tab})

@app.route('/api/nse/auto_add_etf', methods=['POST'])
def api_nse_auto_add_etf():
    """Add distinct ETF names from invest_transactions → nse_master (category=ETF)."""
    conn = get_db()
    etfs = conn.execute("""SELECT DISTINCT stock_name FROM invest_transactions
                           WHERE (UPPER(asset_type) LIKE '%ETF%' OR asset_type = 'ETF')
                           AND stock_name IS NOT NULL AND stock_name != ''""").fetchall()
    added = 0
    for r in etfs:
        sym = r['stock_name'].strip().upper()
        res = conn.execute(
            "INSERT OR IGNORE INTO nse_master (symbol, category) VALUES (?, 'ETF')", (sym,)
        )
        if res.rowcount:
            added += 1
        else:
            # If it already exists, make sure category is set to ETF
            conn.execute(
                "UPDATE nse_master SET category='ETF' WHERE symbol=? AND (category IS NULL OR category='Shares')",
                (sym,)
            )
    conn.commit(); conn.close()
    return jsonify({'success': True, 'added': added, 'total': len(etfs)})

@app.route('/api/nse/sync', methods=['POST'])
def api_nse_sync():
    if not YF_AVAILABLE:
        return jsonify({'error': 'yfinance not installed — run: pip install yfinance'}), 503
    conn = get_db()
    rows = conn.execute("SELECT symbol FROM nse_master ORDER BY symbol").fetchall()
    symbols = [r['symbol'] for r in rows]
    if not symbols:
        conn.close()
        return jsonify({'success': True, 'message': 'No stocks to sync', 'results': []})
    results = []
    for sym in symbols:
        try:
            ticker = yf.Ticker(sym + '.NS')
            info   = ticker.info
            ltp    = float(info.get('currentPrice') or info.get('regularMarketPrice') or 0)
            prev   = float(info.get('previousClose') or 0)
            h52    = float(info.get('fiftyTwoWeekHigh') or 0)
            l52    = float(info.get('fiftyTwoWeekLow')  or 0)
            chg    = round((ltp - prev) / prev * 100, 2) if prev else 0
            frm52  = round((h52 - ltp) / h52 * 100, 2)  if h52  else 0
            name   = str(info.get('longName') or info.get('shortName') or sym)
            sect   = str(info.get('sector') or '')
            vol    = int(info.get('volume') or 0)
            conn.execute("""UPDATE nse_master SET ltp=?,prev_close=?,change_pct=?,
                            high_52w=?,low_52w=?,from_52w_high_pct=?,company_name=?,
                            sector=?,volume=?,updated_at=datetime('now') WHERE symbol=?""",
                         (ltp, prev, chg, h52, l52, frm52, name, sect, vol, sym))
            results.append({'symbol': sym, 'ltp': ltp, 'high_52w': h52, 'change_pct': chg, 'status': 'ok'})
        except Exception as e:
            results.append({'symbol': sym, 'error': str(e), 'status': 'error'})
    # Enrich monthly_investment_calc with live prices
    for r in results:
        if r['status'] == 'ok' and r['ltp'] > 0:
            conn.execute("""UPDATE monthly_investment_calc
                SET current_price=?,
                    current_value=net_qty*?,
                    unrealized_pnl=net_qty*?-total_invested,
                    unrealized_pnl_pct=CASE WHEN total_invested>0
                        THEN (net_qty*?-total_invested)/total_invested*100 ELSE 0 END,
                    updated_at=datetime('now')
                WHERE symbol=?""",
                (r['ltp'], r['ltp'], r['ltp'], r['ltp'], r['symbol']))
    conn.commit(); conn.close()
    ok = sum(1 for r in results if r['status'] == 'ok')
    return jsonify({'success': True, 'synced': ok, 'failed': len(results) - ok, 'results': results})

@app.route('/api/nse/sync_one', methods=['POST'])
def api_nse_sync_one():
    if not YF_AVAILABLE:
        return jsonify({'error': 'yfinance not installed'}), 503
    sym = (request.json or {}).get('symbol', '').strip().upper()
    if not sym:
        return jsonify({'error': 'symbol required'}), 400
    conn = get_db()
    try:
        ticker = yf.Ticker(sym + '.NS')
        info   = ticker.info
        ltp    = float(info.get('currentPrice') or info.get('regularMarketPrice') or 0)
        if ltp == 0:
            ticker2 = yf.Ticker(sym + '.BO')
            info2   = ticker2.info
            ltp     = float(info2.get('currentPrice') or info2.get('regularMarketPrice') or 0)
            if ltp: info = info2
        prev  = float(info.get('previousClose') or 0)
        h52   = float(info.get('fiftyTwoWeekHigh') or 0)
        l52   = float(info.get('fiftyTwoWeekLow')  or 0)
        chg   = round((ltp - prev) / prev * 100, 2) if prev else 0
        frm52 = round((h52 - ltp) / h52 * 100, 2)  if h52  else 0
        name  = str(info.get('longName') or info.get('shortName') or sym)
        conn.execute("""INSERT INTO nse_master(symbol,company_name,ltp,prev_close,change_pct,
                        high_52w,low_52w,from_52w_high_pct,updated_at)
                        VALUES(?,?,?,?,?,?,?,?,datetime('now'))
                        ON CONFLICT(symbol) DO UPDATE SET
                        ltp=excluded.ltp,prev_close=excluded.prev_close,
                        change_pct=excluded.change_pct,high_52w=excluded.high_52w,
                        low_52w=excluded.low_52w,from_52w_high_pct=excluded.from_52w_high_pct,
                        company_name=excluded.company_name,updated_at=excluded.updated_at""",
                     (sym, name, ltp, prev, chg, h52, l52, frm52))
        if ltp > 0:
            conn.execute("""UPDATE assets SET ltp=?,currentvalue=qty*ltp,
                            pnl=qty*?-investedvalue,
                            pnlpct=CASE WHEN investedvalue>0 THEN (qty*?-investedvalue)/investedvalue*100 ELSE 0 END,
                            lastsynced=datetime('now'),updatedat=datetime('now')
                            WHERE MappingID IN (SELECT MappingID FROM AssetMapping WHERE AssetSymbol=?) AND qty>0""",
                         (ltp, ltp, ltp, sym))
            conn.execute("""UPDATE monthly_investment_calc
                SET current_price=?,current_value=net_qty*?,
                    unrealized_pnl=net_qty*?-total_invested,
                    unrealized_pnl_pct=CASE WHEN total_invested>0
                        THEN (net_qty*?-total_invested)/total_invested*100 ELSE 0 END,
                    updated_at=datetime('now')
                WHERE symbol=?""", (ltp, ltp, ltp, ltp, sym))
        conn.commit()
        return jsonify({'success': True, 'symbol': sym, 'ltp': ltp, 'change_pct': chg})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# ── INVEST MAPPING ────────────────────────────────────────────────────────────

@app.route('/api/invest_mapping')
def api_invest_mapping():
    conn = get_db()
    rows = conn.execute("""
        SELECT id, AssetID, AssetClass, AssetCategory, AssetType,
               PriceFetchMode, Symbol, WeightGrams, Purity, InterestRate
        FROM InvestMapping
        ORDER BY AssetClass, AssetCategory, AssetType
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/invest_mapping', methods=['POST'])
def api_invest_mapping_add():
    """Add a new InvestMapping row (AssetClass + AssetCategory + AssetType must be unique)."""
    d = request.json or {}
    cls = (d.get('AssetClass') or '').strip()
    cat = (d.get('AssetCategory') or '').strip()
    typ = (d.get('AssetType') or '').strip()
    if not cls or not cat or not typ:
        return jsonify({'error': 'AssetClass, AssetCategory and AssetType are required'}), 400
    mode   = (d.get('PriceFetchMode') or 'MANUAL').strip().upper()
    sym    = (d.get('Symbol') or '').strip()
    wt     = float(d.get('WeightGrams') or 1)
    purity = (d.get('Purity') or '24K').strip()
    rate   = float(d.get('InterestRate') or 0)
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO InvestMapping
                (AssetClass, AssetCategory, AssetType, PriceFetchMode, Symbol, WeightGrams, Purity, InterestRate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (cls, cat, typ, mode, sym, wt, purity, rate))
        conn.commit()
        row = conn.execute("""
            SELECT id, AssetID, AssetClass, AssetCategory, AssetType,
                   PriceFetchMode, Symbol, WeightGrams, Purity, InterestRate
            FROM InvestMapping WHERE AssetClass=? AND AssetCategory=? AND AssetType=?
        """, (cls, cat, typ)).fetchone()
        conn.close()
        return jsonify(dict(row)), 201
    except Exception as e:
        conn.close()
        if 'UNIQUE' in str(e).upper():
            return jsonify({'error': f'"{cls} › {cat} › {typ}" already exists'}), 409
        return jsonify({'error': str(e)}), 500

@app.route('/api/invest_mapping/<asset_id>', methods=['PATCH'])
def api_invest_mapping_patch(asset_id):
    """Update PriceFetchMode and related fields for a single InvestMapping row."""
    d = request.json or {}
    allowed = {'PriceFetchMode', 'Symbol', 'WeightGrams', 'Purity', 'InterestRate',
               'AssetClass', 'AssetCategory', 'AssetType'}
    updates = {k: v for k, v in d.items() if k in allowed}
    if not updates:
        return jsonify({'error': 'No valid fields to update'}), 400
    conn = get_db()
    row = conn.execute("SELECT id FROM InvestMapping WHERE AssetID=?", (asset_id,)).fetchone()
    if not row:
        conn.close(); return jsonify({'error': 'Not found'}), 404
    set_clause = ', '.join(f'{k}=?' for k in updates)
    conn.execute(f"UPDATE InvestMapping SET {set_clause} WHERE AssetID=?",
                 list(updates.values()) + [asset_id])
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/invest_mapping/<asset_id>', methods=['DELETE'])
def api_invest_mapping_delete(asset_id):
    """Cascade-delete an InvestMapping row and all dependent data:
       assets → AssetMapping → portfolio → InvestMapping
    """
    conn = get_db()
    im_row = conn.execute(
        "SELECT id, AssetClass, AssetCategory, AssetType FROM InvestMapping WHERE AssetID=?",
        (asset_id,)
    ).fetchone()
    if not im_row:
        conn.close(); return jsonify({'error': 'Not found'}), 404

    # 1. Find all AssetMapping rows that reference this InvestMapping
    mapping_rows = conn.execute(
        "SELECT MappingID FROM AssetMapping WHERE AssetId=?", (asset_id,)
    ).fetchall()
    mapping_ids = [r['MappingID'] for r in mapping_rows]

    removed_assets = 0
    if mapping_ids:
        placeholders = ','.join('?' * len(mapping_ids))
        # 2. Count + delete all assets rows linked via AssetMapping
        removed_assets = conn.execute(
            f"SELECT COUNT(*) FROM assets WHERE MappingID IN ({placeholders})", mapping_ids
        ).fetchone()[0]
        conn.execute(
            f"DELETE FROM assets WHERE MappingID IN ({placeholders})", mapping_ids
        )
        # 3. Delete AssetMapping rows
        conn.execute("DELETE FROM AssetMapping WHERE AssetId=?", (asset_id,))

    # 4. Delete portfolio row (FK → InvestMapping.AssetID)
    conn.execute("DELETE FROM portfolio WHERE AssetID=?", (asset_id,))

    # 5. Delete InvestMapping row itself
    conn.execute("DELETE FROM InvestMapping WHERE AssetID=?", (asset_id,))
    conn.commit()

    # 6. Recompute portfolio totals from remaining assets
    _update_portfolio_from_assets(conn)
    conn.commit()
    conn.close()

    label = f"{im_row['AssetClass']} › {im_row['AssetCategory']} › {im_row['AssetType']}"
    return jsonify({
        'success':          True,
        'label':            label,
        'removed_assets':   removed_assets,
        'removed_mappings': len(mapping_ids),
    })

@app.route('/api/market/prices')
def api_market_prices():
    """Return current gold/silver spot prices and USD→INR rate."""
    result = {'gold_per_gram_24k': None, 'silver_per_gram': None, 'usd_inr': None, 'error': None}
    try:
        req_xau = urllib.request.Request('https://api.gold-api.com/price/XAU',
            headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req_xau, timeout=6) as r: xau = _json.loads(r.read())
        req_xag = urllib.request.Request('https://api.gold-api.com/price/XAG',
            headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req_xag, timeout=6) as r: xag = _json.loads(r.read())
        req_fx = urllib.request.Request('https://api.frankfurter.app/latest?from=USD&to=INR',
            headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req_fx, timeout=6) as r: fx = _json.loads(r.read())
        usd_inr = float(fx['rates']['INR'])
        result['usd_inr']           = round(usd_inr, 2)
        result['gold_per_gram_24k'] = round((float(xau['price']) * usd_inr) / 31.1035, 2)
        result['silver_per_gram']   = round((float(xag['price']) * usd_inr) / 31.1035, 2)
    except Exception as e:
        result['error'] = str(e)
    return jsonify(result)

# ── WEALTH TRACKER ────────────────────────────────────────────────────────────

def _wt_cascade(conn, asset_id):
    """Recompute the portfolio row for the given InvestMapping.AssetID.
    Uses assets → AssetMapping → InvestMapping JOIN tree.
    """
    if not asset_id:
        return
    row = conn.execute("""
        SELECT COALESCE(SUM(a.investedvalue), 0) AS iv,
               COALESCE(SUM(a.currentvalue),  0) AS cv
        FROM assets a
        JOIN AssetMapping am ON a.MappingID = am.MappingID
        WHERE am.AssetId = ?
    """, (asset_id,)).fetchone()
    iv = float(row['iv'] or 0)
    cv = float(row['cv'] or 0)
    ret = r2(cv - iv)
    ret_pct = r2(ret / iv * 100) if iv > 0 else 0
    conn.execute("""
        INSERT OR REPLACE INTO portfolio (AssetID, InvestedValue, CurrentValue, ReturnValue, ReturnPCT, UpdateAt)
        VALUES (?, ?, ?, ?, ?, datetime('now'))
    """, (asset_id, iv, cv, ret, ret_pct))

@app.route('/api/wealth')
def api_wealth_list():
    """All wealth goals with computed current_value / achieved_pct from portfolio."""
    from datetime import date as _date, datetime as _datetime
    conn = get_db(); c = conn.cursor()
    import json as _json
    goals = [dict(r) for r in c.execute("SELECT * FROM wealth ORDER BY id").fetchall()]
    for g in goals:
        raw = g.get('linked_asset_ids') or ''
        try: g['linked_asset_ids'] = _json.loads(raw) if raw else []
        except Exception: g['linked_asset_ids'] = []
    today = _date.today()
    # Total portfolio value from the assets table (most up-to-date)
    total_row = c.execute("""
        SELECT COALESCE(SUM(a.currentvalue),0)  total_cv,
               COALESCE(SUM(a.investedvalue),0) total_ti
        FROM assets a
    """).fetchone()
    grand_cv = float(total_row['total_cv'])
    grand_ti = float(total_row['total_ti'])

    for g in goals:
        # Try purpose-specific match first (assets.purpose or portfolio.Purpose)
        row_asset = c.execute("""
            SELECT COALESCE(SUM(a.currentvalue),0)  cv,
                   COALESCE(SUM(a.investedvalue),0) ti
            FROM assets a WHERE LOWER(a.purpose)=LOWER(?)
        """, (g['purpose'],)).fetchone()
        cv = float(row_asset['cv']); ti = float(row_asset['ti'])

        # If no assets tagged to this purpose, fall back to old portfolio table
        if cv == 0:
            row_port = c.execute("""
                SELECT COALESCE(SUM(CurrentValue),0)  cv,
                       COALESCE(SUM(InvestedValue),0) ti
                FROM portfolio WHERE Purpose=?
            """, (g['purpose'],)).fetchone()
            cv = float(row_port['cv']); ti = float(row_port['ti'])

        # If still 0 (no tagged assets anywhere), use grand total portfolio value
        # so the overall progress bar reflects real wealth vs goal
        if cv == 0:
            cv = grand_cv; ti = grand_ti

        g['current_value']  = round(cv, 2)
        g['total_invested'] = round(ti, 2)
        g['total_return']   = round(cv - ti, 2)
        g['achieved_pct']   = round(cv / g['target'] * 100, 2) if g['target'] > 0 else 0
        g['remaining']      = round(max(0.0, g['target'] - cv), 2)
        # Time remaining fields
        td = g.get('target_date')
        if td:
            try:
                target_d = _date.fromisoformat(td[:10])
                days_left = (target_d - today).days
                g['days_remaining'] = days_left
                # Created date for elapsed % calculation
                created_raw = g.get('created_at', '')
                if created_raw:
                    created_d = _datetime.fromisoformat(created_raw[:10]).date()
                    total_days = (target_d - created_d).days
                    elapsed    = (today - created_d).days
                    g['time_elapsed_pct'] = round(min(100, max(0, elapsed / total_days * 100)), 1) if total_days > 0 else 0
                else:
                    g['time_elapsed_pct'] = 0
            except Exception:
                g['days_remaining']   = None
                g['time_elapsed_pct'] = 0
        else:
            g['days_remaining']   = None
            g['time_elapsed_pct'] = 0
    conn.close()
    return jsonify(goals)

def _wealth_assign_assets(conn, purpose, asset_ids):
    """Assign purpose to all assets rows whose MappingID links to the given AssetIDs."""
    import json as _json
    if not asset_ids:
        return
    placeholders = ','.join('?' * len(asset_ids))
    # Update assets table via MappingID → AssetMapping → AssetId
    conn.execute(f"""
        UPDATE assets SET purpose=?
        WHERE MappingID IN (
            SELECT MappingID FROM AssetMapping
            WHERE AssetId IN ({placeholders})
        )
    """, [purpose] + list(asset_ids))
    # Update portfolio table directly via AssetID
    conn.execute(f"""
        UPDATE portfolio SET Purpose=?
        WHERE AssetID IN ({placeholders})
    """, [purpose] + list(asset_ids))

@app.route('/api/wealth', methods=['POST'])
def api_wealth_add():
    import json as _json
    d = request.json or {}
    purpose = (d.get('purpose') or '').strip()
    if not purpose:
        return jsonify({'error': 'purpose required'}), 400
    target      = float(d.get('target', 0))
    target_date = (d.get('target_date') or '').strip() or None
    asset_ids   = d.get('linked_asset_ids') or []          # list of AssetIDs e.g. ["Asset11","Asset09"]
    linked_json = _json.dumps(asset_ids) if asset_ids else ''
    conn = get_db()
    try:
        conn.execute("INSERT INTO wealth (purpose, target, target_date, linked_asset_ids) VALUES (?,?,?,?)",
                     (purpose, target, target_date, linked_json))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400
    nid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    if asset_ids:
        _wealth_assign_assets(conn, purpose, asset_ids)
        conn.commit()
    conn.close()
    return jsonify({'success': True, 'id': nid})

@app.route('/api/wealth/<int:wid>', methods=['PUT'])
def api_wealth_update(wid):
    import json as _json
    d = request.json or {}; conn = get_db()
    fields = []; params = []
    if 'purpose'     in d: fields.append("purpose=?");     params.append(d['purpose'].strip())
    if 'target'      in d: fields.append("target=?");      params.append(float(d['target']))
    if 'target_date' in d: fields.append("target_date=?"); params.append((d['target_date'] or '').strip() or None)
    if 'linked_asset_ids' in d:
        asset_ids = d['linked_asset_ids'] or []
        fields.append("linked_asset_ids=?")
        params.append(_json.dumps(asset_ids) if asset_ids else '')
    if fields:
        params.append(wid)
        conn.execute(f"UPDATE wealth SET {','.join(fields)} WHERE id=?", params)
        conn.commit()
    # Re-assign assets if provided
    if 'linked_asset_ids' in d and 'purpose' in d:
        _wealth_assign_assets(conn, d['purpose'].strip(), d['linked_asset_ids'] or [])
        conn.commit()
    elif 'linked_asset_ids' in d:
        # get current purpose from DB
        row = conn.execute("SELECT purpose FROM wealth WHERE id=?", (wid,)).fetchone()
        if row:
            _wealth_assign_assets(conn, row['purpose'], d['linked_asset_ids'] or [])
            conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/wealth/<int:wid>', methods=['DELETE'])
def api_wealth_delete(wid):
    conn = get_db()
    conn.execute("DELETE FROM wealth WHERE id=?", (wid,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/wt/portfolio')
def api_wt_portfolio_list():
    """Portfolio rows joined to InvestMapping for full asset class/category/type tree."""
    conn = get_db()
    purpose = request.args.get('purpose', '')
    q = """
        SELECT p.AssetID        AS asset_id,
               im.AssetClass    AS asset_class,
               im.AssetCategory AS asset_category,
               im.AssetType     AS asset_type,
               p.InvestedValue  AS invested_value,
               p.CurrentValue   AS current_value,
               p.ReturnValue    AS return_value,
               p.ReturnPCT      AS return_pct,
               p.Purpose        AS purpose
        FROM portfolio p
        JOIN InvestMapping im ON p.AssetID = im.AssetID
        WHERE 1=1
    """
    params = []
    if purpose:
        q += " AND p.Purpose=?"; params.append(purpose)
    q += " ORDER BY im.AssetClass, im.AssetCategory, im.AssetType"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/wt/portfolio/assign', methods=['POST'])
def api_wt_portfolio_assign():
    """Assign a Purpose to a portfolio row by AssetID (preferred) or asset_type fallback."""
    d = request.json or {}; conn = get_db()
    purpose = d.get('purpose', '') or None
    if 'asset_id' in d:
        conn.execute("UPDATE portfolio SET Purpose=? WHERE AssetID=?", (purpose, d['asset_id']))
    elif 'asset_type' in d:
        conn.execute("""UPDATE portfolio SET Purpose=?
                        WHERE AssetID IN (SELECT AssetID FROM InvestMapping WHERE AssetType=?)""",
                     (purpose, d['asset_type']))
    elif 'asset' in d:
        # Legacy: match by AssetType or AssetClass name
        conn.execute("""UPDATE portfolio SET Purpose=?
                        WHERE AssetID IN (
                            SELECT AssetID FROM InvestMapping WHERE AssetType=? OR AssetClass=?
                        )""", (purpose, d['asset'], d['asset']))
    conn.commit(); conn.close()
    return jsonify({'success': True})

def _compute_xirr(cashflows):
    """Newton's method XIRR. cashflows = list of (days_from_t0, amount).
    Returns annualised rate or None if it fails to converge."""
    if not cashflows or sum(a for _, a in cashflows) <= 0:
        return None
    has_neg = any(a < 0 for _, a in cashflows)
    has_pos = any(a > 0 for _, a in cashflows)
    if not (has_neg and has_pos):
        return None
    def npv(r):
        return sum(a / (1 + r) ** (d / 365.0) for d, a in cashflows)
    def dnpv(r):
        return sum(-d / 365.0 * a / (1 + r) ** (d / 365.0 + 1) for d, a in cashflows)
    rate = 0.15
    for _ in range(200):
        f = npv(rate); df = dnpv(rate)
        if abs(df) < 1e-14:
            break
        new_rate = rate - f / df
        if new_rate <= -1:
            new_rate = -0.9999
        if abs(new_rate - rate) < 1e-8:
            rate = new_rate
            break
        rate = new_rate
    if abs(npv(rate)) > 1.0:   # didn't converge cleanly
        return None
    return round(rate * 100, 2)

@app.route('/api/wt/xirr')
def api_wt_xirr():
    """Compute XIRR% per asset_type using invest_transactions cashflows + current portfolio value.
    Optional ?asset_class= filter matches InvestMapping.AssetClass."""
    from datetime import date as _date
    asset_class = request.args.get('asset_class', '').strip()
    conn = get_db()
    today = _date.today()

    # ── Current value per InvestMapping AssetType (from live assets) ──────────
    cur_q = """
        SELECT im.AssetType, SUM(a.currentvalue) AS cur_val
        FROM assets a
        JOIN AssetMapping am ON a.MappingID = am.MappingID
        JOIN InvestMapping im ON am.AssetId = im.AssetID
        WHERE a.currentvalue > 0
    """
    cur_params = []
    if asset_class:
        cur_q += " AND im.AssetClass=?"; cur_params.append(asset_class)
    cur_q += " GROUP BY im.AssetType"
    cur_by_type = {r['AssetType']: float(r['cur_val'])
                   for r in conn.execute(cur_q, cur_params).fetchall()}

    # ── Cashflows from invest_transactions ────────────────────────────────────
    tx_rows = conn.execute("""
        SELECT entry_date, action, invested_value, asset_type
        FROM invest_transactions
        WHERE entry_date IS NOT NULL AND invested_value > 0
        ORDER BY entry_date
    """).fetchall()
    conn.close()

    # Normalise invest_tx asset_type → InvestMapping AssetType key
    def _norm(t):
        t = (t or '').strip()
        if t.lower() == 'stock':   return 'Stocks'
        if t.lower() == 'stocks':  return 'Stocks'
        return t   # 'Mutual Fund' matches as-is

    # Group cashflows by normalised type
    from collections import defaultdict
    cf_by_type = defaultdict(list)
    for row in tx_rows:
        try:
            d = _date.fromisoformat(str(row['entry_date'])[:10])
        except Exception:
            continue
        typ   = _norm(row['asset_type'])
        amt   = float(row['invested_value'])
        action = (row['action'] or '').lower()
        # BUY = outflow (negative), SELL = inflow (positive)
        cf_by_type[typ].append((d, -amt if action in ('buy','purchase') else amt))

    # Build XIRR for each type that has cashflow data
    result = {}
    for typ, flows in cf_by_type.items():
        cur_val = cur_by_type.get(typ, 0)
        if cur_val <= 0:
            # Try fuzzy match (e.g. "Mutual Fund" vs "Liquid Mutual Fund")
            for k, v in cur_by_type.items():
                if typ.lower() in k.lower() or k.lower() in typ.lower():
                    cur_val += v
        if cur_val <= 0:
            continue
        t0 = min(d for d, _ in flows)
        cfs = [(( d - t0).days, a) for d, a in flows]
        # Terminal: receive current portfolio value today
        cfs.append(((today - t0).days, cur_val))
        xirr = _compute_xirr(cfs)
        if xirr is not None:
            result[typ] = xirr

    return jsonify(result)

@app.route('/api/wt/assets')
def api_wt_assets_list():
    conn = get_db()
    purpose     = request.args.get('purpose', '')
    asset_type  = request.args.get('asset_type', '')
    asset_class = request.args.get('asset_class', '')
    q = _ASSETS_SELECT + " WHERE 1=1"; p = []
    if purpose:     q += " AND a.purpose=?";                     p.append(purpose)
    if asset_class: q += " AND im.AssetClass=?";                 p.append(asset_class)
    if asset_type:  q += " AND LOWER(im.AssetType) LIKE ?";      p.append(f'%{asset_type.lower()}%')
    q += " ORDER BY im.AssetType, am.AssetName"
    rows = conn.execute(q, p).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/wt/assets', methods=['POST'])
def api_wt_assets_add():
    d = request.json or {}; conn = get_db()

    # ── New flow: asset_id from InvestMapping ──────────────────────────────────
    invest_id = (d.get('invest_id') or '').strip()
    if invest_id:
        im = conn.execute(
            "SELECT AssetID, AssetClass, AssetCategory, AssetType FROM InvestMapping WHERE AssetID=?",
            (invest_id,)
        ).fetchone()
        if not im:
            conn.close(); return jsonify({'error': f'InvestMapping ID {invest_id!r} not found'}), 400

        asset_name   = (d.get('asset_name') or '').strip()
        asset_symbol = (d.get('asset_symbol') or '').strip()
        if not asset_name:
            conn.close(); return jsonify({'error': 'Asset Name is required'}), 400

        qty           = float(d.get('qty', 0) or 0)
        avg_price     = float(d.get('avg_price', 0) or 0)
        ltp_raw       = d.get('ltp')
        ltp           = float(ltp_raw) if ltp_raw not in (None, '', 0) and float(ltp_raw or 0) > 0 else avg_price
        target_pct    = float(d.get('target_pct', 25) or 25)
        sip_level_pct = float(d.get('sip_level_pct', 10) or 10)
        max_allocation = float(d.get('max_allocation', 50000) or 50000)
        invested = r2(qty * avg_price)
        current = r2(qty * ltp)
        pnl = r2(current - invested)
        pnl_pct = r2(pnl / invested * 100) if invested > 0 else 0

        # Upsert AssetMapping — one row per (AssetId, AssetName, AssetSymbol) combo
        existing = conn.execute(
            "SELECT MappingID FROM AssetMapping WHERE AssetId=? AND AssetName=? AND AssetSymbol=?",
            (invest_id, asset_name, asset_symbol)
        ).fetchone()
        if existing:
            mapping_id = existing['MappingID']
        else:
            cur = conn.execute(
                "INSERT INTO AssetMapping (AssetName, AssetSymbol, AssetId) VALUES (?,?,?)",
                (asset_name, asset_symbol, invest_id)
            )
            mapping_id = cur.lastrowid
        conn.commit()

        conn.execute("""
            INSERT INTO assets
                (MappingID, purpose, qty, avgprice, ltp,
                 investedvalue, currentvalue, pnl, pnlpct,
                 targetpct, sip_level_pct, max_allocation, updatedat)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
        """, (mapping_id, None, qty, avg_price, ltp, invested, current, pnl, pnl_pct,
              target_pct, sip_level_pct, max_allocation))
        conn.commit()
        _wt_cascade(conn, invest_id)
        conn.commit(); conn.close()
        return jsonify({'success': True})

    conn.close()
    return jsonify({'error': 'invest_id is required'}), 400

@app.route('/api/wt/assets/<int:aid>', methods=['PATCH'])
def api_wt_assets_patch(aid):
    d = request.json or {}; conn = get_db()
    row = conn.execute(_ASSETS_SELECT + " WHERE a.AssetEntryID=?", (aid,)).fetchone()
    if not row:
        conn.close(); return jsonify({'error': 'Not found'}), 404
    ltp       = float(d.get('ltp',       row['ltp']       or 0))
    qty       = float(d.get('qty',       row['qty']       or 0))
    avg_price = float(d.get('avg_price', row['avg_price'] or 0))
    purpose   = (d.get('purpose', row['purpose'] or '') or '') or None
    invested = r2(qty * avg_price)
    current = r2(qty * ltp)
    pnl = r2(current - invested)
    pnl_pct = r2(pnl / invested * 100) if invested > 0 else 0
    conn.execute("""
        UPDATE assets SET ltp=?, qty=?, avgprice=?, purpose=?,
            investedvalue=?, currentvalue=?, pnl=?, pnlpct=?,
            lastsynced=datetime('now'), updatedat=datetime('now')
        WHERE AssetEntryID=?
    """, (ltp, qty, avg_price, purpose, invested, current, pnl, pnl_pct, aid))
    conn.commit()
    _wt_cascade(conn, row['invest_id'])
    conn.commit(); conn.close()
    return jsonify({'success': True})

@app.route('/api/wt/assets/<int:aid>', methods=['DELETE'])
def api_wt_assets_delete(aid):
    conn = get_db()
    row = conn.execute(_ASSETS_SELECT + " WHERE a.AssetEntryID=?", (aid,)).fetchone()
    conn.execute("DELETE FROM assets WHERE AssetEntryID=?", (aid,))
    conn.commit()
    if row:
        _wt_cascade(conn, row['invest_id'])
        conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/wt/assets/csv_upload', methods=['POST'])
def api_wt_assets_csv_upload():
    """Bulk-import assets from CSV.
    Required: invest_id, asset_name, qty, avg_price
    Optional: asset_symbol, ltp
    For each row: upserts AssetMapping then inserts into assets."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith('.csv'):
        return jsonify({'error': 'Only .csv files accepted'}), 400
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
    try:
        f.save(tmp.name); tmp.close()
        df = pd.read_csv(tmp.name)
        df.columns = [c.strip().lower() for c in df.columns]
        required = {'invest_id', 'asset_name', 'qty', 'avg_price'}
        missing  = required - set(df.columns)
        if missing:
            return jsonify({'error': f'Missing columns: {", ".join(missing)}'}), 400

        conn = get_db()
        ok = 0; errors = []
        for i, row in df.iterrows():
            try:
                invest_id    = str(row.get('invest_id', '')).strip()
                asset_name   = str(row.get('asset_name', '')).strip()
                asset_symbol = str(row.get('asset_symbol', '')).strip() if 'asset_symbol' in row else ''
                if not invest_id or not asset_name or invest_id.lower() == 'nan' or asset_name.lower() == 'nan':
                    errors.append(f'Row {i+2}: invest_id and asset_name required'); continue

                im = conn.execute(
                    "SELECT AssetID, AssetClass, AssetCategory, AssetType FROM InvestMapping WHERE AssetID=?",
                    (invest_id,)
                ).fetchone()
                if not im:
                    errors.append(f'Row {i+2}: InvestMapping ID {invest_id!r} not found'); continue

                qty       = float(row.get('qty', 0) or 0)
                avg_price = float(row.get('avg_price', 0) or 0)
                ltp_raw   = str(row.get('ltp', '')).strip() if 'ltp' in row else ''
                ltp       = float(ltp_raw) if ltp_raw not in ('', 'nan', '0') and float(ltp_raw or 0) > 0 else avg_price
                tgt_raw   = str(row.get('target_pct', '')).strip() if 'target_pct' in df.columns else ''
                target_pct = float(tgt_raw) if tgt_raw not in ('', 'nan') else 25.0
                sip_raw   = str(row.get('sip_level_pct', '')).strip() if 'sip_level_pct' in df.columns else ''
                sip_level_pct = float(sip_raw) if sip_raw not in ('', 'nan') else 10.0
                ma_raw    = str(row.get('max_allocation', '')).strip() if 'max_allocation' in df.columns else ''
                max_allocation = float(ma_raw) if ma_raw not in ('', 'nan') else 50000.0
                invested = r2(qty * avg_price)
                current = r2(qty * ltp)
                pnl = r2(current - invested)
                pnl_pct = r2(pnl / invested * 100) if invested > 0 else 0.0

                # Upsert AssetMapping
                existing = conn.execute(
                    "SELECT MappingID FROM AssetMapping WHERE AssetId=? AND AssetName=? AND AssetSymbol=?",
                    (invest_id, asset_name, asset_symbol)
                ).fetchone()
                if existing:
                    mapping_id = existing['MappingID']
                else:
                    cur = conn.execute(
                        "INSERT INTO AssetMapping (AssetName, AssetSymbol, AssetId) VALUES (?,?,?)",
                        (asset_name, asset_symbol, invest_id)
                    )
                    mapping_id = cur.lastrowid

                conn.execute("""
                    INSERT INTO assets
                        (MappingID, qty, avgprice, ltp, investedvalue, currentvalue, pnl, pnlpct,
                         targetpct, sip_level_pct, max_allocation, updatedat)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
                """, (mapping_id, qty, avg_price, ltp, invested, current, pnl, pnl_pct,
                      target_pct, sip_level_pct, max_allocation))
                ok += 1
            except Exception as e:
                errors.append(f'Row {i+2}: {e}')

        conn.commit()
        _update_portfolio_from_assets(conn)
        conn.commit(); conn.close()
        return jsonify({'success': True, 'imported': ok, 'errors': errors})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try: os.unlink(tmp.name)
        except: pass

@app.route('/api/wt/assets/sample_csv')
def api_wt_assets_sample_csv():
    """Download a sample CSV template — new schema (invest_id, asset_name, asset_symbol, qty, avg_price, ltp).
    invest_id references InvestMapping.AssetID (Asset01–Asset15)."""
    sample = (
        "invest_id,asset_name,asset_symbol,qty,avg_price,ltp,target_pct,sip_level_pct,max_allocation\n"
        "Asset09,NIFTY 50 ETF,NIFTYBEES,500,200,250,25,10,50000\n"
        "Asset08,Axis ELSS Tax Saver Direct Plan Growth,147070,303.371,75.07,,25,10,50000\n"
        "Asset02,Physical Gold,,100,5500,,25,10,50000\n"
        "Asset13,PPF Account,,1,500000,550000,25,10,50000\n"
        "Asset14,Employer PF,,1,300000,350000,25,10,50000\n"
        "Asset12,NPS Tier 1,,1,400000,450000\n"
        "Asset07,Reliance Industries,RELIANCE,50,2400,2650\n"
        "Asset05,Gold ETF,GOLDBEES,200,55,62\n"
        "Asset10,Plot - Sector 12,,1,2500000,2800000\n"
        "Asset15,PM Kisaan Yojana,,1,50000,50000\n"
    )
    from flask import Response
    return Response(
        sample,
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=wealthflow_assets_sample.csv'}
    )

@app.route('/api/sample/<filename>')
@login_required
def api_sample_csv(filename):
    """Central sample-CSV endpoint. Returns a template CSV for any upload screen."""
    from flask import Response
    samples = {
        # Manual Investment upload (investmentmonth, quantity, price)
        'manual_investment.csv': (
            "investmentmonth,quantity,price\n"
            "2025-01,50,245.50\n"
            "2025-02,30,258.00\n"
            "2025-03,40,271.75\n"
            "2025-04,25,265.20\n"
        ),
        # Zerodha tradebook (CSV export from Console → Tradebook)
        'zerodha_tradebook.csv': (
            "Symbol,ISIN,Trade Date,Exchange,Trade Type,Quantity,Price,Order ID,Order Execution Time\n"
            "RELIANCE,INE002A01018,2024-01-10,NSE,BUY,10,2450.00,1001234567,09:15:00\n"
            "INFY,INE009A01021,2024-01-15,NSE,BUY,20,1520.50,1001234568,10:02:30\n"
            "RELIANCE,INE002A01018,2024-06-20,NSE,SELL,5,2680.00,1001234569,14:35:10\n"
            "HDFCBANK,INE040A01034,2024-02-05,NSE,BUY,15,1650.00,1001234570,11:00:00\n"
        ),
        # Zerodha stock holdings (CSV from Console → Portfolio → Holdings)
        'zerodha_holdings.csv': (
            "Instrument,ISIN,Qty,Avg cost,LTP,Cur val,P&L,Net chg,Day chg\n"
            "RELIANCE,INE002A01018,10,2450.00,2680.00,26800.00,2300.00,9.39%,0.45%\n"
            "INFY,INE009A01021,20,1520.50,1620.00,32400.00,1990.00,6.54%,-0.20%\n"
            "HDFCBANK,INE040A01034,15,1650.00,1710.00,25650.00,900.00,3.64%,0.10%\n"
        ),
        # Groww stock holdings
        'groww_stock_holdings.csv': (
            "Stock Name,Symbol,ISIN,Quantity,Average Price,Current Price,Current Value,P&L,P&L %\n"
            "Reliance Industries,RELIANCE,INE002A01018,10,2450.00,2680.00,26800.00,2300.00,9.39\n"
            "Infosys,INFY,INE009A01021,20,1520.50,1620.00,32400.00,1990.00,6.54\n"
        ),
        # Groww MF holdings
        'groww_mf_holdings.csv': (
            "Scheme Name,ISIN,Units,Average NAV,Current NAV,Current Value,P&L,P&L %\n"
            "SBI Bluechip Fund Direct Growth,INF200K01RO2,500.000,62.30,72.10,36050.00,4900.00,15.73\n"
            "Axis ELSS Tax Saver Direct Growth,INF846K01EW2,300.000,75.07,89.20,26760.00,4239.00,18.83\n"
        ),
        # Groww stock orders
        'groww_stock_orders.csv': (
            "Stock name,Symbol,ISIN,Type,Quantity,Value,Exchange,Exchange Order Id,Execution date and time,Order status\n"
            "Reliance Industries,RELIANCE,INE002A01018,BUY,10,24500.00,NSE,GROW1001,2024-01-10 09:15:00,COMPLETE\n"
            "Infosys,INFY,INE009A01021,BUY,20,30410.00,NSE,GROW1002,2024-01-15 10:02:30,COMPLETE\n"
            "Reliance Industries,RELIANCE,INE002A01018,SELL,5,13400.00,NSE,GROW1003,2024-06-20 14:35:10,COMPLETE\n"
        ),
        # Groww MF orders
        'groww_mf_orders.csv': (
            "Scheme Name,Transaction Type,Units,NAV,Amount,Date\n"
            "SBI Bluechip Fund Direct Growth,BUY,100.000,62.30,6230.00,2024-01-05\n"
            "SBI Bluechip Fund Direct Growth,BUY,80.000,65.10,5208.00,2024-03-05\n"
            "Axis ELSS Tax Saver Direct Growth,BUY,150.000,75.07,11260.50,2024-02-01\n"
        ),
    }
    content = samples.get(filename)
    if not content:
        return 'Not found', 404
    return Response(
        content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/api/wt/last_sync')
def api_wt_last_sync():
    """Return the most recent lastsynced timestamp across all assets."""
    conn = get_db()
    row = conn.execute(
        "SELECT MAX(lastsynced) AS last_synced FROM assets WHERE lastsynced IS NOT NULL"
    ).fetchone()
    conn.close()
    return jsonify({'last_synced': row['last_synced'] if row else None})

@app.route('/api/wt/sync', methods=['POST'])
def api_wt_sync():
    """Sync LTP for all assets, routed by InvestMapping.PriceFetchMode."""
    import time as _time
    results = []
    conn = get_db()

    # Fetch all assets with their InvestMapping fetch configuration
    all_rows = conn.execute("""
        SELECT a.AssetEntryID,
               am.AssetName   AS asset_name,
               am.AssetSymbol AS asset_symbol,
               am.MappingID,
               im.AssetID     AS invest_id,
               im.AssetType   AS asset_type,
               im.AssetClass  AS asset_class,
               im.PriceFetchMode,
               im.Symbol      AS im_symbol,
               im.WeightGrams,
               im.Purity,
               im.InterestRate,
               a.qty,
               a.avgprice,
               a.investedvalue,
               a.currentvalue,
               a.ltp
        FROM assets a
        JOIN AssetMapping am ON a.MappingID = am.MappingID
        JOIN InvestMapping im ON am.AssetId = im.AssetID
        ORDER BY im.PriceFetchMode, im.AssetType
    """).fetchall()

    if not all_rows:
        conn.close()
        return jsonify({'success': True, 'synced': 0, 'failed': 0, 'skipped': 0, 'results': []})

    # Group rows by PriceFetchMode
    by_mode = {}
    for r in all_rows:
        mode = (r['PriceFetchMode'] or 'MANUAL').upper()
        by_mode.setdefault(mode, []).append(r)

    # Shared forex/commodity cache (fetched once per sync call)
    _fx_rate   = None   # USD → INR
    _gold_gram = None   # INR per gram 24K
    _silver_gram = None # INR per gram

    def _get_forex():
        nonlocal _fx_rate
        if _fx_rate is None:
            req = urllib.request.Request('https://api.frankfurter.app/latest?from=USD&to=INR',
                headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=8) as resp:
                _fx_rate = float(_json.loads(resp.read())['rates']['INR'])
        return _fx_rate

    # ── 1. EQUITY — yfinance (stocks, ETF, SGB, listed bonds) ─────────────────
    def _fetch_equity_price(sym):
        """Fetch LTP for an equity ticker. Returns (price, symbol_used).
        Strategy: try sym as-is → if 0 and sym ends .NS, retry with .BO → fast_info fallback."""
        def _price_from_ticker(s):
            info = yf.Ticker(s).info
            p = float(info.get('currentPrice') or info.get('regularMarketPrice') or 0)
            if p == 0:
                try:
                    p = float(yf.Ticker(s).fast_info.last_price or 0)
                except Exception:
                    pass
            return p

        price = _price_from_ticker(sym)
        if price > 0:
            return price, sym
        # NS returned 0 → try BSE
        if sym.upper().endswith('.NS'):
            bo_sym = sym[:-3] + '.BO'
            price = _price_from_ticker(bo_sym)
            if price > 0:
                return price, bo_sym
        # BO returned 0 → try NS
        if sym.upper().endswith('.BO'):
            ns_sym = sym[:-3] + '.NS'
            price = _price_from_ticker(ns_sym)
            if price > 0:
                return price, ns_sym
        return 0, sym

    def _resolve_equity_ticker(raw_sym, asset_name):
        """Return a clean NSE/BSE ticker (with exchange suffix) or None.
        Handles: valid ticker, name-as-symbol (has spaces), AMFI code (all digits),
        ISIN-like strings. Falls back to yf.Search by asset name."""
        s = (raw_sym or '').strip().upper()
        # Already has exchange suffix — trust it
        if s and '.' in s:
            return s
        # Looks like a valid short ticker (no spaces, ≤12 chars, not pure digits)
        if s and ' ' not in s and len(s) <= 12 and not s.isdigit():
            return s + '.NS'
        # Symbol is absent, has spaces (full name entered), or is an AMFI/ISIN code
        # → auto-resolve via yfinance Search using asset name
        search_term = asset_name or s
        try:
            hits = yf.Search(search_term, max_results=5).quotes
            # Prefer NSE (.NS) over BSE (.BO)
            for h in hits:
                ticker_sym = h.get('symbol', '')
                if ticker_sym.endswith('.NS'):
                    return ticker_sym
            for h in hits:
                ticker_sym = h.get('symbol', '')
                if ticker_sym.endswith('.BO'):
                    return ticker_sym
        except Exception:
            pass
        return None

    if 'EQUITY' in by_mode:
        if not YF_AVAILABLE:
            for r in by_mode['EQUITY']:
                results.append({'asset': r['asset_name'], 'status': 'error',
                                 'error': 'yfinance not installed', 'mode': 'EQUITY'})
        else:
            for r in by_mode['EQUITY']:
                raw = (r['asset_symbol'] or r['im_symbol'] or '').strip()
                sym = _resolve_equity_ticker(raw, r['asset_name'])
                if not sym:
                    results.append({'asset': r['asset_name'], 'status': 'no_symbol', 'mode': 'EQUITY',
                                    'hint': 'Set NSE ticker in asset symbol field (e.g. BANKBEES)'})
                    continue
                # If we auto-resolved from a name/bad symbol, save the correct ticker back
                if raw.upper() != sym.replace('.NS', '').replace('.BO', '') and r['MappingID']:
                    clean_ticker = sym.replace('.NS', '').replace('.BO', '')
                    conn.execute("UPDATE AssetMapping SET AssetSymbol=? WHERE MappingID=?",
                                 (clean_ticker, r['MappingID']))
                try:
                    ltp, used_sym = _fetch_equity_price(sym)
                    if ltp > 0:
                        # If BSE fallback was used, persist .BO symbol so next sync is instant
                        if used_sym != sym and r['MappingID']:
                            conn.execute("UPDATE AssetMapping SET AssetSymbol=? WHERE MappingID=?",
                                         (used_sym.replace('.NS','').replace('.BO',''), r['MappingID']))
                        qty = float(r['qty'] or 0); inv = float(r['investedvalue'] or 0)
                        cur = qty * ltp; pnl = r2(cur - inv)
                        pnl_pct = r2(pnl / inv * 100) if inv > 0 else 0
                        conn.execute("""UPDATE assets SET ltp=?,currentvalue=?,pnl=?,pnlpct=?,
                            lastsynced=datetime('now'),updatedat=datetime('now')
                            WHERE AssetEntryID=?""", (ltp, cur, pnl, pnl_pct, r['AssetEntryID']))
                        results.append({'asset': r['asset_name'], 'symbol': used_sym, 'ltp': ltp,
                                        'status': 'ok', 'mode': 'EQUITY'})
                    else:
                        results.append({'asset': r['asset_name'], 'symbol': sym,
                                        'status': 'no_price', 'mode': 'EQUITY',
                                        'hint': 'Not on NSE or BSE — check symbol'})
                except Exception as e:
                    results.append({'asset': r['asset_name'], 'symbol': sym,
                                    'status': 'error', 'error': str(e)[:120], 'mode': 'EQUITY'})
                _time.sleep(0.15)

    # ── 2. MF — MFAPI.in (AMFI scheme code as asset_symbol) ───────────────────
    if 'MF' in by_mode:
        for r in by_mode['MF']:
            sc = (r['asset_symbol'] or '').strip()
            try:
                if not sc or not sc.isdigit():
                    # Auto-search by fund name (with fallback for long names)
                    matches = _mf_search(r['asset_name'])
                    if matches:
                        sc = str(_mf_best_match(r['asset_name'], matches)['schemeCode'])
                        # Save scheme code back to AssetMapping so next sync is instant
                        conn.execute("UPDATE AssetMapping SET AssetSymbol=? WHERE MappingID=?",
                                     (sc, r['MappingID']))
                if sc and sc.isdigit():
                    req2 = urllib.request.Request(f'https://api.mfapi.in/mf/{sc}',
                        headers={'User-Agent': 'Mozilla/5.0'})
                    with urllib.request.urlopen(req2, timeout=8) as resp2:
                        nav_data = _json.loads(resp2.read())
                    nav = float(nav_data['data'][0]['nav'])
                    if nav > 0:
                        qty = float(r['qty'] or 0); inv = float(r['investedvalue'] or 0)
                        cur = qty * nav; pnl = r2(cur - inv)
                        pnl_pct = r2(pnl / inv * 100) if inv > 0 else 0
                        conn.execute("""UPDATE assets SET ltp=?,currentvalue=?,pnl=?,pnlpct=?,
                            lastsynced=datetime('now'),updatedat=datetime('now')
                            WHERE AssetEntryID=?""", (nav, cur, pnl, pnl_pct, r['AssetEntryID']))
                        results.append({'asset': r['asset_name'], 'scheme_code': sc, 'nav': nav,
                                        'status': 'ok', 'mode': 'MF'})
                    else:
                        results.append({'asset': r['asset_name'], 'scheme_code': sc,
                                        'status': 'no_nav', 'mode': 'MF'})
                else:
                    results.append({'asset': r['asset_name'], 'status': 'no_scheme_code',
                                    'mode': 'MF', 'hint': 'Set AMFI scheme code in asset symbol field'})
                _time.sleep(0.2)
            except Exception as e:
                results.append({'asset': r['asset_name'], 'status': 'error',
                                 'error': str(e)[:120], 'mode': 'MF'})

    # ── 3. COMMODITY_GOLD — physical gold priced by purity + weight ────────────
    if 'COMMODITY_GOLD' in by_mode:
        try:
            req_xau = urllib.request.Request('https://api.gold-api.com/price/XAU',
                headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req_xau, timeout=8) as rr:
                xau_usd = float(_json.loads(rr.read())['price'])
            fx = _get_forex()
            gold_per_gram_24k = (xau_usd * fx) / 31.1035   # INR per gram pure gold
            _gold_gram = gold_per_gram_24k
            PURITY = {'24K': 1.0, '22K': 22/24, '20K': 20/24, '18K': 18/24}
            for r in by_mode['COMMODITY_GOLD']:
                pf  = PURITY.get((r['Purity'] or '24K').strip(), 1.0)
                wt  = float(r['WeightGrams'] or 1)
                ltp = round(gold_per_gram_24k * pf * wt, 2)
                qty = float(r['qty'] or 0); inv = float(r['investedvalue'] or 0)
                cur = qty * ltp; pnl = r2(cur - inv)
                pnl_pct = r2(pnl / inv * 100) if inv > 0 else 0
                conn.execute("""UPDATE assets SET ltp=?,currentvalue=?,pnl=?,pnlpct=?,
                    lastsynced=datetime('now'),updatedat=datetime('now')
                    WHERE AssetEntryID=?""", (ltp, cur, pnl, pnl_pct, r['AssetEntryID']))
                results.append({'asset': r['asset_name'], 'ltp': ltp,
                                 'purity': r['Purity'], 'weight_g': wt,
                                 'status': 'ok', 'mode': 'COMMODITY_GOLD'})
        except Exception as e:
            for r in by_mode.get('COMMODITY_GOLD', []):
                results.append({'asset': r['asset_name'], 'status': 'error',
                                 'error': str(e)[:120], 'mode': 'COMMODITY_GOLD'})

    # ── 4. COMMODITY_SILVER — physical silver priced per gram ──────────────────
    if 'COMMODITY_SILVER' in by_mode:
        try:
            req_xag = urllib.request.Request('https://api.gold-api.com/price/XAG',
                headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req_xag, timeout=8) as rr:
                xag_usd = float(_json.loads(rr.read())['price'])
            fx = _get_forex()
            silver_per_gram = (xag_usd * fx) / 31.1035
            _silver_gram = silver_per_gram
            for r in by_mode['COMMODITY_SILVER']:
                wt  = float(r['WeightGrams'] or 1)
                ltp = round(silver_per_gram * wt, 2)
                qty = float(r['qty'] or 0); inv = float(r['investedvalue'] or 0)
                cur = qty * ltp; pnl = r2(cur - inv)
                pnl_pct = r2(pnl / inv * 100) if inv > 0 else 0
                conn.execute("""UPDATE assets SET ltp=?,currentvalue=?,pnl=?,pnlpct=?,
                    lastsynced=datetime('now'),updatedat=datetime('now')
                    WHERE AssetEntryID=?""", (ltp, cur, pnl, pnl_pct, r['AssetEntryID']))
                results.append({'asset': r['asset_name'], 'ltp': ltp, 'weight_g': wt,
                                 'status': 'ok', 'mode': 'COMMODITY_SILVER'})
        except Exception as e:
            for r in by_mode.get('COMMODITY_SILVER', []):
                results.append({'asset': r['asset_name'], 'status': 'error',
                                 'error': str(e)[:120], 'mode': 'COMMODITY_SILVER'})

    # ── 5. RATE_BASED — PPF / EPF / SSY / NPS / FD (annual interest accrual) ──
    if 'RATE_BASED' in by_mode:
        for r in by_mode['RATE_BASED']:
            rate = float(r['InterestRate'] or 0)
            if rate <= 0:
                results.append({'asset': r['asset_name'], 'status': 'no_rate',
                                 'mode': 'RATE_BASED',
                                 'hint': 'Set InterestRate % in InvestMapping config'})
                continue
            try:
                inv = float(r['investedvalue'] or 0)
                # Current value = principal × (1 + annual_rate/100)
                # User maintains the principal (balance); we apply one year's accrual
                cur     = inv * (1 + rate / 100)
                pnl = r2(cur - inv)
                pnl_pct = rate   # return % = the interest rate itself
                ltp     = rate   # ltp field stores the current interest rate for display
                conn.execute("""UPDATE assets SET ltp=?,currentvalue=?,pnl=?,pnlpct=?,
                    lastsynced=datetime('now'),updatedat=datetime('now')
                    WHERE AssetEntryID=?""", (ltp, cur, pnl, pnl_pct, r['AssetEntryID']))
                results.append({'asset': r['asset_name'], 'rate': rate, 'current_value': cur,
                                 'status': 'ok', 'mode': 'RATE_BASED'})
            except Exception as e:
                results.append({'asset': r['asset_name'], 'status': 'error',
                                 'error': str(e)[:120], 'mode': 'RATE_BASED'})

    # ── 6. MANUAL — user manages values; skip silently ─────────────────────────
    for r in by_mode.get('MANUAL', []):
        results.append({'asset': r['asset_name'], 'status': 'skipped', 'mode': 'MANUAL'})

    conn.commit()
    _update_portfolio_from_assets(conn)
    conn.commit(); conn.close()

    ok      = sum(1 for r in results if r['status'] == 'ok')
    skipped = sum(1 for r in results if r['status'] == 'skipped')
    failed  = len(results) - ok - skipped
    return jsonify({
        'success': True, 'synced': ok, 'failed': failed, 'skipped': skipped,
        'results': results,
    })

# ── MONTHLY INVESTMENT CALC ───────────────────────────────────────────────────
@app.route('/api/monthly_investment_calc')
def api_monthly_calc():
    month = request.args.get('month', '')
    conn  = get_db()
    q = "SELECT * FROM monthly_investment_calc WHERE 1=1"; p = []
    if month: q += " AND month LIKE ?"; p.append(f'{month}%')
    q += " ORDER BY month DESC, total_invested DESC"
    rows = conn.execute(q, p).fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/monthly_investment_calc/refresh', methods=['POST'])
def api_monthly_calc_refresh():
    conn = get_db()
    _refresh_monthly_calc(conn)
    cnt = conn.execute("SELECT COUNT(*) FROM monthly_investment_calc").fetchone()[0]
    conn.close()
    return jsonify({'success': True, 'rows': cnt})

# ══════════════════════════════════════════════════════════════════════════════
# TRADING STRATEGY
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/trading_strategy')
def api_trading_strategy():
    """Return shares from assets table enriched with NSE live data + last tx."""
    conn = get_db()
    # Use the canonical 3-table JOIN (assets → AssetMapping → InvestMapping)
    # then LEFT JOIN nse_master for live price data
    assets = conn.execute("""
        SELECT
            a.AssetEntryID          AS id,
            am.AssetName            AS asset,
            am.AssetSymbol          AS symbol,
            im.AssetType            AS asset_type,
            a.qty,
            a.avgprice              AS avg_price,
            a.ltp                   AS db_ltp,
            COALESCE(a.targetpct, 25)        AS target_pct,
            COALESCE(a.sip_level_pct, 10)    AS sip_level_pct,
            COALESCE(a.max_allocation, 50000) AS max_allocation,
            a.investedvalue         AS invested_value,
            a.currentvalue          AS current_value,
            a.lastsynced            AS last_synced,
            COALESCE(n.ltp,        a.ltp, 0)            AS ltp,
            COALESCE(n.high_52w,   0)                    AS high_52w,
            COALESCE(n.low_52w,    0)                    AS low_52w,
            COALESCE(n.from_52w_high_pct, 0)            AS from_52w_high_pct,
            COALESCE(n.change_pct, 0)                    AS day_change_pct,
            COALESCE(n.company_name, am.AssetName)       AS company_name,
            n.updated_at AS nse_updated
        FROM assets a
        JOIN AssetMapping am ON a.MappingID = am.MappingID
        JOIN InvestMapping im ON am.AssetId = im.AssetID
        LEFT JOIN nse_master n ON UPPER(TRIM(am.AssetSymbol)) = UPPER(TRIM(n.symbol))
        WHERE LOWER(im.AssetType) IN ('stocks','stock','shares','equity','etf')
           OR LOWER(im.AssetType) LIKE '%etf%'
           OR LOWER(im.AssetType) LIKE '%share%'
           OR LOWER(im.AssetType) LIKE '%stock%'
        ORDER BY am.AssetSymbol
    """).fetchall()

    # Last buy per symbol from invest_transactions
    last_buys = conn.execute("""
        SELECT t.stock_name, t.price, t.entry_date, t.quantity
        FROM invest_transactions t
        INNER JOIN (
            SELECT stock_name, MAX(entry_date) AS max_date
            FROM invest_transactions WHERE LOWER(action)='buy'
            GROUP BY stock_name
        ) m ON t.stock_name = m.stock_name AND t.entry_date = m.max_date
        WHERE LOWER(t.action) = 'buy'
    """).fetchall()
    last_buy_map = {r['stock_name']: dict(r) for r in last_buys}

    # Last sell per symbol
    last_sells = conn.execute("""
        SELECT t.stock_name, t.price, t.entry_date, t.quantity
        FROM invest_transactions t
        INNER JOIN (
            SELECT stock_name, MAX(entry_date) AS max_date
            FROM invest_transactions WHERE LOWER(action)='sell'
            GROUP BY stock_name
        ) m ON t.stock_name = m.stock_name AND t.entry_date = m.max_date
        WHERE LOWER(t.action) = 'sell'
    """).fetchall()
    last_sell_map = {r['stock_name']: dict(r) for r in last_sells}

    rows = []
    for a in assets:
        sym  = (a['symbol'] or '').strip().upper()
        name = (a['asset'] or '').strip()
        lb   = last_buy_map.get(sym) or last_buy_map.get(name) or {}
        ls   = last_sell_map.get(sym) or last_sell_map.get(name) or {}
        rows.append({
            'id':                a['id'],
            'symbol':            sym,
            'asset':             name,
            'company_name':      a['company_name'] or name,
            'qty':               float(a['qty'] or 0),
            'avg_price':         float(a['avg_price'] or 0),
            'ltp':               float(a['ltp'] or 0),
            'invested_value':    float(a['invested_value'] or 0),
            'current_value':     float(a['current_value'] or 0),
            'high_52w':          float(a['high_52w'] or 0),
            'low_52w':           float(a['low_52w'] or 0),
            'from_52w_high_pct': float(a['from_52w_high_pct'] or 0),
            'day_change_pct':    float(a['day_change_pct'] or 0),
            'target_pct':        float(a['target_pct'] or 25),
            'nse_updated':       a['nse_updated'] or '',
            'last_buy':  {'price': float(lb.get('price') or 0),
                          'date':  lb.get('entry_date',''),
                          'qty':   float(lb.get('quantity') or 0)},
            'last_sell': {'price': float(ls.get('price') or 0),
                          'date':  ls.get('entry_date',''),
                          'qty':   float(ls.get('quantity') or 0)},
        })
    conn.close()
    return jsonify(rows)


# ─────────────────────────────────────────────────────────────────────────────
# BROKER RAW DATA UPLOAD
# ─────────────────────────────────────────────────────────────────────────────

# Parsing configuration for each source type
_BROKER_PARSERS = {
    'groww_mf_orders': {
        'broker': 'groww', 'instrument': 'mutual_fund', 'data_type': 'orders',
        'sheet': 0,           # first sheet
        'header_hint': 'Scheme Name',   # scan rows until this cell found in col 0
        'col_map': {
            'Scheme Name': 'name', 'Transaction Type': 'trade_type',
            'Units': 'quantity', 'NAV': 'price', 'Amount': 'amount', 'Date': 'trade_date'
        }
    },
    'groww_stock_orders': {
        'broker': 'groww', 'instrument': 'stocks', 'data_type': 'orders',
        'sheet': 0,
        'header_hint': 'Stock name',
        'col_map': {
            'Stock name': 'name', 'Symbol': 'symbol', 'ISIN': 'isin',
            'Type': 'trade_type', 'Quantity': 'quantity', 'Value': 'amount',
            'Exchange': 'exchange', 'Exchange Order Id': 'order_id',
            'Execution date and time': 'trade_date', 'Order status': 'status'
        }
    },
    'zerodha_stock_trades': {
        'broker': 'zerodha', 'instrument': 'stocks', 'data_type': 'trades',
        'sheet': 0,
        'header_hint': 'Symbol',
        'col_map': {
            'Symbol': 'symbol', 'ISIN': 'isin', 'Trade Date': 'trade_date',
            'Exchange': 'exchange', 'Trade Type': 'trade_type',
            'Quantity': 'quantity', 'Price': 'price', 'Order ID': 'order_id',
            'Order Execution Time': 'status'
        }
    },
    'groww_mf_holdings': {
        'broker': 'groww', 'instrument': 'mutual_fund', 'data_type': 'holdings',
        'sheet': 0, 'header_hint': None, 'col_map': {}
    },
    'groww_stock_holdings': {
        'broker': 'groww', 'instrument': 'stocks', 'data_type': 'holdings',
        'sheet': 0, 'header_hint': None, 'col_map': {}
    },
    'zerodha_stock_holdings': {
        'broker': 'zerodha', 'instrument': 'stocks', 'data_type': 'holdings',
        'sheet': 0, 'header_hint': None, 'col_map': {}
    },
}

def _normalise_date(raw):
    """Convert any common date string to YYYY-MM-DD.
    Handles: ISO YYYY-MM-DD, DD-MM-YYYY HH:MM AM/PM (Groww),
    DD-MM-YYYY, DD/MM/YYYY, DD Mon YYYY, and openpyxl datetime objects."""
    from datetime import datetime as _dt
    if not raw:
        return ''
    if isinstance(raw, _dt):
        return raw.strftime('%Y-%m-%d')
    s = str(raw).strip()
    if not s:
        return ''

    # Try full datetime formats first — covers "DD-MM-YYYY HH:MM AM/PM" (Groww CSV)
    # and ISO datetime variants before we strip the time component.
    for fmt in (
        '%d-%m-%Y %I:%M %p',   # "06-05-2026 09:34 AM"  ← Groww stock orders
        '%d-%m-%Y %H:%M',      # "06-05-2026 09:34"
        '%d/%m/%Y %I:%M %p',   # "06/05/2026 09:34 AM"
        '%d/%m/%Y %H:%M',      # "06/05/2026 09:34"
        '%Y-%m-%d %H:%M:%S',   # "2026-05-06 09:34:00"
        '%Y-%m-%d %H:%M',      # "2026-05-06 09:34"
        '%d %b %Y %H:%M:%S',
        '%d %b %Y %I:%M %p',
    ):
        try:
            return _dt.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass

    # Strip trailing time / punctuation to get bare date token
    date_part = s.split()[0].rstrip(',.;') if s else s

    # ISO YYYY-MM-DD — validate both separators to avoid "YYYY-M-D" pass-through
    if (len(date_part) >= 10
            and date_part[4] == '-' and date_part[7] == '-'):
        try:
            _dt.strptime(date_part[:10], '%Y-%m-%d')
            return date_part[:10]
        except ValueError:
            pass

    # DD Mon YYYY  e.g. "26 Aug 2021"
    for fmt in ('%d %b %Y', '%d %B %Y', '%d-%b-%Y', '%d-%B-%Y'):
        try:
            return _dt.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass

    # DD/MM/YYYY or DD-MM-YYYY (prefer day-first / Indian format over US)
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return _dt.strptime(date_part, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass

    # Fallback: return the date token as-is (will be caught by _MONTH_SQL NULL handling)
    return date_part if date_part else (s[:10] if len(s) >= 10 else s)


def _parse_broker_file(file_storage, source_type):
    """Parse uploaded Excel/CSV file into list of dicts based on source_type config."""
    import openpyxl, io, csv, json as _json
    cfg = _BROKER_PARSERS.get(source_type)
    if not cfg:
        raise ValueError(f"Unknown source_type: {source_type}")

    fname = file_storage.filename or ''
    content = file_storage.read()

    rows_raw = []
    if fname.lower().endswith('.csv'):
        text = content.decode('utf-8-sig', errors='replace')
        reader = csv.reader(io.StringIO(text))
        rows_raw = list(reader)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet_idx = cfg.get('sheet', 0)
        ws = wb.worksheets[sheet_idx] if isinstance(sheet_idx, int) else wb[sheet_idx]
        rows_raw = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]

    # Find header row
    header_row_idx = 0
    hint = cfg.get('header_hint')
    if hint:
        for i, row in enumerate(rows_raw):
            row_vals = [str(v).strip() if v is not None else '' for v in row]
            if hint in row_vals:
                header_row_idx = i
                break

    headers = [str(v).strip() if v is not None else '' for v in rows_raw[header_row_idx]]

    col_map = cfg.get('col_map', {})
    broker   = cfg['broker']
    instrument = cfg['instrument']
    data_type  = cfg['data_type']

    result = []
    for row in rows_raw[header_row_idx + 1:]:
        # Skip completely blank rows
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        raw = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                raw[headers[j]] = str(val).strip() if val is not None else ''
        if not raw:
            continue

        def g(col): return raw.get(col, '')
        def gf(col):
            try: return float(str(raw.get(col, '') or '').replace(',', ''))
            except: return None

        if col_map:
            record = {
                'broker': broker, 'instrument': instrument, 'data_type': data_type,
                'name':       g(next((k for k,v in col_map.items() if v=='name'), '')),
                'symbol':     g(next((k for k,v in col_map.items() if v=='symbol'), '')),
                'isin':       g(next((k for k,v in col_map.items() if v=='isin'), '')),
                'trade_type': g(next((k for k,v in col_map.items() if v=='trade_type'), '')),
                'trade_date': g(next((k for k,v in col_map.items() if v=='trade_date'), '')),
                'quantity':   gf(next((k for k,v in col_map.items() if v=='quantity'), '')),
                'price':      gf(next((k for k,v in col_map.items() if v=='price'), '')),
                'amount':     gf(next((k for k,v in col_map.items() if v=='amount'), '')),
                'exchange':   g(next((k for k,v in col_map.items() if v=='exchange'), '')),
                'order_id':   g(next((k for k,v in col_map.items() if v=='order_id'), '')),
                'status':     g(next((k for k,v in col_map.items() if v=='status'), '')),
                'raw_data':   _json.dumps(raw),
            }
        else:
            # Generic: store everything in raw_data
            record = {
                'broker': broker, 'instrument': instrument, 'data_type': data_type,
                'name': '', 'symbol': '', 'isin': '', 'trade_type': '',
                'trade_date': '', 'quantity': None, 'price': None, 'amount': None,
                'exchange': '', 'order_id': '', 'status': '',
                'raw_data': _json.dumps(raw),
            }
        result.append(record)
    return result


@app.route('/api/broker_upload/<source_type>', methods=['POST'])
def api_broker_upload(source_type):
    """Upload a broker file, parse it, store rows in raw_upload_data."""
    if source_type not in _BROKER_PARSERS:
        return jsonify({'error': f'Unknown source type: {source_type}'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'Empty filename'}), 400

    try:
        rows = _parse_broker_file(f, source_type)
    except Exception as e:
        return jsonify({'error': f'Parse error: {str(e)}'}), 400

    conn = get_db()
    # Delete previous uploads for this source_type
    old_ids = [r[0] for r in conn.execute(
        "SELECT id FROM raw_upload_meta WHERE source_type=?", (source_type,)).fetchall()]
    if old_ids:
        placeholders = ','.join('?' * len(old_ids))
        conn.execute(f"DELETE FROM raw_upload_data WHERE upload_id IN ({placeholders})", old_ids)
        conn.execute(f"DELETE FROM raw_upload_meta WHERE id IN ({placeholders})", old_ids)

    # Insert meta
    conn.execute(
        "INSERT INTO raw_upload_meta (source_type, file_name, row_count) VALUES (?,?,?)",
        (source_type, f.filename, len(rows)))
    upload_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # Insert rows
    for row in rows:
        conn.execute("""
            INSERT INTO raw_upload_data
                (upload_id, source_type, broker, instrument, data_type,
                 name, symbol, isin, trade_type, trade_date,
                 quantity, price, amount, exchange, order_id, status, raw_data)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (upload_id, source_type, row['broker'], row['instrument'], row['data_type'],
              row['name'], row['symbol'], row['isin'], row['trade_type'], row['trade_date'],
              row['quantity'], row['price'], row['amount'], row['exchange'],
              row['order_id'], row['status'], row['raw_data']))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'rows': len(rows), 'upload_id': upload_id})


@app.route('/api/broker_uploads')
def api_broker_uploads_meta():
    """Return metadata for all uploads, grouped by source_type."""
    conn = get_db()
    rows = conn.execute(
        "SELECT source_type, file_name, row_count, uploaded_at FROM raw_upload_meta ORDER BY uploaded_at DESC"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/broker_uploads/<source_type>')
def api_broker_upload_rows(source_type):
    """Return up to 50 rows for preview for a given source_type."""
    conn = get_db()
    rows = conn.execute("""
        SELECT d.name, d.symbol, d.isin, d.trade_type, d.trade_date,
               d.quantity, d.price, d.amount, d.exchange, d.order_id, d.status, d.raw_data
        FROM raw_upload_data d
        JOIN raw_upload_meta m ON m.id = d.upload_id
        WHERE d.source_type=?
        ORDER BY d.id DESC LIMIT 50
    """, (source_type,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/broker_uploads/<source_type>', methods=['DELETE'])
def api_broker_upload_delete(source_type):
    """Clear all uploaded data for a source_type."""
    conn = get_db()
    old_ids = [r[0] for r in conn.execute(
        "SELECT id FROM raw_upload_meta WHERE source_type=?", (source_type,)).fetchall()]
    if old_ids:
        placeholders = ','.join('?' * len(old_ids))
        conn.execute(f"DELETE FROM raw_upload_data WHERE upload_id IN ({placeholders})", old_ids)
        conn.execute(f"DELETE FROM raw_upload_meta WHERE id IN ({placeholders})", old_ids)
    conn.commit(); conn.close()
    return jsonify({'success': True})


@app.route('/api/manual_investment/asset_list')
def api_manual_investment_asset_list():
    """Return InvestMapping rows for the asset selector dropdown."""
    conn = get_db()
    rows = conn.execute("""
        SELECT im.AssetID, im.AssetClass, im.AssetCategory, im.AssetType,
               am.AssetName, am.AssetSymbol
        FROM InvestMapping im
        LEFT JOIN AssetMapping am ON am.AssetId = im.AssetID
        ORDER BY im.AssetClass, im.AssetCategory, im.AssetType
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/manual_investment/preview', methods=['POST'])
def api_manual_investment_preview():
    """Parse uploaded CSV/Excel (columns: investmentmonth, quantity, price) and return rows."""
    import openpyxl, io, csv as _csv
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400

    filename = f.filename.lower()
    rows = []
    try:
        if filename.endswith('.csv'):
            content = f.read().decode('utf-8-sig', errors='replace')
            reader = _csv.DictReader(io.StringIO(content))
            for row in reader:
                rows.append({k.strip().lower(): (v or '').strip() for k, v in row.items()})
        else:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip().lower() if c else '' for c in row]
                else:
                    if all(c is None for c in row):
                        continue
                    rows.append({headers[j]: (str(row[j]).strip() if row[j] is not None else '') for j in range(len(headers))})
    except Exception as e:
        return jsonify({'error': f'Parse error: {e}'}), 400

    # Normalise column names (accept variants)
    _aliases = {
        'investmentmonth': 'month', 'investment month': 'month', 'buy month': 'month',
        'buymonth': 'month', 'date': 'month', 'month': 'month',
        'qty': 'quantity', 'units': 'quantity', 'quantity': 'quantity',
        'price': 'price', 'nav': 'price', 'rate': 'price',
    }
    normalised = []
    for row in rows:
        nr = {}
        for k, v in row.items():
            mapped = _aliases.get(k.lower().strip())
            if mapped:
                nr[mapped] = v
        if nr.get('month') or nr.get('quantity') or nr.get('price'):
            normalised.append(nr)

    if not normalised:
        return jsonify({'error': 'Could not find columns: investmentmonth, quantity, price'}), 400

    return jsonify({'rows': normalised, 'count': len(normalised)})


@app.route('/api/manual_investment/commit', methods=['POST'])
def api_manual_investment_commit():
    """
    Commit manually uploaded investment rows for a given AssetID.
    Body JSON: { asset_id, rows: [{month, quantity, price}, ...] }
    """
    body     = request.get_json(force=True)
    asset_id = (body.get('asset_id') or '').strip()
    rows     = body.get('rows', [])

    if not asset_id:
        return jsonify({'error': 'asset_id is required'}), 400
    if not rows:
        return jsonify({'error': 'No rows to commit'}), 400

    conn = get_db()
    try:
        # Resolve AssetMapping → get asset name, symbol, instrument info
        mapping = conn.execute("""
            SELECT im.AssetID, im.AssetClass, im.AssetCategory, im.AssetType,
                   am.AssetName, am.AssetSymbol
            FROM InvestMapping im
            LEFT JOIN AssetMapping am ON am.AssetId = im.AssetID
            WHERE im.AssetID = ?
        """, (asset_id,)).fetchone()

        if not mapping:
            return jsonify({'error': f'AssetID {asset_id} not found'}), 400

        asset_name   = mapping['AssetName'] or asset_id
        symbol_val   = mapping['AssetSymbol'] or ''
        asset_type   = mapping['AssetType']   or ''
        asset_cat    = mapping['AssetCategory'] or ''
        tx_note      = f'Manual Upload — {asset_name}'

        # Map AssetCategory → transaction category
        _cat_map = {
            'mutual fund': 'Mutual Fund', 'mf': 'Mutual Fund',
            'gold': 'Gold', 'sgb': 'Gold', 'bond': 'Gold',
            'equity': 'Stocks', 'stocks': 'Stocks',
            'etf': 'ETF',
            'fixed return': 'Fixed Return', 'epf': 'Fixed Return', 'ppf': 'Fixed Return',
            'retirement': 'Retirement', 'nps': 'Retirement',
            'real estate': 'Real Estate',
        }
        tx_cat = _cat_map.get(asset_cat.lower()) or _cat_map.get(asset_type.lower()) or asset_cat or 'Investment'
        tx_sub = f'{asset_type} Purchase' if asset_type else 'Purchase'

        inserted = 0
        skipped  = 0

        for r in rows:
            raw_month = str(r.get('month') or '').strip()
            try:
                qty   = float(str(r.get('quantity') or 0).replace(',', ''))
                price = float(str(r.get('price')    or 0).replace(',', ''))
            except (ValueError, TypeError):
                skipped += 1; continue

            if qty <= 0 or price <= 0:
                skipped += 1; continue

            # Normalise month → YYYY-MM
            import re as _re
            mon_str = ''
            # Try "Apr-21" / "Apr-2021" style
            m = _re.match(r'^([A-Za-z]{3})[- ](\d{2,4})$', raw_month)
            if m:
                mo_name, yr = m.group(1).capitalize(), m.group(2)
                yr = ('20' + yr) if len(yr) == 2 else yr
                try:
                    from datetime import datetime as _dt
                    mon_str = _dt.strptime(f'01 {mo_name} {yr}', '%d %b %Y').strftime('%Y-%m')
                except ValueError:
                    pass
            if not mon_str:
                # Try YYYY-MM or YYYY-MM-DD
                m2 = _re.match(r'^(\d{4})-(\d{2})', raw_month)
                if m2:
                    mon_str = f'{m2.group(1)}-{m2.group(2)}'
            if not mon_str:
                skipped += 1; continue

            entry_date   = mon_str + '-01'
            amount       = round(qty * price, 2)
            instrument   = 'bond' if 'bond' in asset_type.lower() else \
                           'mutual_fund' if 'mutual' in asset_cat.lower() else \
                           'etf' if 'etf' in asset_type.lower() else 'stocks'

            # invest_transactions
            conn.execute("""
                INSERT INTO invest_transactions
                    (entry_date, stock_name, asset_type, quantity, action,
                     price, invested_value, month)
                VALUES (?,?,?,?,?,?,?,?)
            """, (entry_date, asset_name, asset_type, qty, 'buy', price, amount, mon_str))
            inserted += 1

            # transactions table — upsert monthly aggregate
            existing = conn.execute("""
                SELECT id FROM transactions
                WHERE type='investment' AND category=? AND date=? AND note=?
            """, (tx_cat, entry_date, tx_note)).fetchone()
            if existing:
                conn.execute("UPDATE transactions SET amount=amount+?, sub_category=? WHERE id=?",
                             (amount, tx_sub, existing[0]))
            else:
                conn.execute("""
                    INSERT INTO transactions (type, category, sub_category, amount, date, note)
                    VALUES ('investment',?,?,?,?,?)
                """, (tx_cat, tx_sub, amount, entry_date, tx_note))

        # Recompute asset portfolio from full invest_transactions history
        assets_updated = 0
        if inserted > 0:
            _, is_new = _recompute_asset_from_transactions(conn, asset_name, symbol_val, instrument)
            assets_updated = 1
            _refresh_monthly_calc(conn)
            _update_portfolio_from_assets(conn)

        conn.commit()
        return jsonify({
            'success': True, 'inserted': inserted, 'skipped': skipped,
            'assets_updated': assets_updated, 'asset_name': asset_name
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


def _normalize_action(trade_type):
    """Normalise broker trade_type strings to 'buy' or 'sell'."""
    t = (trade_type or '').strip().upper()
    if t in ('BUY', 'PURCHASE', 'BUY NEW', 'BUYNEW', 'SWITCH_IN', 'SIP'):
        return 'buy'
    if t in ('SELL', 'REDEEM', 'REDEMPTION', 'SWITCH_OUT', 'REPURCHASE'):
        return 'sell'
    return t.lower() or None


def _instrument_to_asset_type(instrument):
    """Map broker instrument string to invest_transactions asset_type label."""
    m = {
        'stocks':       'Stock',
        'mutual_fund':  'Mutual Fund',
        'etf':          'ETF',
        'bonds':        'Bond',
    }
    return m.get((instrument or '').lower(), instrument or 'Stock')


# Instrument → default InvestMapping.AssetID for new assets created via broker import.
# Stocks → Asset07 (Growth/Equity/Stocks)
# Mutual Fund → Asset08 (Growth/Equity/Mutual Fund)
# ETF → Asset09 (Growth/Equity/ETF)
_INSTRUMENT_INVEST_ASSET_ID = {
    'stocks':      'Asset07',
    'mutual_fund': 'Asset08',
    'etf':         'Asset09',
}


def _get_or_create_asset_mapping(conn, name, symbol, instrument):
    """
    Look up AssetMapping by symbol (preferred) then name.
    If no row found, insert one using the default AssetId for this instrument type.
    Returns MappingID.
    """
    lookup_sym  = (symbol or '').strip().upper()
    lookup_name = (name   or '').strip().upper()

    # 1. Try by symbol
    row = None
    if lookup_sym:
        row = conn.execute(
            "SELECT MappingID FROM AssetMapping WHERE UPPER(TRIM(AssetSymbol))=? LIMIT 1",
            (lookup_sym,)
        ).fetchone()

    # 2. Try by name
    if not row and lookup_name:
        row = conn.execute(
            "SELECT MappingID FROM AssetMapping WHERE UPPER(TRIM(AssetName))=? LIMIT 1",
            (lookup_name,)
        ).fetchone()

    if row:
        return row['MappingID']

    # 3. Not found → create new AssetMapping entry
    invest_asset_id = _INSTRUMENT_INVEST_ASSET_ID.get(
        (instrument or '').lower(), 'Asset07'
    )
    conn.execute(
        "INSERT INTO AssetMapping (AssetName, AssetSymbol, AssetId) VALUES (?,?,?)",
        ((name or symbol or '').strip(), (symbol or '').strip(), invest_asset_id)
    )
    return conn.execute("SELECT last_insert_rowid()").fetchone()[0]


def _get_or_create_asset_row(conn, mapping_id, name, symbol):
    """
    Find the assets row for mapping_id; create it if absent (qty=0, avgprice=0).
    Returns a dict with AssetEntryID, qty, avgprice, investedvalue.
    """
    row = conn.execute(
        "SELECT AssetEntryID, qty, avgprice, investedvalue FROM assets WHERE MappingID=? LIMIT 1",
        (mapping_id,)
    ).fetchone()

    if row:
        # Ensure denormalised columns are filled
        conn.execute(
            "UPDATE assets SET assetname=COALESCE(NULLIF(assetname,''),?), symbol=COALESCE(NULLIF(symbol,''),?) WHERE AssetEntryID=?",
            ((name or '').strip(), (symbol or '').strip(), row['AssetEntryID'])
        )
        return dict(row)

    # Create new asset row
    conn.execute("""
        INSERT INTO assets
            (MappingID, qty, avgprice, ltp, investedvalue, currentvalue,
             pnl, pnlpct, assetname, symbol, updatedat)
        VALUES (?,0,0,0,0,0,0,0,?,?,datetime('now'))
    """, (mapping_id, (name or '').strip(), (symbol or '').strip()))
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return {'AssetEntryID': new_id, 'qty': 0.0, 'avgprice': 0.0, 'investedvalue': 0.0}


def _recompute_asset_from_transactions(conn, stock_name, symbol_val, instrument):
    """
    Recompute qty / avgprice / investedvalue for one asset by aggregating ALL rows
    in invest_transactions that match stock_name (case-insensitive).

    Logic:
      net_qty        = SUM(BUY qty) − SUM(SELL qty)          [clamped to ≥ 0]
      avg_price      = SUM(BUY invested_value) / SUM(BUY qty) [weighted average cost]
      invested_value = avg_price × net_qty                    [remaining cost basis]

    Ensures or creates the matching AssetMapping + assets rows, then writes
    the recomputed values.  Returns a tuple (asset_entry_id, is_new_asset).
    """
    agg = conn.execute("""
        SELECT
            SUM(CASE WHEN UPPER(action) = 'BUY'  THEN quantity       ELSE 0 END) AS buy_qty,
            SUM(CASE WHEN UPPER(action) = 'SELL' THEN quantity       ELSE 0 END) AS sell_qty,
            SUM(CASE WHEN UPPER(action) = 'BUY'  THEN invested_value ELSE 0 END) AS buy_amount
        FROM invest_transactions
        WHERE UPPER(TRIM(stock_name)) = UPPER(TRIM(?))
    """, (stock_name,)).fetchone()

    buy_qty    = float(agg['buy_qty']    or 0)
    sell_qty   = float(agg['sell_qty']   or 0)
    buy_amount = float(agg['buy_amount'] or 0)

    net_qty   = max(0.0, buy_qty - sell_qty)
    avg_price = buy_amount / buy_qty if buy_qty > 0 else 0.0
    inv_value = round(avg_price * net_qty, 6)

    mapping_id = _get_or_create_asset_mapping(conn, stock_name, symbol_val, instrument)
    asset_row  = _get_or_create_asset_row(conn, mapping_id, stock_name, symbol_val)
    is_new     = asset_row['qty'] == 0 and asset_row['avgprice'] == 0

    conn.execute("""
        UPDATE assets
        SET qty=?, avgprice=?, investedvalue=?, currentvalue=?, updatedat=datetime('now')
        WHERE AssetEntryID=?
    """, (net_qty, avg_price, inv_value, inv_value, asset_row['AssetEntryID']))

    return asset_row['AssetEntryID'], is_new


@app.route('/api/broker_uploads/<source_type>/commit', methods=['POST'])
def api_broker_upload_commit(source_type):
    """
    Commit staged broker rows →
      1. Insert every valid row into invest_transactions
      2. After all inserts, recompute net qty / avg_price / investedvalue for each
         unique stock from the FULL invest_transactions history (BUY − SELL)
      3. Find or create AssetMapping + assets rows as needed
      4. Clear staging tables on success
    """
    conn = get_db()
    try:
        # ── Fetch staged rows ────────────────────────────────────────────────
        rows = conn.execute("""
            SELECT d.name, d.symbol, d.isin, d.trade_type, d.trade_date,
                   d.quantity, d.price, d.amount, d.exchange, d.order_id,
                   d.status, d.instrument
            FROM raw_upload_data d
            JOIN raw_upload_meta m ON m.id = d.upload_id
            WHERE d.source_type = ?
            ORDER BY d.trade_date, d.id
        """, (source_type,)).fetchall()

        if not rows:
            return jsonify({'error': 'No staged data found. Upload a file first.'}), 400

        inserted   = 0
        skipped    = 0

        # Track unique stocks seen in this batch (name → {symbol, instrument})
        affected = {}

        for row in rows:
            action = _normalize_action(row['trade_type'])
            if action not in ('buy', 'sell'):
                skipped += 1
                continue

            qty    = float(row['quantity'] or 0)
            price  = float(row['price']    or 0)
            amount = float(row['amount']   or 0)
            if qty <= 0:
                skipped += 1
                continue

            if price == 0 and qty > 0 and amount > 0:
                price = amount / qty

            raw_date   = (row['trade_date'] or '').strip()
            entry_date = _normalise_date(raw_date)
            month_str  = entry_date[:7] if len(entry_date) >= 7 else ''

            name       = (row['name']   or row['symbol'] or '').strip()
            symbol_val = (row['symbol'] or '').strip()
            instrument = (row['instrument'] or 'stocks').lower()
            asset_type = _instrument_to_asset_type(instrument)

            # ── 1. Insert into invest_transactions ───────────────────────────
            conn.execute("""
                INSERT INTO invest_transactions
                    (entry_date, stock_name, asset_type, quantity, action,
                     price, invested_value, month)
                VALUES (?,?,?,?,?,?,?,?)
            """, (entry_date, name, asset_type, qty, action, price, amount, month_str))
            inserted += 1

            # Track for recompute pass
            key = name.upper()
            if key not in affected:
                affected[key] = {'name': name, 'symbol': symbol_val, 'instrument': instrument}

        # ── 2. Recompute qty/avg/invested for each unique stock from full history ─
        assets_created = 0
        assets_updated = 0
        for key, info in affected.items():
            _, is_new = _recompute_asset_from_transactions(
                conn, info['name'], info['symbol'], info['instrument']
            )
            if is_new:
                assets_created += 1
            else:
                assets_updated += 1

        # ── 3. Refresh derived tables ─────────────────────────────────────────
        _refresh_monthly_calc(conn)
        _update_portfolio_from_assets(conn)   # sync portfolio aggregates

        # ── 3b. Upsert monthly investment rows into transactions table ─────────
        # Aggregate BUY amounts per (month, category) and upsert into transactions.
        # For MF → category='Mutual Fund'; for stocks → resolve per-row from
        # InvestMapping (ETF stays 'ETF', stocks become 'Stocks').
        _cfg        = _BROKER_PARSERS.get(source_type, {})
        _instrument = _cfg.get('instrument', '')

        if _instrument in ('mutual_fund', 'etf', 'stocks'):
            _tx_note = f'Imported from {source_type.replace("_", " ").title()}'

            # Build a name→asset_type lookup from InvestMapping for stock rows
            _name_to_atype = {}
            if _instrument == 'stocks':
                _im_rows = conn.execute("""
                    SELECT am.AssetName, im.AssetType
                    FROM AssetMapping am
                    JOIN InvestMapping im ON am.AssetId = im.AssetID
                """).fetchall()
                for _im in _im_rows:
                    _name_to_atype[(_im['AssetName'] or '').strip().upper()] = (_im['AssetType'] or '').strip()

            # Category/sub helpers
            def _resolve_cat(name, instrument):
                if instrument == 'mutual_fund':
                    return 'Mutual Fund', 'MF SIP'
                if instrument == 'etf':
                    return 'ETF', 'ETF Purchase'
                # stocks: check actual asset type from mapping
                atype = _name_to_atype.get((name or '').strip().upper(), '').lower()
                if 'etf' in atype:
                    return 'ETF', 'ETF Purchase'
                return 'Stocks', 'Stock Purchase'

            # Aggregate total BUY amount per (month, category)
            _monthly = {}  # key: (month, category, sub) → total
            for _row in rows:
                _action = _normalize_action(_row['trade_type'] or '')
                if _action != 'buy':
                    continue
                _amt = float(_row['amount'] or 0)
                if _amt <= 0:
                    _qty = float(_row['quantity'] or 0)
                    _prc = float(_row['price'] or 0)
                    _amt = _qty * _prc
                if _amt <= 0:
                    continue
                _rd  = _normalise_date((_row['trade_date'] or '').strip())
                _mon = _rd[:7] if len(_rd) >= 7 else ''
                if not _mon:
                    continue
                _cat, _sub = _resolve_cat(_row['name'], _instrument)
                _key = (_mon, _cat, _sub)
                _monthly[_key] = _monthly.get(_key, 0) + _amt

            for (_mon, _cat, _sub), _total in _monthly.items():
                _tx_date = _mon + '-01'
                _existing = conn.execute("""
                    SELECT id FROM transactions
                    WHERE type='investment' AND category=? AND date=? AND note=?
                """, (_cat, _tx_date, _tx_note)).fetchone()
                if _existing:
                    conn.execute("""
                        UPDATE transactions SET amount=?, sub_category=?
                        WHERE id=?
                    """, (round(_total, 2), _sub, _existing[0]))
                else:
                    conn.execute("""
                        INSERT INTO transactions (type, category, sub_category, amount, date, note)
                        VALUES ('investment', ?, ?, ?, ?, ?)
                    """, (_cat, _sub, round(_total, 2), _tx_date, _tx_note))

        # ── 4. Clear staging data ─────────────────────────────────────────────
        old_ids = [r[0] for r in conn.execute(
            "SELECT id FROM raw_upload_meta WHERE source_type=?", (source_type,)).fetchall()]
        if old_ids:
            ph = ','.join('?' * len(old_ids))
            conn.execute(f"DELETE FROM raw_upload_data WHERE upload_id IN ({ph})", old_ids)
            conn.execute(f"DELETE FROM raw_upload_meta WHERE id IN ({ph})", old_ids)

        conn.commit()
        return jsonify({
            'success':        True,
            'inserted':       inserted,
            'assets_updated': assets_updated,
            'assets_created': assets_created,
            'skipped':        skipped,
        })

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/reset/<scope>', methods=['POST'])
def api_reset_data(scope):
    """
    Permanently delete records for the requested scope.
    Returns {deleted: {table: row_count}} for every table touched.

    Scopes and what they delete:
      transactions       → transactions
      invest_transactions→ invest_transactions, monthly_investment_calc
      assets             → assets, portfolio, raw_upload_data, raw_upload_meta
      loans              → loans, loan_master
      broker_staging     → raw_upload_data, raw_upload_meta
      nse_master         → nse_master
      all                → all of the above (preserves InvestMapping, AssetMapping,
                           alerts, wealth, wealth goals/dates, magnet_status,
                           um_vision_cards)
      asset_mapping      → AssetMapping, assets, portfolio
      factory            → EVERY table (full blank slate)
    """
    # Define table lists per scope
    # Each entry: (table_name, optional_where_clause)
    _SCOPE_TABLES = {
        'transactions':       ['transactions'],
        'invest_transactions':['invest_transactions', 'monthly_investment_calc'],
        'assets':             ['assets', 'portfolio', 'raw_upload_data', 'raw_upload_meta'],
        'loans':              ['loans', 'loan_master'],
        'broker_staging':     ['raw_upload_data', 'raw_upload_meta'],
        'nse_master':         ['nse_master'],
        'all': [
            'transactions', 'invest_transactions', 'monthly_investment_calc',
            'assets', 'portfolio',
            'loans', 'loan_master',
            'raw_upload_data', 'raw_upload_meta',
            'nse_master',
        ],
        'asset_mapping': ['assets', 'portfolio', 'AssetMapping'],
        # Wizard reset scopes (module-level)
        'wizard_vision': ['um_vision_cards', 'magnet_status'],
        'wizard_wealth': [
            'assets', 'portfolio', 'AssetMapping',
            'invest_transactions', 'monthly_investment_calc',
            'raw_upload_data', 'raw_upload_meta',
            'nse_master',
            'demat_wallet', 'stock_holdings', 'stock_transactions',
            'stock_dividends', 'stock_pnl',
            'wealth',
        ],
        'wizard_money': ['transactions', 'loans', 'loan_master', 'alerts'],
        'wizard_docs':  ['doc_wallet', 'doc_attachments', 'doc_wallet_types'],
        'factory': [
            # Money Tracker
            'transactions', 'loans', 'loan_master', 'alerts',
            # Wealth Engine
            'invest_transactions', 'monthly_investment_calc',
            'assets', 'portfolio', 'AssetMapping',
            'raw_upload_data', 'raw_upload_meta',
            'nse_master',
            'demat_wallet', 'stock_holdings', 'stock_transactions',
            'stock_dividends', 'stock_pnl',
            'wealth',
            # Vision & Magnet
            'um_vision_cards', 'magnet_status',
            # Docs Wallet
            'doc_wallet', 'doc_attachments', 'doc_wallet_types',
            # Onboarding state (but NOT app_password / theme / currency)
            # app_settings is handled separately below
            # InvestMapping intentionally excluded — system reference table
        ],
    }

    tables = _SCOPE_TABLES.get(scope)
    if not tables:
        return jsonify({'error': f'Unknown reset scope: {scope}'}), 400

    conn = get_db()
    deleted = {}
    try:
        # Disable FK constraints temporarily so order doesn't matter
        conn.execute("PRAGMA foreign_keys = OFF")

        for tbl in tables:
            try:
                count_before = conn.execute(f'SELECT COUNT(*) FROM [{tbl}]').fetchone()[0]
                conn.execute(f'DELETE FROM [{tbl}]')
                deleted[tbl] = count_before
            except Exception as tbl_err:
                # Table may not exist in older DBs — skip gracefully
                deleted[tbl] = f'error: {tbl_err}'

        # Re-enable FK constraints
        conn.execute("PRAGMA foreign_keys = ON")

        # After factory reset: clear user app_settings + re-seed all default data
        if scope == 'factory':
            conn.execute("""
                DELETE FROM app_settings
                WHERE key NOT IN ('app_password','theme','currency','user_name')
            """)
            deleted['app_settings (user keys)'] = 'cleared'
            _restore_seed_data(conn)

        conn.commit()
        return jsonify({'success': True, 'scope': scope, 'deleted': deleted})

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.close()


@app.route('/api/assets/recompute_from_transactions', methods=['POST'])
def api_recompute_assets_from_transactions():
    """
    For every unique stock_name in invest_transactions, recompute the net qty,
    avg_price and investedvalue in the assets table (BUY increases, SELL decreases).
    Safe to call at any time — fully idempotent.
    """
    conn = get_db()
    try:
        stocks = conn.execute("""
            SELECT DISTINCT stock_name, asset_type
            FROM invest_transactions
            WHERE stock_name IS NOT NULL AND stock_name != ''
        """).fetchall()

        updated = 0
        created = 0
        for s in stocks:
            stock_name = s['stock_name']
            asset_type = (s['asset_type'] or 'Stock')
            # Map asset_type label back to instrument key for AssetMapping lookup
            instrument = 'mutual_fund' if 'mutual' in asset_type.lower() \
                         else 'etf'    if 'etf'    in asset_type.lower() \
                         else 'stocks'
            # Try to find existing symbol from AssetMapping
            am = conn.execute("""
                SELECT am.AssetSymbol FROM AssetMapping am
                WHERE UPPER(TRIM(am.AssetName)) = UPPER(TRIM(?)) LIMIT 1
            """, (stock_name,)).fetchone()
            symbol_val = am['AssetSymbol'] if am else ''

            _, is_new = _recompute_asset_from_transactions(
                conn, stock_name, symbol_val, instrument
            )
            if is_new:
                created += 1
            else:
                updated += 1

        _refresh_monthly_calc(conn)
        _update_portfolio_from_assets(conn)   # sync portfolio aggregates
        conn.commit()
        return jsonify({'success': True, 'updated': updated, 'created': created})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# DEMAT PORTFOLIO — FIFO ENGINE + CHARGE CALCULATOR
# ─────────────────────────────────────────────────────────────────────────────

def _zerodha_charges(action, qty, price):
    """Estimate Zerodha equity delivery charges (INR)."""
    tv = qty * price
    stt          = round(tv * 0.001, 2)          # 0.1% buy & sell
    exchange_chg = round(tv * 0.0000345, 2)
    gst          = round(exchange_chg * 0.18, 2)
    sebi         = round(tv * 0.000001, 2)
    stamp        = round(tv * 0.00015, 2) if action.upper() == 'BUY' else 0
    total        = stt + exchange_chg + gst + sebi + stamp
    return {'stt': stt, 'exchange': exchange_chg, 'gst': gst,
            'sebi': sebi, 'stamp': stamp, 'total': round(total, 2)}


def _fifo_sell(conn, symbol, trade_type, qty_to_sell, sell_price, sell_date, sell_txn_id):
    """Consume FIFO lots for symbol+trade_type.
    Returns (consumed_list, unfilled_qty).  consumed_list contains one dict per lot used."""
    lots = conn.execute("""
        SELECT id, buy_date, qty_remaining, buy_price
        FROM stock_holdings
        WHERE UPPER(symbol)=UPPER(?) AND trade_type=? AND qty_remaining > 0
        ORDER BY buy_date ASC
    """, (symbol, trade_type)).fetchall()

    remaining = qty_to_sell
    consumed  = []
    for lot in lots:
        if remaining <= 0:
            break
        use     = min(remaining, lot['qty_remaining'])
        new_rem = round(lot['qty_remaining'] - use, 6)
        conn.execute("UPDATE stock_holdings SET qty_remaining=? WHERE id=?", (new_rem, lot['id']))
        try:
            d1 = datetime.strptime(sell_date[:10], '%Y-%m-%d')
            d2 = datetime.strptime(lot['buy_date'][:10], '%Y-%m-%d')
            holding_days = max(0, (d1 - d2).days)
        except Exception:
            holding_days = 0
        gross_pnl = round((sell_price - lot['buy_price']) * use, 2)
        consumed.append({
            'lot_id': lot['id'], 'qty': use,
            'buy_price': lot['buy_price'], 'buy_date': lot['buy_date'],
            'gross_pnl': gross_pnl, 'holding_days': holding_days,
            'tax_category': 'LTCG' if holding_days >= 365 else 'STCG'
        })
        remaining = round(remaining - use, 6)
    return consumed, remaining


def _demat_wallet_balance(conn):
    """Compute running demat cash balance from ledger."""
    row = conn.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN type IN ('DEPOSIT','SELL','DIVIDEND') THEN amount ELSE 0 END), 0) -
            COALESCE(SUM(CASE WHEN type IN ('WITHDRAW','BUY','CHARGE') THEN amount ELSE 0 END), 0)
        AS balance FROM demat_wallet
    """).fetchone()
    return round(float(row[0] or 0), 2)


# ─────────────────────────────────────────────────────────────────────────────
# DEMAT API ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/demat/onboarding_status')
def api_demat_onboarding_status():
    conn = get_db()
    row  = conn.execute("SELECT value FROM app_settings WHERE key='demat_onboarded'").fetchone()
    conn.close()
    return jsonify({'onboarded': row and row['value'] == '1'})


@app.route('/api/demat/onboard', methods=['POST'])
def api_demat_onboard():
    """One-time seeding: existing holdings + starting wallet balance."""
    d    = request.get_json(force=True)
    conn = get_db()
    try:
        # Wallet seed
        wallet_balance = float(d.get('wallet_balance') or 0)
        invested_capital = float(d.get('invested_capital') or 0)
        onboard_date = (d.get('onboard_date') or datetime.now().strftime('%Y-%m-%d'))[:10]

        # Store invested capital in settings (not as wallet entry) to avoid double-counting
        if invested_capital > 0:
            conn.execute(
                "INSERT OR REPLACE INTO app_settings (key, value) VALUES ('demat_invested_capital', ?)",
                (str(invested_capital),)
            )
        # Only actual cash in broker account goes into wallet ledger
        if wallet_balance > 0:
            conn.execute(
                "INSERT INTO demat_wallet (txn_date, type, amount, note) VALUES (?,?,?,?)",
                (onboard_date, 'DEPOSIT', wallet_balance, 'Onboarding: current demat cash balance')
            )

        # Holdings seed
        holdings = d.get('holdings') or []
        for h in holdings:
            symbol     = (h.get('symbol') or '').strip().upper()
            isin       = (h.get('isin')   or '').strip()
            trade_type = (h.get('trade_type') or 'LONG').upper()
            qty        = float(h.get('qty') or 0)
            avg_price  = float(h.get('avg_price') or 0)
            buy_date   = (h.get('buy_date') or onboard_date)[:10]
            if not symbol or qty <= 0:
                continue
            conn.execute("""
                INSERT INTO stock_holdings
                    (symbol, isin, trade_type, buy_date, qty_original, qty_remaining, buy_price, lot_source)
                VALUES (?,?,?,?,?,?,?,'ONBOARDING')
            """, (symbol, isin, trade_type, buy_date, qty, qty, avg_price))
            # Also log as a synthetic BUY transaction
            conn.execute("""
                INSERT INTO stock_transactions
                    (trade_date, symbol, isin, action, qty, price, trade_type, net_amount, lot_source)
                VALUES (?,?,?,'BUY',?,?,?,?,'ONBOARDING')
            """, (buy_date, symbol, isin, qty, avg_price, trade_type, round(qty * avg_price, 2)))
            # Mirror into invest_transactions so Asset07 stays in sync
            conn.execute("""
                INSERT INTO invest_transactions
                    (entry_date, stock_name, asset_type, action, quantity, price, invested_value)
                VALUES (?,?,?,?,?,?,?)
            """, (buy_date, symbol, 'Stock', 'BUY', qty, avg_price,
                  round(qty * avg_price, 2)))

        conn.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('demat_onboarded','1')")
        conn.commit()
        # Sync holdings → Wealth Engine assets table
        _sync_demat_to_wealth(conn)
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/demat/wallet')
def api_demat_wallet():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, txn_date, type, amount, note FROM demat_wallet ORDER BY txn_date DESC, id DESC LIMIT 100"
    ).fetchall()
    balance = _demat_wallet_balance(conn)
    conn.close()
    return jsonify({'balance': balance, 'entries': [dict(r) for r in rows]})


@app.route('/api/demat/wallet', methods=['POST'])
def api_demat_wallet_add():
    d      = request.get_json(force=True)
    conn   = get_db()
    wtype  = (d.get('type') or '').upper()
    amount = float(d.get('amount') or 0)
    date   = (d.get('date') or datetime.now().strftime('%Y-%m-%d'))[:10]
    note   = d.get('note', '')
    if wtype not in ('DEPOSIT', 'WITHDRAW') or amount <= 0:
        conn.close()
        return jsonify({'error': 'type must be DEPOSIT or WITHDRAW, amount > 0'}), 400

    # Mirror to transactions table for P&L / cash-flow tracking
    if wtype == 'DEPOSIT':
        conn.execute(
            "INSERT INTO transactions (type, category, sub_category, amount, date, note) VALUES (?,?,?,?,?,?)",
            ('investment', 'Stocks', 'Demat Deposit', amount, date, note or 'Zerodha Pay-in')
        )
        # Increment invested capital
        ic_row = conn.execute("SELECT value FROM app_settings WHERE key='demat_invested_capital'").fetchone()
        new_ic = round(float(ic_row['value'] if ic_row else 0) + amount, 2)
        conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('demat_invested_capital',?)", (str(new_ic),))
    else:
        conn.execute(
            "INSERT INTO transactions (type, category, sub_category, amount, date, note) VALUES (?,?,?,?,?,?)",
            ('income', 'Investment Return', 'Demat Withdrawal', amount, date, note or 'Zerodha Pay-out')
        )
        # Decrement invested capital
        ic_row = conn.execute("SELECT value FROM app_settings WHERE key='demat_invested_capital'").fetchone()
        new_ic = round(max(0, float(ic_row['value'] if ic_row else 0) - amount), 2)
        conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('demat_invested_capital',?)", (str(new_ic),))
    conn.commit()
    ref_txn_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    conn.execute(
        "INSERT INTO demat_wallet (txn_date, type, amount, note, ref_txn_id) VALUES (?,?,?,?,?)",
        (date, wtype, amount, note, ref_txn_id)
    )
    conn.commit()
    balance = _demat_wallet_balance(conn)
    # Keep portfolio Asset07 InvestedValue in sync with updated pay-in total
    _sync_demat_to_wealth(conn)
    conn.close()
    return jsonify({'success': True, 'balance': balance})


def _it_equity_symbols(conn):
    """Return set of stock symbols already covered by stock_holdings (qty > 0)."""
    rows = conn.execute(
        "SELECT UPPER(TRIM(symbol)) AS s FROM stock_holdings WHERE qty_remaining > 0"
    ).fetchall()
    return {r['s'] for r in rows}

def _it_stocks_as_holdings(conn, exclude_symbols=None):
    """
    Return Wealth Engine Asset07 stock rows (assets + AssetMapping) as a holdings-like list,
    supplemented by invest_transactions(Stock) for any symbol not yet in assets.
    Each item: symbol, trade_type='LONG', qty_remaining, buy_price, current_price, company_name.
    Excludes symbols already in stock_holdings (to avoid double-counting).
    """
    ex = exclude_symbols or set()

    # Primary source: assets table linked to AssetMapping where AssetId='Asset07'
    rows = conn.execute("""
        SELECT UPPER(TRIM(COALESCE(a.symbol, am.AssetSymbol, am.AssetName))) AS symbol,
               am.AssetName         AS company_name,
               a.qty                AS qty,
               COALESCE(a.avgprice, CASE WHEN a.qty > 0 THEN a.investedvalue / a.qty ELSE 0 END) AS avg_price,
               COALESCE(a.ltp, 0)   AS ltp,
               a.investedvalue      AS invested_value
        FROM assets a
        JOIN AssetMapping am ON a.MappingID = am.MappingID
        WHERE am.AssetId = 'Asset07'
          AND a.qty > 0
          AND TRIM(COALESCE(a.symbol, am.AssetSymbol, '')) != ''
        ORDER BY am.AssetSymbol
    """).fetchall()

    seen = set()
    result = []
    for r in rows:
        sym = r['symbol']
        if sym in ex or sym in seen:
            continue
        seen.add(sym)
        qty = float(r['qty'] or 0)
        avg_p = float(r['avg_price'] or 0)
        result.append({
            'symbol': sym,
            'isin': '',
            'trade_type': 'LONG',
            'buy_date': '',
            'qty_original': qty,
            'qty_remaining': qty,
            'buy_price': avg_p,
            'lot_source': 'WEALTH_ENGINE',
            'current_price': float(r['ltp'] or 0),
            'company_name': r['company_name'] or sym,
        })

    # Supplement with invest_transactions(Stock) for any symbol not yet captured above
    wce_syms = {r['symbol'] for r in result}
    try:
        it_rows = conn.execute("""
            SELECT UPPER(TRIM(stock_name)) AS symbol,
                   SUM(CASE WHEN LOWER(action)='buy'  THEN quantity ELSE 0 END)
                 - SUM(CASE WHEN LOWER(action)='sell' THEN quantity ELSE 0 END) AS net_qty,
                   SUM(CASE WHEN LOWER(action)='buy'  THEN invested_value ELSE 0 END)
                 / NULLIF(SUM(CASE WHEN LOWER(action)='buy' THEN quantity ELSE 0 END), 0) AS avg_buy_price,
                   COALESCE(n.ltp, 0) AS ltp
            FROM invest_transactions it
            LEFT JOIN nse_master n ON UPPER(TRIM(n.symbol)) = UPPER(TRIM(it.stock_name))
            WHERE it.asset_type = 'Stock'
              AND it.stock_name IS NOT NULL AND TRIM(it.stock_name) != ''
            GROUP BY UPPER(TRIM(it.stock_name))
            HAVING net_qty > 0
        """).fetchall()
        for r in it_rows:
            sym = r['symbol']
            if sym in ex or sym in seen or sym in wce_syms:
                continue
            seen.add(sym)
            result.append({
                'symbol': sym, 'isin': '', 'trade_type': 'LONG', 'buy_date': '',
                'qty_original': float(r['net_qty'] or 0),
                'qty_remaining': float(r['net_qty'] or 0),
                'buy_price': float(r['avg_buy_price'] or 0),
                'lot_source': 'INVEST_TX',
                'current_price': float(r['ltp'] or 0),
                'company_name': sym,
            })
    except Exception:
        pass

    return result


@app.route('/api/demat/holdings')
def api_demat_holdings():
    """Open lots grouped by symbol+trade_type. Merges stock_holdings + invest_transactions(Stock)."""
    conn = get_db()
    sh_symbols = _it_equity_symbols(conn)

    lots = conn.execute("""
        SELECT h.id, h.symbol, h.isin, h.trade_type, h.buy_date,
               h.qty_original, h.qty_remaining, h.buy_price, h.lot_source,
               COALESCE(n.ltp, 0) AS current_price,
               COALESCE(n.company_name, h.symbol) AS company_name
        FROM stock_holdings h
        LEFT JOIN nse_master n ON UPPER(n.symbol) = UPPER(h.symbol)
        WHERE h.qty_remaining > 0
        ORDER BY h.symbol, h.trade_type, h.buy_date
    """).fetchall()

    # Group stock_holdings by symbol+trade_type
    groups = {}
    for lot in lots:
        key = f"{lot['symbol'].upper()}|{lot['trade_type']}"
        if key not in groups:
            groups[key] = {
                'symbol': lot['symbol'].upper(),
                'company_name': lot['company_name'],
                'trade_type': lot['trade_type'],
                'current_price': float(lot['current_price']),
                'total_qty': 0, 'total_invested': 0, 'lots': []
            }
        g = groups[key]
        qty = float(lot['qty_remaining'])
        g['total_qty']      = round(g['total_qty'] + qty, 6)
        g['total_invested'] = round(g['total_invested'] + qty * lot['buy_price'], 2)
        g['lots'].append({
            'id': lot['id'], 'buy_date': lot['buy_date'],
            'qty_remaining': qty, 'buy_price': float(lot['buy_price']),
            'lot_source': lot['lot_source']
        })

    result = []
    for g in groups.values():
        qty = g['total_qty']
        invested = g['total_invested']
        avg_price = r2(invested / qty) if qty > 0 else 0
        cp = g['current_price']
        current_val = round(qty * cp, 2)
        pnl = round(current_val - invested, 2)
        pnl_pct = round(pnl / invested * 100, 2) if invested > 0 else 0
        result.append({
            'symbol': g['symbol'], 'company_name': g['company_name'],
            'trade_type': g['trade_type'],
            'total_qty': qty, 'avg_price': avg_price,
            'total_invested': invested, 'current_price': cp,
            'current_value': current_val, 'unrealized_pnl': pnl, 'pnl_pct': pnl_pct,
            'lots': g['lots']
        })

    # Append invest_transactions stocks not already in stock_holdings
    for it in _it_stocks_as_holdings(conn, exclude_symbols=sh_symbols):
        qty = it['qty_remaining']
        invested = round(qty * it['buy_price'], 2)
        cp = it['current_price']
        current_val = round(qty * cp, 2) if cp > 0 else invested
        pnl = round(current_val - invested, 2)
        pnl_pct = round(pnl / invested * 100, 2) if invested > 0 else 0
        result.append({
            'symbol': it['symbol'], 'company_name': it['company_name'],
            'trade_type': it['trade_type'],
            'total_qty': qty, 'avg_price': it['buy_price'],
            'total_invested': invested, 'current_price': cp,
            'current_value': current_val, 'unrealized_pnl': pnl, 'pnl_pct': pnl_pct,
            'lots': [{'id': None, 'buy_date': '', 'qty_remaining': qty,
                      'buy_price': it['buy_price'], 'lot_source': 'INVEST_TX'}]
        })

    conn.close()
    return jsonify(result)


@app.route('/api/demat/portfolio_summary')
def api_demat_portfolio_summary():
    conn = get_db()

    # One-time migration: old onboarding stored invested_capital as a DEPOSIT in demat_wallet
    # with note 'Onboarding: historical capital baseline'. Move it to app_settings and remove
    # that wallet entry so it doesn't inflate the cash balance.
    ic_row = conn.execute("SELECT value FROM app_settings WHERE key='demat_invested_capital'").fetchone()
    if not ic_row:
        old_entries = conn.execute(
            "SELECT id, amount FROM demat_wallet WHERE type='DEPOSIT' AND note LIKE '%historical capital%'"
        ).fetchall()
        if old_entries:
            migrated_capital = round(sum(float(e['amount']) for e in old_entries), 2)
            conn.execute(
                "INSERT OR REPLACE INTO app_settings (key,value) VALUES ('demat_invested_capital',?)",
                (str(migrated_capital),)
            )
            # Remove these entries so they don't count toward cash balance
            for e in old_entries:
                conn.execute("DELETE FROM demat_wallet WHERE id=?", (e['id'],))
            conn.commit()
            ic_row = conn.execute("SELECT value FROM app_settings WHERE key='demat_invested_capital'").fetchone()

    invested_capital = round(float(ic_row['value']) if ic_row else 0, 2)
    cash_balance = _demat_wallet_balance(conn)

    # Holdings per trade_type from stock_holdings
    sh_symbols = _it_equity_symbols(conn)
    lots = conn.execute("""
        SELECT h.symbol, h.trade_type, h.qty_remaining, h.buy_price,
               COALESCE(n.ltp, 0) AS ltp
        FROM stock_holdings h
        LEFT JOIN nse_master n ON UPPER(n.symbol) = UPPER(h.symbol)
        WHERE h.qty_remaining > 0
    """).fetchall()
    # Add invest_transactions stocks (all treated as LONG)
    it_lots = [
        {'trade_type': it['trade_type'], 'qty_remaining': it['qty_remaining'],
         'buy_price': it['buy_price'], 'ltp': it['current_price']}
        for it in _it_stocks_as_holdings(conn, exclude_symbols=sh_symbols)
    ]
    all_lots = list(lots) + it_lots

    def _agg(tt):
        rows = [l for l in all_lots if l['trade_type'] == tt]
        inv = round(sum(float(l['qty_remaining']) * float(l['buy_price']) for l in rows), 2)
        ltp_sum = round(sum(float(l['qty_remaining']) * float(l['ltp']) for l in rows if float(l['ltp']) > 0), 2)
        # For rows with no LTP, fall back to invested value
        no_ltp = round(sum(float(l['qty_remaining']) * float(l['buy_price']) for l in rows if float(l['ltp']) == 0), 2)
        cur = round(ltp_sum + no_ltp, 2)
        return inv, cur

    long_inv,  long_cur  = _agg('LONG')
    swing_inv, swing_cur = _agg('SWING')
    holdings_value   = round(long_cur + swing_cur, 2)
    current_portfolio = round(cash_balance + holdings_value, 2)
    total_invested_holdings = round(long_inv + swing_inv, 2)
    unrealized_pnl = round(holdings_value - total_invested_holdings, 2)

    long_unreal  = round(long_cur  - long_inv,  2)
    swing_unreal = round(swing_cur - swing_inv, 2)
    long_unreal_pct  = round(long_unreal  / long_inv  * 100, 2) if long_inv  > 0 else 0
    swing_unreal_pct = round(swing_unreal / swing_inv * 100, 2) if swing_inv > 0 else 0

    # Realized P&L per trade_type
    pnl_rows = conn.execute("""
        SELECT trade_type, COALESCE(SUM(net_pnl),0) AS pnl FROM stock_pnl GROUP BY trade_type
    """).fetchall()
    realized = {r['trade_type']: round(float(r['pnl']), 2) for r in pnl_rows}

    conn.close()
    return jsonify({
        'invested_capital':    invested_capital,
        'cash_balance':        cash_balance,
        'holdings_value':      holdings_value,
        'current_portfolio':   current_portfolio,
        'unrealized_pnl':      unrealized_pnl,
        # Long term breakdown
        'long_invested':       long_inv,
        'long_current':        long_cur,
        'long_unrealized':     long_unreal,
        'long_unrealized_pct': long_unreal_pct,
        'long_realized':       realized.get('LONG', 0),
        # Swing breakdown
        'swing_invested':       swing_inv,
        'swing_current':        swing_cur,
        'swing_unrealized':     swing_unreal,
        'swing_unrealized_pct': swing_unreal_pct,
        'swing_realized':       realized.get('SWING', 0),
        # Legacy keys
        'realized_pnl_swing': realized.get('SWING', 0),
        'realized_pnl_long':  realized.get('LONG', 0),
    })


@app.route('/api/demat/pnl')
def api_demat_pnl():
    trade_type = request.args.get('trade_type', '')
    conn = get_db()
    q = "SELECT * FROM stock_pnl"
    p = []
    if trade_type:
        q += " WHERE trade_type=?"; p.append(trade_type.upper())
    q += " ORDER BY sell_date DESC LIMIT 200"
    rows = conn.execute(q, p).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


def _sync_demat_to_wealth(conn):
    """Write stock_holdings + invest_transactions(Stock) totals into assets/portfolio for Wealth Engine.
    Returns list of synced symbols.
    """
    sh_symbols = _it_equity_symbols(conn)

    # From stock_holdings
    sh_rows = conn.execute("""
        SELECT h.symbol,
               SUM(h.qty_remaining)              AS total_qty,
               SUM(h.qty_remaining * h.buy_price) AS total_invested,
               COALESCE(n.ltp, 0)                AS ltp
        FROM stock_holdings h
        LEFT JOIN nse_master n ON UPPER(n.symbol) = UPPER(h.symbol)
        WHERE h.qty_remaining > 0
        GROUP BY h.symbol
    """).fetchall()

    # Build combined list: stock_holdings first, then invest_transactions extras
    it_rows = _it_stocks_as_holdings(conn, exclude_symbols=sh_symbols)
    holdings_combined = [
        {'symbol': r['symbol'], 'total_qty': r['total_qty'],
         'total_invested': r['total_invested'], 'ltp': r['ltp']}
        for r in sh_rows
    ] + [
        {'symbol': it['symbol'], 'total_qty': it['qty_remaining'],
         'total_invested': it['qty_remaining'] * it['buy_price'], 'ltp': it['current_price']}
        for it in it_rows
    ]

    synced = []
    for h in holdings_combined:
        sym       = h['symbol'].strip().upper()
        total_qty = round(float(h['total_qty']), 6)
        invested  = round(float(h['total_invested']), 2)
        ltp       = float(h['ltp'])
        avg_price = r2(invested / total_qty) if total_qty > 0 else 0
        cur_value = round(total_qty * ltp, 2) if ltp > 0 else invested

        mapping_row = conn.execute(
            "SELECT MappingID FROM AssetMapping WHERE UPPER(TRIM(AssetSymbol))=? LIMIT 1", (sym,)
        ).fetchone()
        if mapping_row:
            mapping_id = mapping_row['MappingID']
            conn.execute("UPDATE AssetMapping SET AssetId='Asset07' WHERE MappingID=?", (mapping_id,))
        else:
            conn.execute(
                "INSERT INTO AssetMapping (AssetName, AssetSymbol, AssetId) VALUES (?,?,'Asset07')", (sym, sym)
            )
            mapping_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        existing = conn.execute("SELECT AssetEntryID FROM assets WHERE MappingID=? LIMIT 1", (mapping_id,)).fetchone()
        if existing:
            conn.execute("""
                UPDATE assets SET qty=?, avgprice=?, investedvalue=?, currentvalue=?,
                    ltp=?, assetname=?, symbol=?, updatedat=datetime('now')
                WHERE AssetEntryID=?
            """, (total_qty, avg_price, invested, cur_value, ltp, sym, sym, existing['AssetEntryID']))
        else:
            conn.execute("""
                INSERT INTO assets (MappingID, qty, avgprice, ltp, investedvalue, currentvalue,
                                    pnl, pnlpct, assetname, symbol, updatedat)
                VALUES (?,?,?,?,?,?,0,0,?,?,datetime('now'))
            """, (mapping_id, total_qty, avg_price, ltp, invested, cur_value, sym, sym))
        synced.append(sym)

    # Purge corrupt Asset07 rows: no symbol + not a real stock holding
    if synced:
        placeholders = ','.join('?' * len(synced))
        conn.execute(f"""
            DELETE FROM assets WHERE AssetEntryID IN (
                SELECT a.AssetEntryID FROM assets a
                JOIN AssetMapping am ON a.MappingID = am.MappingID
                WHERE am.AssetId = 'Asset07'
                  AND (TRIM(COALESCE(am.AssetSymbol,'')) = '')
            )
        """)

    _update_portfolio_from_assets(conn)

    # Override Asset07 InvestedValue with total pay-ins (demat_invested_capital) so that
    # the Wealth Engine shows true capital deployed, not fluctuating cost basis of open lots.
    # CurrentValue = cash_balance + market value of all holdings.
    ic_row = conn.execute("SELECT value FROM app_settings WHERE key='demat_invested_capital'").fetchone()
    if ic_row and float(ic_row['value'] or 0) > 0:
        ic      = round(float(ic_row['value']), 2)
        # total current value = cash in wallet + market value of all open holdings
        cash_bal = _demat_wallet_balance(conn)
        mkt_lots = conn.execute("""
            SELECT h.qty_remaining, COALESCE(n.ltp, h.buy_price) AS price
            FROM stock_holdings h
            LEFT JOIN nse_master n ON UPPER(n.symbol)=UPPER(h.symbol)
            WHERE h.qty_remaining > 0
        """).fetchall()
        mkt_val  = round(sum(float(l['qty_remaining']) * float(l['price']) for l in mkt_lots), 2)
        cur_total = round(cash_bal + mkt_val, 2)
        ret       = round(cur_total - ic, 2)
        ret_pct   = round(ret / ic * 100, 2) if ic > 0 else 0
        conn.execute("""
            UPDATE portfolio
            SET InvestedValue=?, CurrentValue=?, ReturnValue=?, ReturnPCT=?, UpdateAt=datetime('now')
            WHERE AssetID='Asset07'
        """, (ic, cur_total, ret, ret_pct))

    conn.commit()
    return synced


@app.route('/api/demat/sync_to_wealth', methods=['POST'])
def api_demat_sync_to_wealth():
    """Sync demat stock_holdings → Wealth Engine assets/portfolio."""
    conn = get_db()
    try:
        synced = _sync_demat_to_wealth(conn)
        conn.close()
        return jsonify({'success': True, 'synced_symbols': len(synced), 'symbols': synced})
    except Exception as e:
        conn.rollback(); conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/demat/sync_prices', methods=['POST'])
def api_demat_sync_prices():
    """Fetch live LTP for all stock_holdings symbols via yfinance, update nse_master."""
    import yfinance as yf
    conn = get_db()
    try:
        symbols = conn.execute(
            "SELECT DISTINCT UPPER(symbol) AS symbol FROM stock_holdings WHERE qty_remaining > 0"
        ).fetchall()
        updated = []
        failed  = []
        for row in symbols:
            sym = row['symbol']
            # Ensure symbol exists in nse_master
            conn.execute(
                "INSERT OR IGNORE INTO nse_master (symbol, company_name) VALUES (?,?)",
                (sym, sym)
            )
            ltp = None
            for suffix in ['.NS', '.BO']:
                try:
                    ticker = yf.Ticker(sym + suffix)
                    info   = ticker.fast_info
                    price  = getattr(info, 'last_price', None) or getattr(info, 'previous_close', None)
                    if price and float(price) > 0:
                        ltp = round(float(price), 2)
                        break
                except Exception:
                    pass
            if ltp:
                conn.execute("UPDATE nse_master SET ltp=? WHERE UPPER(symbol)=?", (ltp, sym))
                updated.append({'symbol': sym, 'ltp': ltp})
            else:
                failed.append(sym)
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'updated': updated, 'failed': failed})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/demat/reset_tradebook', methods=['POST'])
def api_demat_reset_tradebook():
    """Clear all tradebook-sourced data so a fresh re-upload can be done."""
    conn = get_db()
    try:
        # Remove all trade transactions
        conn.execute("DELETE FROM stock_transactions")
        # Remove all realized P&L
        conn.execute("DELETE FROM stock_pnl")
        # Remove lots added by tradebook; restore ONBOARDING lots to full qty
        conn.execute("DELETE FROM stock_holdings WHERE lot_source != 'ONBOARDING'")
        conn.execute("UPDATE stock_holdings SET qty_remaining = qty_original WHERE lot_source = 'ONBOARDING'")
        # Remove trade-related wallet entries (BUY/SELL/CHARGE from trades)
        conn.execute("DELETE FROM demat_wallet WHERE type IN ('BUY','SELL','CHARGE')")
        # Remove mirrored invest_transactions from tradebook/onboarding
        conn.execute("DELETE FROM invest_transactions WHERE rationale IS NULL AND stock_name IN (SELECT DISTINCT symbol FROM stock_holdings)")
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Tradebook data cleared. Ready for fresh upload.'})
    except Exception as e:
        conn.rollback(); conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/demat/tradebook_upload', methods=['POST'])
def api_demat_tradebook_upload():
    """Upload Zerodha tradebook CSV. trade_type param: LONG or SWING (applies to all rows)."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f          = request.files['file']
    trade_type = (request.form.get('trade_type') or 'LONG').upper()
    if trade_type not in ('LONG', 'SWING'):
        trade_type = 'LONG'

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
    try:
        f.save(tmp.name); tmp.close()
        df = pd.read_csv(tmp.name)
        df.columns = [c.strip() for c in df.columns]

        # Normalize to lowercase first, then alias to canonical names
        df.columns = [c.lower().strip() for c in df.columns]
        # Column alias map — handles both title-case export and lowercase Zerodha CSV
        _COL = {
            'trade_type': 'action',    # Zerodha: trade_type = buy/sell
            'quantity':   'qty',       # Zerodha: quantity
            'order_execution_time': 'exec_time',
            'trade_id':   'trade_id',
        }
        df.rename(columns=_COL, inplace=True)

        required = {'symbol', 'trade_date', 'action', 'qty', 'price'}
        missing  = required - set(df.columns)
        if missing:
            return jsonify({'error': f'Missing columns: {", ".join(missing)}', 'columns_found': list(df.columns)}), 400

        conn     = get_db()
        inserted = skipped = dupes = 0
        errors   = []
        preview  = []
        affected_symbols = set()

        for i, row in df.iterrows():
            try:
                symbol    = str(row.get('symbol','') or '').strip().upper()
                isin      = str(row.get('isin','')   or '').strip()
                action    = str(row.get('action','') or '').strip().upper()
                qty       = float(row.get('qty',0)   or 0)
                price     = float(row.get('price',0) or 0)
                # Prefer trade_id (unique per fill) over order_id for dedup
                trade_id  = str(row.get('trade_id','') or '').strip()
                order_id  = str(row.get('order_id','') or '').strip()
                dedup_key = trade_id or order_id   # trade_id is per-fill; order_id may be shared
                exchange  = str(row.get('exchange','NSE') or 'NSE').strip()

                raw_date   = str(row.get('trade_date','') or '').strip()
                trade_date = _normalise_date(raw_date)

                if not symbol or qty <= 0 or action not in ('BUY','SELL'):
                    skipped += 1; continue

                # Duplicate guard using trade_id (unique per fill) to avoid skipping partial fills
                if dedup_key:
                    dup = conn.execute(
                        "SELECT id FROM stock_transactions WHERE zerodha_order_id=? AND symbol=?",
                        (dedup_key, symbol)
                    ).fetchone()
                    if dup:
                        dupes += 1; continue

                charges = _zerodha_charges(action, qty, price)
                net_amt = round(qty * price, 2)
                if action == 'BUY':
                    net_amt_signed = -(net_amt + charges['total'])
                else:
                    net_amt_signed = net_amt - charges['total']

                # Insert stock_transaction
                cur = conn.execute("""
                    INSERT INTO stock_transactions
                        (trade_date, symbol, isin, action, qty, price, trade_type,
                         exchange, stt, other_charges, total_charges, net_amount, zerodha_order_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (trade_date, symbol, isin, action, qty, price, trade_type,
                      exchange, charges['stt'],
                      charges['exchange'] + charges['gst'] + charges['sebi'] + charges['stamp'],
                      charges['total'], net_amt_signed, dedup_key))
                txn_id = cur.lastrowid

                if action == 'BUY':
                    # Create FIFO lot
                    conn.execute("""
                        INSERT INTO stock_holdings
                            (symbol, isin, trade_type, buy_date, qty_original, qty_remaining, buy_price)
                        VALUES (?,?,?,?,?,?,?)
                    """, (symbol, isin, trade_type, trade_date, qty, qty, price))
                    # Mirror into invest_transactions for Asset07
                    conn.execute("""
                        INSERT INTO invest_transactions (entry_date, stock_name, asset_type, action, quantity, price, invested_value)
                        VALUES (?,?,?,?,?,?,?)
                    """, (trade_date, symbol, 'Stock', 'BUY', qty, price, round(qty*price,2)))
                    affected_symbols.add(symbol)
                    # Deduct from wallet
                    conn.execute(
                        "INSERT INTO demat_wallet (txn_date, type, amount, note, ref_txn_id) VALUES (?,?,?,?,?)",
                        (trade_date, 'BUY', net_amt, f'BUY {qty} {symbol} @ {price}', txn_id)
                    )
                    if charges['total'] > 0:
                        conn.execute(
                            "INSERT INTO demat_wallet (txn_date, type, amount, note, ref_txn_id) VALUES (?,?,?,?,?)",
                            (trade_date, 'CHARGE', charges['total'], f'Charges {symbol} {trade_date}', txn_id)
                        )
                else:  # SELL
                    # Mirror into invest_transactions for Asset07
                    conn.execute("""
                        INSERT INTO invest_transactions (entry_date, stock_name, asset_type, action, quantity, price, invested_value)
                        VALUES (?,?,?,?,?,?,?)
                    """, (trade_date, symbol, 'Stock', 'SELL', qty, price, round(qty*price,2)))
                    affected_symbols.add(symbol)
                    consumed, unfilled = _fifo_sell(conn, symbol, trade_type, qty, price, trade_date, txn_id)
                    lot_ids = [c['lot_id'] for c in consumed]
                    conn.execute("UPDATE stock_transactions SET lot_ids_affected=? WHERE id=?",
                                 (_json.dumps(lot_ids), txn_id))

                    # Write P&L records
                    for c in consumed:
                        net_pnl = round(c['gross_pnl'] - (charges['total'] * c['qty'] / qty), 2)
                        conn.execute("""
                            INSERT INTO stock_pnl
                                (sell_txn_id, symbol, sell_date, trade_type, qty_sold,
                                 avg_buy_price, sell_price, gross_pnl, charges, net_pnl,
                                 holding_days, tax_category)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                        """, (txn_id, symbol, trade_date, trade_type, c['qty'],
                              c['buy_price'], price, c['gross_pnl'],
                              round(charges['total'] * c['qty'] / qty, 2),
                              net_pnl, c['holding_days'], c['tax_category']))

                    # Add proceeds to wallet
                    proceeds = round(net_amt - charges['total'], 2)
                    conn.execute(
                        "INSERT INTO demat_wallet (txn_date, type, amount, note, ref_txn_id) VALUES (?,?,?,?,?)",
                        (trade_date, 'SELL', proceeds, f'SELL {qty} {symbol} @ {price}', txn_id)
                    )
                    if unfilled > 0:
                        errors.append(f'Row {i+2}: {symbol} — sold {unfilled} more shares than available lots ({trade_type})')

                inserted += 1
                if len(preview) < 20:
                    preview.append({
                        'symbol': symbol, 'action': action, 'qty': qty, 'price': price,
                        'trade_date': trade_date, 'trade_type': trade_type,
                        'charges': charges['total']
                    })
            except Exception as e:
                errors.append(f'Row {i+2}: {e}')

        conn.commit()
        # Sync all holdings → Wealth Engine
        _sync_demat_to_wealth(conn)
        balance = _demat_wallet_balance(conn)
        conn.close()
        return jsonify({
            'success': True, 'inserted': inserted, 'skipped': skipped,
            'duplicates': dupes, 'errors': errors[:20],
            'wallet_balance': balance, 'preview': preview
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try: os.unlink(tmp.name)
        except: pass


@app.route('/api/demat/corporate_action', methods=['POST'])
def api_demat_corporate_action():
    """Handle SPLIT or BONUS for a symbol."""
    d      = request.get_json(force=True)
    symbol = (d.get('symbol') or '').strip().upper()
    action = (d.get('action') or '').upper()  # SPLIT or BONUS
    conn   = get_db()
    try:
        if action == 'SPLIT':
            ratio = float(d.get('ratio') or 1)  # e.g. 2 means 2-for-1 split
            if ratio <= 0:
                return jsonify({'error': 'ratio must be > 0'}), 400
            lots = conn.execute(
                "SELECT id, qty_remaining, buy_price FROM stock_holdings WHERE UPPER(symbol)=? AND qty_remaining>0",
                (symbol,)
            ).fetchall()
            for lot in lots:
                new_qty   = round(lot['qty_remaining'] * ratio, 6)
                new_price = round(lot['buy_price'] / ratio, 4)
                conn.execute("UPDATE stock_holdings SET qty_remaining=?, buy_price=? WHERE id=?",
                             (new_qty, new_price, lot['id']))
            # Also update qty_original proportionally
            conn.execute("""
                UPDATE stock_holdings SET qty_original = qty_original * ?
                WHERE UPPER(symbol)=?
            """, (ratio, symbol))
            conn.execute("""
                INSERT INTO stock_transactions (trade_date, symbol, action, qty, price, trade_type, lot_source)
                VALUES (?,?,'SPLIT',?,?,?,?)
            """, (d.get('date') or datetime.now().strftime('%Y-%m-%d'), symbol, ratio, 0, 'NA', 'CORPORATE'))
            conn.commit()
            return jsonify({'success': True, 'lots_updated': len(lots)})

        elif action == 'BONUS':
            trade_type = (d.get('trade_type') or 'LONG').upper()
            qty        = float(d.get('qty') or 0)
            bonus_date = (d.get('date') or datetime.now().strftime('%Y-%m-%d'))[:10]
            if qty <= 0:
                return jsonify({'error': 'qty must be > 0'}), 400
            conn.execute("""
                INSERT INTO stock_holdings (symbol, trade_type, buy_date, qty_original, qty_remaining, buy_price, lot_source)
                VALUES (?,?,?,?,?,0,'BONUS')
            """, (symbol, trade_type, bonus_date, qty, qty))
            conn.execute("""
                INSERT INTO stock_transactions (trade_date, symbol, action, qty, price, trade_type, lot_source)
                VALUES (?,?,'BONUS',?,0,?,'CORPORATE')
            """, (bonus_date, symbol, qty, trade_type))
            conn.commit()
            return jsonify({'success': True})

        else:
            return jsonify({'error': 'action must be SPLIT or BONUS'}), 400
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/demat/dividends')
def api_demat_dividends_list():
    conn = get_db()
    rows = conn.execute(
        "SELECT id, symbol, date, per_share, total_amount FROM stock_dividends ORDER BY date DESC LIMIT 100"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/demat/dividend', methods=['POST'])
def api_demat_dividend_add():
    """Record a dividend: adds to demat_wallet (DIVIDEND) + transactions (income)."""
    d      = request.get_json(force=True)
    symbol = (d.get('symbol') or '').strip().upper()
    total  = float(d.get('total_amount') or 0)
    pps    = float(d.get('per_share')    or 0)
    date   = (d.get('date') or datetime.now().strftime('%Y-%m-%d'))[:10]
    if not symbol or total <= 0:
        return jsonify({'error': 'symbol and total_amount required'}), 400
    conn = get_db()
    try:
        # Record in dividend log
        conn.execute(
            "INSERT INTO stock_dividends (symbol, date, per_share, total_amount) VALUES (?,?,?,?)",
            (symbol, date, pps, total)
        )
        # Add to demat wallet (increases cash balance)
        conn.execute(
            "INSERT INTO demat_wallet (txn_date, type, amount, note) VALUES (?,?,?,?)",
            (date, 'DIVIDEND', total, f'Dividend from {symbol}')
        )
        # Mirror to transactions as income (Investment Return)
        conn.execute(
            "INSERT INTO transactions (type, category, sub_category, amount, date, note) VALUES (?,?,?,?,?,?)",
            ('income', 'Investment Return', 'Dividend', total, date, f'Dividend from {symbol}')
        )
        conn.commit()
        balance = _demat_wallet_balance(conn)
        conn.close()
        return jsonify({'success': True, 'balance': balance})
    except Exception as e:
        conn.rollback(); conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/demat/update_price', methods=['POST'])
def api_demat_update_price():
    """Manually update current price for a symbol in nse_master (used for portfolio valuation)."""
    d      = request.get_json(force=True)
    symbol = (d.get('symbol') or '').strip().upper()
    price  = float(d.get('price') or 0)
    if not symbol or price <= 0:
        return jsonify({'error': 'symbol and price required'}), 400
    conn = get_db()
    existing = conn.execute("SELECT symbol FROM nse_master WHERE UPPER(symbol)=?", (symbol,)).fetchone()
    if existing:
        conn.execute("UPDATE nse_master SET ltp=?, updated_at=datetime('now') WHERE UPPER(symbol)=?",
                     (price, symbol))
    else:
        conn.execute("INSERT INTO nse_master (symbol, ltp, company_name) VALUES (?,?,?)",
                     (symbol, price, symbol))
    conn.commit(); conn.close()
    return jsonify({'success': True})


# ══════════════════════════════════════════════════════════════════════════════
# REPORTS
# ══════════════════════════════════════════════════════════════════════════════
# Reusable SQL fragment: normalise any date column → 'YYYY-MM'
# Priority: strftime (handles ISO 8601 inc. non-zero-padded), then DD-MM-YYYY variants
_MONTH_SQL = """COALESCE(
    strftime('%Y-%m', {col}),
    CASE
        -- Already YYYY-MM (e.g. invest_transactions.month column)
        WHEN length({col})=7 AND {col} GLOB '????-??'
             AND CAST(substr({col},1,4) AS INTEGER) BETWEEN 1900 AND 2200
        THEN {col}
        -- Already YYYY-M (single-digit month, no leading zero)
        WHEN length({col})=6 AND {col} GLOB '????-?'
             AND CAST(substr({col},1,4) AS INTEGER) BETWEEN 1900 AND 2200
        THEN substr({col},1,4)||'-0'||substr({col},6,1)
        -- DD-MM-YYYY: validate year is a sane 4-digit integer (rejects "2-01" style corruption)
        WHEN {col} GLOB '??-??-????*'
             AND CAST(substr({col},7,4) AS INTEGER) BETWEEN 1900 AND 2200
        THEN substr({col},7,4)||'-'||substr({col},4,2)
        -- D-MM-YYYY
        WHEN {col} GLOB '?-??-????*'
             AND CAST(substr({col},6,4) AS INTEGER) BETWEEN 1900 AND 2200
        THEN substr({col},6,4)||'-'||substr({col},3,2)
        -- DD-M-YYYY
        WHEN {col} GLOB '??-?-????*'
             AND CAST(substr({col},6,4) AS INTEGER) BETWEEN 1900 AND 2200
        THEN substr({col},6,4)||'-0'||substr({col},4,1)
        -- D-M-YYYY
        WHEN {col} GLOB '?-?-????*'
             AND CAST(substr({col},5,4) AS INTEGER) BETWEEN 1900 AND 2200
        THEN substr({col},5,4)||'-0'||substr({col},3,1)
        -- DD/MM/YYYY (slash-separated)
        WHEN {col} GLOB '??/??/????*'
             AND CAST(substr({col},7,4) AS INTEGER) BETWEEN 1900 AND 2200
        THEN substr({col},7,4)||'-'||substr({col},4,2)
    END
)"""

def _month_sql(col='date'):
    return _MONTH_SQL.format(col=col)

@app.route('/api/reports/net_worth')
def api_report_net_worth():
    conn = get_db()
    rows = conn.execute("""
        SELECT im.AssetClass AS asset_class, im.AssetCategory AS asset_category,
               SUM(p.InvestedValue) AS invested, SUM(p.CurrentValue) AS current_value
        FROM portfolio p
        JOIN InvestMapping im ON p.AssetID = im.AssetID
        GROUP BY im.AssetClass, im.AssetCategory
        ORDER BY current_value DESC
    """).fetchall()
    total_invested = sum(r['invested'] or 0 for r in rows)
    total_current  = sum(r['current_value'] or 0 for r in rows)
    conn.close()
    return jsonify({
        'rows': [dict(r) for r in rows],
        'total_invested': total_invested,
        'total_current': total_current,
        'total_pnl': total_current - total_invested,
        'total_pnl_pct': r2((total_current - total_invested) / total_invested * 100) if total_invested > 0 else 0 if total_invested else 0
    })

@app.route('/api/reports/monthly_cashflow')
def api_report_monthly_cashflow():
    conn = get_db()
    rows = conn.execute(f"""
        SELECT {_month_sql()} AS month,
               SUM(CASE WHEN type='income'  THEN amount ELSE 0 END) AS income,
               SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) AS expense,
               SUM(CASE WHEN type='savings' THEN amount ELSE 0 END) AS savings
        FROM transactions
        WHERE date IS NOT NULL AND date != ''
        GROUP BY month
        ORDER BY month DESC
        LIMIT 24
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/reports/expense_by_category')
def api_report_expense_category():
    conn = get_db()
    rows = conn.execute("""
        SELECT COALESCE(NULLIF(TRIM(category),''), 'Uncategorised') AS category,
               SUM(amount) AS total
        FROM transactions
        WHERE type='expense'
        GROUP BY category
        ORDER BY total DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/reports/investment_monthly')
def api_report_investment_monthly():
    conn = get_db()
    rows = conn.execute("""
        SELECT substr(COALESCE(entry_date, month), 1, 7) AS month,
               asset_type,
               SUM(CASE WHEN UPPER(action)='BUY'  THEN invested_value ELSE 0 END) AS invested,
               SUM(CASE WHEN UPPER(action)='SELL' THEN invested_value ELSE 0 END) AS redeemed
        FROM invest_transactions
        WHERE (entry_date IS NOT NULL OR month IS NOT NULL)
        GROUP BY month, asset_type
        ORDER BY month DESC
        LIMIT 120
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/reports/stock_pnl')
def api_report_stock_pnl():
    conn = get_db()
    # Realized P&L from stock_pnl table if present
    try:
        rows = conn.execute("""
            SELECT symbol, SUM(realized_pnl) AS realized_pnl,
                   SUM(qty_sold) AS qty_sold,
                   MIN(sell_date) AS first_sell, MAX(sell_date) AS last_sell
            FROM stock_pnl
            GROUP BY symbol
            ORDER BY realized_pnl DESC
        """).fetchall()
    except Exception:
        rows = []
    # Unrealized: stock_holdings (demat tradebook) + Wealth Engine Asset07 stocks
    try:
        sh_syms = _it_equity_symbols(conn)
        holdings_sh = conn.execute("""
            SELECT sh.symbol,
                   SUM(sh.qty_remaining * sh.buy_price)                   AS cost_basis,
                   SUM(sh.qty_remaining * COALESCE(nm.ltp, sh.buy_price)) AS market_value,
                   SUM(sh.qty_remaining)                                   AS qty
            FROM stock_holdings sh
            LEFT JOIN nse_master nm ON UPPER(nm.symbol) = UPPER(sh.symbol)
            WHERE sh.qty_remaining > 0
            GROUP BY sh.symbol
        """).fetchall()
    except Exception:
        sh_syms = set()
        holdings_sh = []
    try:
        it_extra = _it_stocks_as_holdings(conn, exclude_symbols=sh_syms)
        holdings_it = [
            {'symbol': it['symbol'],
             'cost_basis':   round(it['qty_remaining'] * it['buy_price'], 2),
             'market_value': round(it['qty_remaining'] * it['current_price'], 2)
                             if it['current_price'] > 0
                             else round(it['qty_remaining'] * it['buy_price'], 2),
             'qty':          it['qty_remaining']}
            for it in it_extra
        ]
    except Exception:
        holdings_it = []
    holdings = list(holdings_sh) + holdings_it
    # deduplicate by symbol (keep first occurrence)
    seen_s = set()
    holdings_deduped = []
    for h in holdings:
        sym = (h['symbol'] if isinstance(h, dict) else h['symbol']).upper()
        if sym not in seen_s:
            seen_s.add(sym)
            holdings_deduped.append(h)
    holdings = holdings_deduped
    conn.close()
    return jsonify({
        'realized': [dict(r) for r in rows],
        'unrealized': [dict(h) for h in holdings]
    })

@app.route('/api/reports/loan_summary')
def api_report_loan_summary():
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT loan_name, loan_type,
                   loan_amount AS principal_amount,
                   total_repayment,
                   start_date, target_close_date AS end_date, status
            FROM loan_master
            ORDER BY loan_amount DESC
        """).fetchall()
    except Exception:
        rows = []
    # Monthly loan payments from loans table
    try:
        payments = conn.execute("""
            SELECT loan_type, SUM(amount) AS paid
            FROM loans GROUP BY loan_type
        """).fetchall()
        paid_map = {r['loan_type']: float(r['paid'] or 0) for r in payments}
    except Exception:
        paid_map = {}
    result = []
    for r in rows:
        d = dict(r)
        d['outstanding_amount'] = max(0, float(d['principal_amount'] or 0) - paid_map.get(d['loan_type'], 0))
        d['emi_amount'] = 0  # not stored separately
        d['interest_rate'] = 0
        d['borrower_name'] = d.get('loan_name', '—')
        result.append(d)
    total_principal   = sum(r['principal_amount'] or 0 for r in rows)
    total_outstanding = sum(r['outstanding_amount'] for r in result)
    total_emi         = 0
    conn.close()
    return jsonify({
        'loans': result,
        'total_principal': total_principal,
        'total_outstanding': total_outstanding,
        'total_emi': total_emi
    })

@app.route('/api/reports/demat_summary')
def api_report_demat_summary():
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT sh.symbol,
                   SUM(sh.qty_remaining) AS qty,
                   SUM(sh.qty_remaining * sh.buy_price) AS cost_basis,
                   SUM(sh.qty_remaining * COALESCE(nm.ltp, sh.buy_price)) AS market_value,
                   COALESCE(nm.ltp, 0) AS ltp,
                   MAX(sh.buy_date) AS last_buy
            FROM stock_holdings sh
            LEFT JOIN nse_master nm ON UPPER(nm.symbol) = UPPER(sh.symbol)
            WHERE sh.qty_remaining > 0
            GROUP BY sh.symbol
            ORDER BY market_value DESC
        """).fetchall()
    except Exception:
        rows = []
    invested = float(conn.execute(
        "SELECT COALESCE(value,0) FROM app_settings WHERE key='demat_invested_capital'"
    ).fetchone()['value'] if conn.execute(
        "SELECT value FROM app_settings WHERE key='demat_invested_capital'"
    ).fetchone() else 0)
    conn.close()
    return jsonify({
        'holdings': [dict(r) for r in rows],
        'invested_capital': invested
    })

@app.route('/api/reports/income_by_source')
def api_report_income_by_source():
    conn = get_db()
    by_cat = conn.execute("""
        SELECT COALESCE(NULLIF(TRIM(category),''),'Other') AS category,
               SUM(amount) AS total, COUNT(*) AS txn_count
        FROM transactions WHERE type='income'
        GROUP BY category ORDER BY total DESC
    """).fetchall()
    monthly = conn.execute(f"""
        SELECT {_month_sql()} AS month, SUM(amount) AS total
        FROM transactions WHERE type='income' AND date IS NOT NULL
        GROUP BY month ORDER BY month DESC LIMIT 24
    """).fetchall()
    conn.close()
    return jsonify({'by_category': [dict(r) for r in by_cat], 'monthly': [dict(r) for r in monthly]})

@app.route('/api/reports/expense_trends')
def api_report_expense_trends():
    conn = get_db()
    monthly = conn.execute(f"""
        SELECT {_month_sql()} AS month,
               COALESCE(NULLIF(TRIM(category),''),'Other') AS category,
               SUM(amount) AS total
        FROM transactions WHERE type='expense' AND date IS NOT NULL
        GROUP BY month, category ORDER BY month DESC LIMIT 200
    """).fetchall()
    by_month = conn.execute(f"""
        SELECT {_month_sql()} AS month, SUM(amount) AS total
        FROM transactions WHERE type='expense' AND date IS NOT NULL
        GROUP BY month ORDER BY month DESC LIMIT 24
    """).fetchall()
    top_cat = conn.execute("""
        SELECT COALESCE(NULLIF(TRIM(category),''),'Other') AS category,
               COUNT(*) AS txn_count, SUM(amount) AS total, AVG(amount) AS avg_amt
        FROM transactions WHERE type='expense'
        GROUP BY category ORDER BY total DESC LIMIT 20
    """).fetchall()
    conn.close()
    return jsonify({
        'monthly_by_cat': [dict(r) for r in monthly],
        'monthly_total':  [dict(r) for r in by_month],
        'top_categories': [dict(r) for r in top_cat]
    })

@app.route('/api/reports/comparisons')
def api_report_comparisons():
    conn = get_db()
    # Monthly income, expense, savings, investment, loan_emi rolled up
    cf = conn.execute(f"""
        SELECT {_month_sql()} AS month,
               SUM(CASE WHEN type='income'  THEN amount ELSE 0 END) AS income,
               SUM(CASE WHEN type='expense' THEN amount ELSE 0 END) AS expense,
               SUM(CASE WHEN type='savings' THEN amount ELSE 0 END) AS savings
        FROM transactions WHERE date IS NOT NULL
        GROUP BY month HAVING month IS NOT NULL ORDER BY month ASC
    """).fetchall()
    inv = conn.execute(f"""
        SELECT {_month_sql('COALESCE(entry_date,month)')} AS month,
               SUM(CASE WHEN UPPER(action)='BUY' THEN invested_value ELSE 0 END) AS invested
        FROM invest_transactions WHERE (entry_date IS NOT NULL OR month IS NOT NULL)
        GROUP BY month HAVING month IS NOT NULL ORDER BY month ASC
    """).fetchall()
    # Per-month actual EMI paid (sum of transactions in category 'Loan EMI')
    try:
        emi_rows = conn.execute(f"""
            SELECT {_month_sql()} AS month, SUM(amount) AS emi_total
            FROM transactions
            WHERE type='expense' AND LOWER(category) IN ('loan emi','loan_emi','emi')
              AND date IS NOT NULL
            GROUP BY month HAVING month IS NOT NULL
        """).fetchall()
    except Exception:
        emi_rows = []
    conn.close()
    inv_map = {r['month']: float(r['invested'] or 0) for r in inv if r['month']}
    cf_map  = {r['month']: r for r in cf if r['month']}
    emi_map = {r['month']: float(r['emi_total'] or 0) for r in emi_rows if r['month']}
    all_months = sorted(set(cf_map.keys()) | set(inv_map.keys()))
    result = []
    for m in all_months:
        r = cf_map.get(m)
        result.append({
            'month':    m,
            'income':   float(r['income']  or 0) if r else 0,
            'expense':  float(r['expense'] or 0) if r else 0,
            'savings':  float(r['savings'] or 0) if r else 0,
            'invested': inv_map.get(m, 0),
            'loan_emi': emi_map.get(m, 0),
        })
    return jsonify(result)

@app.route('/api/reports/loan_emi_timeline')
def api_report_loan_emi_timeline():
    conn = get_db()
    try:
        loans = conn.execute("""
            SELECT loan_name, loan_type, loan_amount AS principal_amount,
                   total_repayment, start_date, target_close_date AS end_date, status
            FROM loan_master ORDER BY loan_amount DESC
        """).fetchall()
        payments = conn.execute("""
            SELECT month, SUM(amount) AS paid FROM loans
            GROUP BY month ORDER BY month DESC LIMIT 24
        """).fetchall()
    except Exception:
        loans, payments = [], []
    conn.close()
    return jsonify({
        'loans':     [dict(r) for r in loans],
        'payments':  [dict(r) for r in payments],
        'total_emi': 0
    })

@app.route('/api/reports/transactions_full')
def api_report_transactions_full():
    """Full raw transaction tables for Excel download."""
    conn = get_db()
    txns   = conn.execute("SELECT * FROM transactions ORDER BY date DESC LIMIT 5000").fetchall()
    inv    = conn.execute("SELECT * FROM invest_transactions ORDER BY entry_date DESC LIMIT 2000").fetchall()
    loans  = conn.execute("SELECT * FROM loans ORDER BY month DESC LIMIT 2000").fetchall()
    loan_m = conn.execute("SELECT * FROM loan_master ORDER BY created_at DESC LIMIT 2000").fetchall()
    conn.close()
    return jsonify({
        'transactions':        [dict(r) for r in txns],
        'invest_transactions': [dict(r) for r in inv],
        'loans':               [dict(r) for r in loans],
        'loan_master':         [dict(r) for r in loan_m],
    })

@app.route('/api/reports/assets_detail')
def api_report_assets_detail():
    """One row per asset entry — full join of assets + AssetMapping + InvestMapping."""
    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT
                a.AssetEntryID          AS entry_id,
                am.AssetName            AS asset_name,
                am.AssetSymbol          AS symbol,
                im.AssetID              AS asset_id,
                im.AssetClass           AS asset_class,
                im.AssetCategory        AS asset_category,
                im.AssetType            AS asset_type,
                im.PriceFetchMode       AS price_fetch_mode,
                im.Symbol               AS mapping_symbol,
                im.WeightGrams          AS weight_grams,
                im.Purity               AS purity,
                im.InterestRate         AS interest_rate,
                a.purpose               AS purpose,
                a.qty                   AS quantity,
                a.avgprice              AS avg_price,
                a.ltp                   AS ltp,
                a.investedvalue         AS invested_value,
                a.currentvalue          AS current_value,
                a.pnl                   AS pnl,
                a.pnlpct                AS pnl_pct,
                a.targetpct             AS target_allocation_pct,
                a.lastsynced            AS last_synced,
                a.updatedat             AS updated_at
            FROM assets a
            JOIN AssetMapping am ON a.MappingID = am.MappingID
            JOIN InvestMapping im ON am.AssetId  = im.AssetID
            ORDER BY im.AssetClass, im.AssetCategory, im.AssetType, am.AssetName
        """).fetchall()
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e), 'rows': []})
    conn.close()
    return jsonify([dict(r) for r in rows])

# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# DOCUMENT WALLET
# ══════════════════════════════════════════════════════════════════════════════
import os, uuid as _uuid_mod, base64 as _b64

DOCS_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'documents')
os.makedirs(DOCS_FOLDER, exist_ok=True)

def _init_docwallet_schema(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS doc_wallet (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_type    TEXT NOT NULL DEFAULT '',
            doc_name    TEXT NOT NULL,
            doc_details TEXT,
            created_at  TEXT DEFAULT (datetime('now')),
            updated_at  TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS doc_attachments (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_id      INTEGER NOT NULL,
            file_name   TEXT,
            file_ext    TEXT,
            mime_type   TEXT,
            file_size   INTEGER DEFAULT 0,
            thumb_name  TEXT,
            label       TEXT DEFAULT '',
            uploaded_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS doc_wallet_types (
            id    INTEGER PRIMARY KEY AUTOINCREMENT,
            label TEXT NOT NULL UNIQUE
        )
    """)
    # migrate old single-file rows to doc_attachments
    try:
        old_cols = [r[1] for r in conn.execute("PRAGMA table_info(doc_wallet)").fetchall()]
        if 'file_name' in old_cols:
            for row in conn.execute("SELECT id,file_name,file_ext,mime_type,file_size FROM doc_wallet WHERE file_name IS NOT NULL AND file_name!=''").fetchall():
                exists = conn.execute("SELECT id FROM doc_attachments WHERE doc_id=? AND file_name=?", (row['id'], row['file_name'])).fetchone()
                if not exists:
                    conn.execute("INSERT INTO doc_attachments (doc_id,file_name,file_ext,mime_type,file_size) VALUES (?,?,?,?,?)",
                                 (row['id'], row['file_name'], row['file_ext'], row['mime_type'], row['file_size']))
    except Exception:
        pass
    conn.commit()

def _save_uploaded_file(f):
    """Save an uploaded FileStorage to DOCS_FOLDER. Returns (safe_name, ext, mime, size)."""
    orig_ext  = os.path.splitext(f.filename)[1].lower()
    safe_name = _uuid_mod.uuid4().hex + orig_ext
    fp = os.path.join(DOCS_FOLDER, safe_name)
    f.save(fp)
    return safe_name, orig_ext.lstrip('.'), f.content_type or '', os.path.getsize(fp)

def _save_thumbnail_b64(data_url: str):
    """Save a base64 data URL as a JPEG thumbnail. Returns thumb_name or None."""
    try:
        header, b64 = data_url.split(',', 1)
        raw = _b64.b64decode(b64)
        ext = '.jpg' if 'jpeg' in header else '.png'
        name = 'thumb_' + _uuid_mod.uuid4().hex + ext
        with open(os.path.join(DOCS_FOLDER, name), 'wb') as fh:
            fh.write(raw)
        return name
    except Exception:
        return None

def _delete_doc_files(conn, doc_id):
    """Delete all physical files for a doc and remove attachment rows."""
    atts = conn.execute("SELECT file_name,thumb_name FROM doc_attachments WHERE doc_id=?", (doc_id,)).fetchall()
    for a in atts:
        for fn in (a['file_name'], a['thumb_name']):
            if fn:
                fp = os.path.join(DOCS_FOLDER, fn)
                if os.path.exists(fp):
                    try: os.remove(fp)
                    except Exception: pass
    conn.execute("DELETE FROM doc_attachments WHERE doc_id=?", (doc_id,))

def _doc_with_attachments(conn, doc_id):
    doc = conn.execute("SELECT * FROM doc_wallet WHERE id=?", (doc_id,)).fetchone()
    if not doc: return None
    atts = conn.execute("SELECT * FROM doc_attachments WHERE doc_id=? ORDER BY uploaded_at DESC", (doc_id,)).fetchall()
    d = dict(doc)
    d['attachments'] = [dict(a) for a in atts]
    d['latest_attachment'] = d['attachments'][0] if d['attachments'] else None
    return d

# ── List all docs (summary) ──────────────────────────────────────────────────
@app.route('/api/docwallet', methods=['GET'])
def api_docwallet_list():
    conn = get_db(); _init_docwallet_schema(conn)
    docs = conn.execute("SELECT * FROM doc_wallet ORDER BY doc_type, doc_name").fetchall()
    result = []
    for doc in docs:
        atts = conn.execute("SELECT * FROM doc_attachments WHERE doc_id=? ORDER BY uploaded_at DESC", (doc['id'],)).fetchall()
        d = dict(doc)
        d['attachments'] = [dict(a) for a in atts]
        d['latest_attachment'] = d['attachments'][0] if d['attachments'] else None
        result.append(d)
    conn.close()
    return jsonify(result)

# ── Create doc ───────────────────────────────────────────────────────────────
@app.route('/api/docwallet', methods=['POST'])
def api_docwallet_add():
    conn = get_db(); _init_docwallet_schema(conn)
    doc_type    = request.form.get('doc_type','').strip()
    doc_name    = request.form.get('doc_name','').strip()
    doc_details = request.form.get('doc_details','').strip()
    if not doc_name:
        conn.close(); return jsonify({'error': 'Document name required'}), 400
    cur = conn.execute(
        "INSERT INTO doc_wallet (doc_type,doc_name,doc_details) VALUES (?,?,?)",
        (doc_type, doc_name, doc_details)
    )
    new_id = cur.lastrowid
    # Save all uploaded files
    for key in request.files:
        f = request.files[key]
        if f and f.filename:
            safe_name, ext, mime, size = _save_uploaded_file(f)
            thumb = None
            thumb_b64 = request.form.get('thumb_' + key)
            if thumb_b64: thumb = _save_thumbnail_b64(thumb_b64)
            label = request.form.get('label_' + key, '')
            conn.execute("INSERT INTO doc_attachments (doc_id,file_name,file_ext,mime_type,file_size,thumb_name,label) VALUES (?,?,?,?,?,?,?)",
                         (new_id, safe_name, ext, mime, size, thumb, label))
    conn.commit()
    result = _doc_with_attachments(conn, new_id)
    conn.close()
    return jsonify(result)

# ── Update doc metadata ──────────────────────────────────────────────────────
@app.route('/api/docwallet/<int:did>', methods=['PATCH'])
def api_docwallet_patch(did):
    conn = get_db(); _init_docwallet_schema(conn)
    d = request.json or {}
    conn.execute("""UPDATE doc_wallet SET doc_type=COALESCE(?,doc_type), doc_name=COALESCE(?,doc_name),
                    doc_details=COALESCE(?,doc_details), updated_at=datetime('now') WHERE id=?""",
                 (d.get('doc_type'), d.get('doc_name'), d.get('doc_details'), did))
    conn.commit()
    result = _doc_with_attachments(conn, did)
    conn.close()
    return jsonify(result or {'error': 'Not found'})

# ── Delete doc + all files ───────────────────────────────────────────────────
@app.route('/api/docwallet/<int:did>', methods=['DELETE'])
def api_docwallet_delete(did):
    conn = get_db(); _init_docwallet_schema(conn)
    _delete_doc_files(conn, did)
    conn.execute("DELETE FROM doc_wallet WHERE id=?", (did,))
    conn.commit(); conn.close()
    return jsonify({'success': True})

# ── Add attachment to existing doc ───────────────────────────────────────────
@app.route('/api/docwallet/<int:did>/attachments', methods=['POST'])
def api_docwallet_add_attachment(did):
    conn = get_db(); _init_docwallet_schema(conn)
    if not conn.execute("SELECT id FROM doc_wallet WHERE id=?", (did,)).fetchone():
        conn.close(); return jsonify({'error': 'Doc not found'}), 404

    replace_id = request.form.get('replace_id','').strip()
    if replace_id:
        old = conn.execute("SELECT file_name,thumb_name FROM doc_attachments WHERE id=? AND doc_id=?", (replace_id, did)).fetchone()
        if old:
            for fn in (old['file_name'], old['thumb_name']):
                if fn:
                    fp = os.path.join(DOCS_FOLDER, fn)
                    if os.path.exists(fp): os.remove(fp)
            conn.execute("DELETE FROM doc_attachments WHERE id=?", (replace_id,))

    added = []
    for key in request.files:
        f = request.files[key]
        if f and f.filename:
            safe_name, ext, mime, size = _save_uploaded_file(f)
            thumb = None
            thumb_b64 = request.form.get('thumb_' + key)
            if thumb_b64: thumb = _save_thumbnail_b64(thumb_b64)
            label = request.form.get('label_' + key, '')
            cur = conn.execute("INSERT INTO doc_attachments (doc_id,file_name,file_ext,mime_type,file_size,thumb_name,label) VALUES (?,?,?,?,?,?,?)",
                               (did, safe_name, ext, mime, size, thumb, label))
            added.append(cur.lastrowid)

    conn.execute("UPDATE doc_wallet SET updated_at=datetime('now') WHERE id=?", (did,))
    conn.commit()
    result = _doc_with_attachments(conn, did)
    conn.close()
    return jsonify(result)

# ── Save/update thumbnail for an attachment ──────────────────────────────────
@app.route('/api/docwallet/attach/<int:aid>/thumb', methods=['POST'])
def api_docwallet_save_thumb(aid):
    conn = get_db(); _init_docwallet_schema(conn)
    row = conn.execute("SELECT thumb_name FROM doc_attachments WHERE id=?", (aid,)).fetchone()
    if not row: conn.close(); return jsonify({'error': 'Not found'}), 404
    # Delete old thumb
    if row['thumb_name']:
        fp = os.path.join(DOCS_FOLDER, row['thumb_name'])
        if os.path.exists(fp): os.remove(fp)
    data_url = (request.json or {}).get('data_url','')
    thumb = _save_thumbnail_b64(data_url) if data_url else None
    conn.execute("UPDATE doc_attachments SET thumb_name=? WHERE id=?", (thumb, aid))
    conn.commit(); conn.close()
    return jsonify({'success': True, 'thumb_name': thumb})

# ── Delete single attachment ─────────────────────────────────────────────────
@app.route('/api/docwallet/attach/<int:aid>', methods=['DELETE'])
def api_docwallet_delete_attachment(aid):
    conn = get_db(); _init_docwallet_schema(conn)
    row = conn.execute("SELECT file_name,thumb_name,doc_id FROM doc_attachments WHERE id=?", (aid,)).fetchone()
    if not row: conn.close(); return jsonify({'error': 'Not found'}), 404
    for fn in (row['file_name'], row['thumb_name']):
        if fn:
            fp = os.path.join(DOCS_FOLDER, fn)
            if os.path.exists(fp): os.remove(fp)
    conn.execute("DELETE FROM doc_attachments WHERE id=?", (aid,))
    conn.execute("UPDATE doc_wallet SET updated_at=datetime('now') WHERE id=?", (row['doc_id'],))
    conn.commit()
    result = _doc_with_attachments(conn, row['doc_id'])
    conn.close()
    return jsonify(result)

# ── Serve attachment file ────────────────────────────────────────────────────
@app.route('/api/docwallet/attach/<int:aid>/file')
def api_docwallet_attach_file(aid):
    conn = get_db(); _init_docwallet_schema(conn)
    row = conn.execute("""SELECT da.file_name,da.file_ext,da.mime_type,dw.doc_name
                          FROM doc_attachments da JOIN doc_wallet dw ON dw.id=da.doc_id
                          WHERE da.id=?""", (aid,)).fetchone()
    conn.close()
    if not row or not row['file_name']:
        return jsonify({'error': 'No file'}), 404
    fp = os.path.join(DOCS_FOLDER, row['file_name'])
    if not os.path.exists(fp): return jsonify({'error': 'File missing'}), 404
    from flask import send_file
    dl_name = (row['doc_name'] or 'document') + '.' + (row['file_ext'] or 'bin')
    return send_file(fp, mimetype=row['mime_type'] or 'application/octet-stream',
                     as_attachment=False, download_name=dl_name)

# ── Serve thumbnail ──────────────────────────────────────────────────────────
@app.route('/api/docwallet/attach/<int:aid>/thumb')
def api_docwallet_attach_thumb(aid):
    conn = get_db(); _init_docwallet_schema(conn)
    row = conn.execute("SELECT thumb_name FROM doc_attachments WHERE id=?", (aid,)).fetchone()
    conn.close()
    if not row or not row['thumb_name']: return jsonify({'error': 'No thumb'}), 404
    fp = os.path.join(DOCS_FOLDER, row['thumb_name'])
    if not os.path.exists(fp): return jsonify({'error': 'Missing'}), 404
    from flask import send_file
    return send_file(fp, mimetype='image/jpeg')

# ── Custom types ─────────────────────────────────────────────────────────────
@app.route('/api/docwallet/types', methods=['GET'])
def api_docwallet_types_get():
    conn = get_db(); _init_docwallet_schema(conn)
    rows = conn.execute("SELECT label FROM doc_wallet_types ORDER BY label").fetchall()
    conn.close()
    return jsonify([r['label'] for r in rows])

@app.route('/api/docwallet/types', methods=['POST'])
def api_docwallet_types_add():
    conn = get_db(); _init_docwallet_schema(conn)
    label = (request.json or {}).get('label','').strip()
    if not label: conn.close(); return jsonify({'error': 'Label required'}), 400
    conn.execute("INSERT OR IGNORE INTO doc_wallet_types (label) VALUES (?)", (label,))
    conn.commit()
    rows = conn.execute("SELECT label FROM doc_wallet_types ORDER BY label").fetchall()
    conn.close()
    return jsonify([r['label'] for r in rows])

@app.route('/api/docwallet/types/<label>', methods=['DELETE'])
def api_docwallet_types_delete(label):
    conn = get_db(); _init_docwallet_schema(conn)
    conn.execute("DELETE FROM doc_wallet_types WHERE label=?", (label,))
    conn.commit()
    rows = conn.execute("SELECT label FROM doc_wallet_types ORDER BY label").fetchall()
    conn.close()
    return jsonify([r['label'] for r in rows])


# ─────────────────────────────────────────────────────────────────────────────
# ONBOARDING
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/onboarding/rerun', methods=['POST'])
@login_required
def api_onboarding_rerun():
    """Clear the onboarding_complete flag so the wizard shows again."""
    conn = get_db()
    conn.execute("DELETE FROM app_settings WHERE key='onboarding_complete'")
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/onboarding/skip', methods=['POST'])
@login_required
def api_onboarding_skip():
    """Mark onboarding as dismissed (overlay hidden) WITHOUT marking wizard_done.
    Re-run button stays enabled so the user can complete setup later."""
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('onboarding_complete','1')")
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


AVATAR_PATH = os.path.join(os.path.dirname(__file__), 'static', 'user_avatar.png')
DEFAULT_AVATAR_PATH = os.path.join(os.path.dirname(__file__), 'static', 'logo.png')

@app.route('/api/profile/avatar')
def api_profile_avatar():
    """Serve user avatar; fallback to default logo."""
    path = AVATAR_PATH if os.path.exists(AVATAR_PATH) else DEFAULT_AVATAR_PATH
    return send_file(path, mimetype='image/png', max_age=0)

@app.route('/api/profile/avatar', methods=['POST'])
def api_profile_avatar_upload():
    """Accept a base64 PNG data-URL or a file upload and save as user_avatar.png."""
    import base64, re as _re
    data = request.get_json(silent=True)
    if data and data.get('data_url'):
        # base64 data URL: data:image/png;base64,....
        m = _re.match(r'data:image/\w+;base64,(.+)', data['data_url'])
        if not m:
            return jsonify({'ok': False, 'error': 'Invalid data URL'}), 400
        img_bytes = base64.b64decode(m.group(1))
        with open(AVATAR_PATH, 'wb') as f:
            f.write(img_bytes)
        return jsonify({'ok': True})
    f = request.files.get('avatar')
    if f:
        f.save(AVATAR_PATH)
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'No image provided'}), 400


@app.route('/api/profile/avatar/remove', methods=['POST'])
@login_required
def api_profile_avatar_remove():
    try:
        if os.path.exists(AVATAR_PATH):
            os.remove(AVATAR_PATH)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/onboarding/save_credentials', methods=['POST'])
def api_onboarding_save_credentials():
    """Save name + password during onboarding Step 1 (no session required — used before restore)."""
    data = request.get_json(force=True) or {}
    name     = str(data.get('name', '') or '').strip()
    password = str(data.get('password', '') or '').strip()
    if not name or not password:
        return jsonify({'ok': False, 'error': 'Name and password are required'}), 400
    try:
        conn = get_db()
        conn.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('user_name', ?)", (name,))
        _set_password(password, conn)
        conn.commit()
        conn.close()
        # Update session so the user stays logged in after password change
        session['logged_in'] = True
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/onboarding/status')
@login_required
def api_onboarding_status():
    conn = get_db()
    row   = conn.execute("SELECT value FROM app_settings WHERE key='onboarding_complete'").fetchone()
    wdone = conn.execute("SELECT value FROM app_settings WHERE key='onboarding_wizard_done'").fetchone()
    name  = conn.execute("SELECT value FROM app_settings WHERE key='user_name'").fetchone()
    conn.close()
    return jsonify({
        'complete':     bool(row   and row['value']   == '1'),
        'wizard_done':  bool(wdone and wdone['value'] == '1'),
        'user_name':    name['value'] if name else '',
    })


def _create_onboarding_alerts(conn, income, expense, loan_emi, alert_config):
    """
    Create smart default alerts based on the user's income.
    alert_config is a list of dicts from the frontend:
      [{ "key": "expense_limit", "enabled": true, "threshold": 80000 }, ...]
    Only creates alerts that are enabled and don't already exist.
    """
    if income <= 0:
        return

    # Build a lookup from key → user-chosen threshold / enabled flag
    cfg = {a['key']: a for a in (alert_config or [])}

    def _should(key):
        return cfg.get(key, {}).get('enabled', True)

    def _thresh(key, default):
        return float(cfg.get(key, {}).get('threshold', default))

    def _exists(name_like):
        return conn.execute(
            "SELECT id FROM alerts WHERE name LIKE ?", ('%' + name_like + '%',)
        ).fetchone() is not None

    alerts_to_create = []

    # 1. Total monthly expense limit
    if _should('expense_limit') and not _exists('Expense Limit'):
        alerts_to_create.append((
            'Monthly Expense Limit',
            'expense', 'expense_limit',
            _thresh('expense_limit', round(income * 0.80 / 100) * 100),
            None, 'monthly'
        ))

    # 2. Household / Rent budget
    if _should('household') and not _exists('Household'):
        alerts_to_create.append((
            'Household Budget',
            'expense', 'category_exceeds',
            _thresh('household', round(income * 0.30 / 100) * 100),
            'Household', 'monthly'
        ))

    # 3. Transport budget
    if _should('transport') and not _exists('Transport'):
        alerts_to_create.append((
            'Transport Budget',
            'expense', 'category_exceeds',
            _thresh('transport', round(income * 0.08 / 100) * 100),
            'Transport', 'monthly'
        ))

    # 4. Shopping budget
    if _should('shopping') and not _exists('Shopping'):
        alerts_to_create.append((
            'Shopping Budget',
            'expense', 'category_exceeds',
            _thresh('shopping', round(income * 0.10 / 100) * 100),
            'Shopping', 'monthly'
        ))

    # 5. Lifestyle budget
    if _should('lifestyle') and not _exists('Lifestyle'):
        alerts_to_create.append((
            'Lifestyle Budget',
            'expense', 'category_exceeds',
            _thresh('lifestyle', round(income * 0.08 / 100) * 100),
            'Lifestyle', 'monthly'
        ))

    # 6. Monthly savings floor
    if _should('savings_floor') and not _exists('Savings Floor'):
        alerts_to_create.append((
            'Net Savings Floor',
            'savings', 'savings_below',
            _thresh('savings_floor', round(income * 0.10 / 100) * 100),
            None, 'monthly'
        ))

    # 7. Monthly investment target
    if _should('invest_target') and not _exists('Investment Target'):
        alerts_to_create.append((
            'Monthly Investment Target',
            'investment', 'total_below',
            _thresh('invest_target', round(income * 0.20 / 100) * 100),
            None, 'monthly'
        ))

    # 8. Loan EMI limit
    if _should('loan_emi') and not _exists('Loan EMI'):
        default_emi = loan_emi if loan_emi > 0 else round(income * 0.30 / 100) * 100
        alerts_to_create.append((
            'Monthly Loan EMI',
            'system', 'loan_emi_exceeds',
            _thresh('loan_emi', default_emi),
            None, 'monthly'
        ))

    # 9. Emergency fund floor
    if _should('emergency_fund') and not _exists('Emergency Fund'):
        ef_target = expense * 6 if expense > 0 else income * 3
        alerts_to_create.append((
            'Emergency Fund Target',
            'system', 'emergency_fund_below',
            _thresh('emergency_fund', round(ef_target / 1000) * 1000),
            None, 'monthly'
        ))

    for (name, typ, cond, thresh, cat, period) in alerts_to_create:
        conn.execute(
            "INSERT INTO alerts (name,type,condition,threshold,category,period,is_active) VALUES (?,?,?,?,?,?,1)",
            (name, typ, cond, thresh, cat, period)
        )


@app.route('/api/onboarding/setup', methods=['POST'])
@login_required
def api_onboarding_setup():
    """
    5-step onboarding seed. Expected JSON body:
    {
      "monthly_income":   <number>,
      "monthly_expense":  <number>,
      "emergency_fund":   <number>,   // current liquidity balance
      "loan_entries":     [ { "name", "loan_type", "loan_amount", "total_repayment", "start_date", "target_close_date" } ],
      "invest_entries":   [ { "stock_name", "asset_type", "quantity", "price", "action", "entry_date", "rationale" } ]
    }
    """
    data = request.get_json(force=True) or {}
    try:
        monthly_income  = float(data.get('monthly_income',  0) or 0)
        monthly_expense = float(data.get('monthly_expense', 0) or 0)
        emergency_fund  = float(data.get('emergency_fund',  0) or 0)
        loan_emi_total  = float(data.get('loan_emi_total',  0) or 0)
    except (TypeError, ValueError):
        monthly_income = monthly_expense = emergency_fund = loan_emi_total = 0
    loan_entries    = data.get('loan_entries', [])  or []
    invest_entries  = data.get('invest_entries', []) or []
    user_name       = str(data.get('user_name', '') or '').strip()
    loan_emi_total  = loan_emi_total  # already set above
    alert_config_raw = data.get('alert_config', []) or []
    # Guard against null items sent by frontend when skipping steps
    alert_config = [a for a in alert_config_raw if isinstance(a, dict)]
    new_password = str(data.get('password', '') or '').strip()

    conn = get_db()
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        month = datetime.now().strftime('%Y-%m')

        # 1. Seed income / expense transactions for current month
        if monthly_income > 0:
            conn.execute(
                "INSERT INTO transactions (type,category,sub_category,amount,date,note,source_ref,auto_created) VALUES (?,?,?,?,?,?,?,?)",
                ('income', 'Salary', 'Onboarding Seed', monthly_income, today,
                 'Onboarding: monthly income baseline', 'onboarding', 1)
            )
        if monthly_expense > 0:
            conn.execute(
                "INSERT INTO transactions (type,category,sub_category,amount,date,note,source_ref,auto_created) VALUES (?,?,?,?,?,?,?,?)",
                ('expense', 'Living', 'Onboarding Seed', monthly_expense, today,
                 'Onboarding: monthly expense baseline', 'onboarding', 1)
            )

        # 2. Seed Emergency Fund asset
        if emergency_fund > 0:
            im = conn.execute(
                "SELECT AssetID FROM InvestMapping WHERE AssetClass='Liquidity' AND AssetCategory='Emergency Fund' AND AssetType='Savings Account'"
            ).fetchone()
            if im:
                asset_id_str = im['AssetID']
                am = conn.execute("SELECT MappingID FROM AssetMapping WHERE AssetId=?", (asset_id_str,)).fetchone()
                if not am:
                    conn.execute(
                        "INSERT INTO AssetMapping (AssetName, AssetSymbol, AssetId) VALUES (?,?,?)",
                        ('Emergency Fund', '', asset_id_str)
                    )
                    conn.commit()
                    am = conn.execute("SELECT MappingID FROM AssetMapping WHERE AssetId=?", (asset_id_str,)).fetchone()
                if am:
                    mid = am['MappingID']
                    existing = conn.execute(
                        "SELECT AssetEntryID FROM assets WHERE MappingID=? AND onboarding_seed=1", (mid,)
                    ).fetchone()
                    if existing:
                        conn.execute(
                            "UPDATE assets SET investedvalue=?, currentvalue=?, updatedat=? WHERE AssetEntryID=?",
                            (emergency_fund, emergency_fund, today, existing['AssetEntryID'])
                        )
                    else:
                        conn.execute(
                            "INSERT INTO assets (MappingID,purpose,qty,avgprice,ltp,investedvalue,currentvalue,pnl,pnlpct,onboarding_seed,updatedat) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                            (mid, 'Emergency Fund', 1, emergency_fund, emergency_fund,
                             emergency_fund, emergency_fund, 0, 0, 1, today)
                        )

        # 3. Seed loan_master rows
        for ln in (loan_entries or []):
            if not isinstance(ln, dict): continue
            name  = str(ln.get('name', '') or '').strip()
            ltype = str(ln.get('loan_type', '') or '').strip()
            if not name or not ltype:
                continue
            existing = conn.execute(
                "SELECT id FROM loan_master WHERE loan_name=? AND status='active'", (name,)
            ).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO loan_master (loan_name,loan_type,loan_amount,total_repayment,start_date,target_close_date,status) VALUES (?,?,?,?,?,?,?)",
                    (name, ltype,
                     float(ln.get('loan_amount', 0) or 0),
                     float(ln.get('total_repayment', 0) or 0),
                     ln.get('start_date', today) or today,
                     ln.get('target_close_date', today) or today,
                     'active')
                )

        # 4. Seed invest_transactions + AssetMapping + assets rows
        # Map CSV asset_type values → InvestMapping.AssetType names
        _ATYPE_MAP = {
            'stocks': 'Stocks', 'stock': 'Stocks',
            'mutual fund': 'Mutual Fund', 'mf': 'Mutual Fund', 'mutualfund': 'Mutual Fund',
            'etf': 'ETF',
            'ppf': 'PPF',
            'epf': 'EPF', 'employer pf': 'Employer PF', 'pf': 'Employer PF',
            'digital gold': 'Digital Gold', 'gold': 'Physical Gold',
            'physical gold': 'Physical Gold',
            'pension': 'Pension Scheme', 'pension scheme': 'Pension Scheme',
            'government bond': 'Government Bond/Yojana', 'bond': 'Government Bond/Yojana',
            'fd': 'Short-term FD', 'short-term fd': 'Short-term FD',
            'liquid mutual fund': 'Liquid Mutual Fund',
            'savings account': 'Savings Account',
            'plot': 'Plot', 'flat': 'Flat',
        }

        for inv in (invest_entries or []):
            if not isinstance(inv, dict): continue
            sname = str(inv.get('stock_name', '') or '').strip()
            if not sname:
                continue
            qty        = float(inv.get('quantity', 0) or 0)
            price      = float(inv.get('price', 0) or 0)
            invested   = qty * price
            raw_atype  = str(inv.get('asset_type', '') or '').strip()
            atype_norm = _ATYPE_MAP.get(raw_atype.lower(), raw_atype)
            entry_date = inv.get('entry_date', today) or today

            # Insert invest_transaction
            conn.execute(
                "INSERT INTO invest_transactions (entry_date,stock_name,asset_type,quantity,action,price,invested_value,current_value,profit,profit_pct,rationale,month) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (entry_date, sname, atype_norm, qty,
                 inv.get('action', 'BUY') or 'BUY',
                 price, invested, invested, 0, 0,
                 inv.get('rationale', 'Onboarding seed') or 'Onboarding seed',
                 month)
            )

            # Find matching InvestMapping row by AssetType
            im = conn.execute(
                "SELECT AssetID FROM InvestMapping WHERE AssetType=? LIMIT 1", (atype_norm,)
            ).fetchone()
            if not im:
                continue  # unknown asset type — skip mapping
            asset_id_str = im['AssetID']

            # Ensure AssetMapping row exists for this asset name + type
            am = conn.execute(
                "SELECT MappingID FROM AssetMapping WHERE AssetName=? AND AssetId=?",
                (sname, asset_id_str)
            ).fetchone()
            if not am:
                conn.execute(
                    "INSERT INTO AssetMapping (AssetName, AssetSymbol, AssetId) VALUES (?,?,?)",
                    (sname, sname, asset_id_str)
                )
                conn.commit()
                am = conn.execute(
                    "SELECT MappingID FROM AssetMapping WHERE AssetName=? AND AssetId=?",
                    (sname, asset_id_str)
                ).fetchone()
            if not am:
                continue

            mid = am['MappingID']
            # Upsert assets row
            existing_asset = conn.execute(
                "SELECT AssetEntryID FROM assets WHERE MappingID=? AND onboarding_seed=1", (mid,)
            ).fetchone()
            if existing_asset:
                conn.execute(
                    "UPDATE assets SET qty=qty+?,avgprice=?,investedvalue=investedvalue+?,currentvalue=currentvalue+?,updatedat=? WHERE AssetEntryID=?",
                    (qty, price, invested, invested, today, existing_asset['AssetEntryID'])
                )
            else:
                conn.execute(
                    "INSERT INTO assets (MappingID,purpose,qty,avgprice,ltp,investedvalue,currentvalue,pnl,pnlpct,onboarding_seed,updatedat,assetname,symbol) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (mid, atype_norm, qty, price, price, invested, invested, 0, 0, 1, today, sname, sname)
                )

            # For Stocks/ETF also seed a Long Term stock_holdings lot
            if atype_norm in ('Stocks', 'ETF') and qty > 0 and price > 0:
                existing_lot = conn.execute(
                    "SELECT id FROM stock_holdings WHERE UPPER(symbol)=UPPER(?) AND trade_type='LONG' AND lot_source='ONBOARDING'",
                    (sname,)
                ).fetchone()
                if not existing_lot:
                    conn.execute(
                        "INSERT INTO stock_holdings (symbol, trade_type, buy_date, qty_original, qty_remaining, buy_price, lot_source) VALUES (?,?,?,?,?,?,?)",
                        (sname, 'LONG', entry_date, qty, qty, price, 'ONBOARDING')
                    )

        # 5. Save user profile to app_settings
        if user_name:
            conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('user_name',?)", (user_name,))
        if monthly_income > 0:
            conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('monthly_income_target',?)",
                         (str(int(monthly_income)),))
        if monthly_expense > 0:
            conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('monthly_expense_target',?)",
                         (str(int(monthly_expense)),))

        # 6. Create smart alerts based on income & user-selected config
        _create_onboarding_alerts(conn, monthly_income, monthly_expense, loan_emi_total, alert_config)

        # 7. Set password if provided during onboarding
        if new_password and len(new_password) >= 6:
            _set_password(new_password, conn=conn)

        # 8. Mark onboarding complete
        conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('onboarding_complete','1')")
        # wizard_done = '1' only for full completion (not skip); controls Re-run button state
        conn.execute("INSERT OR REPLACE INTO app_settings (key,value) VALUES ('onboarding_wizard_done','1')")
        conn.commit()
        return jsonify({'ok': True, 'msg': 'Onboarding complete'})

    except Exception as e:
        import traceback; traceback.print_exc()
        try: conn.rollback()
        except Exception: pass
        return jsonify({'ok': False, 'msg': f'Setup error: {e}'}), 500
    finally:
        conn.close()




# ─────────────────────────────────────────────────────────────────────────────
# CROSS-LINK HELPERS  (called automatically when entries are saved)
# ─────────────────────────────────────────────────────────────────────────────

def _crosslink_investment(conn, txn_id, stock_name, asset_type, qty, price, action, note, entry_date):
    """Mirror an investment transaction into invest_transactions + monthly_investment_calc."""
    month = entry_date[:7] if entry_date else datetime.now().strftime('%Y-%m')
    invested = qty * price if action.upper() == 'BUY' else 0

    conn.execute(
        "INSERT INTO invest_transactions (entry_date,stock_name,asset_type,quantity,action,price,invested_value,current_value,profit,profit_pct,rationale,month) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        (entry_date, stock_name, asset_type, qty, action.upper(), price,
         invested, invested, 0, 0, note or '', month)
    )

    # Upsert monthly_investment_calc
    row = conn.execute(
        "SELECT id, qty_bought, qty_sold, total_invested FROM monthly_investment_calc WHERE month=? AND symbol=?",
        (month, stock_name)
    ).fetchone()
    if row:
        if action.upper() == 'BUY':
            new_qty_b = row['qty_bought'] + qty
            new_inv   = row['total_invested'] + invested
            new_avg   = new_inv / new_qty_b if new_qty_b else 0
            new_net   = new_qty_b - row['qty_sold']
            conn.execute(
                "UPDATE monthly_investment_calc SET qty_bought=?, total_invested=?, avg_buy_price=?, net_qty=?, updated_at=? WHERE id=?",
                (new_qty_b, new_inv, new_avg, new_net, entry_date, row['id'])
            )
        else:
            new_qty_s = row['qty_sold'] + qty
            new_net   = row['qty_bought'] - new_qty_s
            conn.execute(
                "UPDATE monthly_investment_calc SET qty_sold=?, net_qty=?, updated_at=? WHERE id=?",
                (new_qty_s, new_net, entry_date, row['id'])
            )
    else:
        new_avg = price if action.upper() == 'BUY' else 0
        conn.execute(
            "INSERT INTO monthly_investment_calc (month,symbol,asset_type,qty_bought,qty_sold,net_qty,avg_buy_price,total_invested,current_price,current_value,unrealized_pnl,unrealized_pnl_pct,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (month, stock_name, asset_type,
             qty if action.upper() == 'BUY' else 0,
             qty if action.upper() != 'BUY' else 0,
             qty if action.upper() == 'BUY' else -qty,
             new_avg, invested, price, invested, 0, 0, entry_date)
        )


def _crosslink_loan_emi(conn, loan_master_id, amount, date, note):
    """Record an EMI payment as an expense transaction and link it back."""
    lm = conn.execute("SELECT loan_name, loan_type FROM loan_master WHERE id=?", (loan_master_id,)).fetchone()
    if not lm:
        return
    cur = conn.execute(
        "INSERT INTO transactions (type,category,sub_category,amount,date,note,source_ref,auto_created) VALUES (?,?,?,?,?,?,?,?)",
        ('expense', 'Loan EMI', lm['loan_name'], amount, date, note or f'EMI: {lm["loan_name"]}', f'loan:{loan_master_id}', 1)
    )
    txn_id = cur.lastrowid
    conn.execute("INSERT INTO loans (month,loan_type,amount,txn_id) VALUES (?,?,?,?)",
                 (date[:7], lm['loan_type'], amount, txn_id))


def _crosslink_savings_to_emergency(conn, month, savings_amount):
    """Add monthly savings to the Emergency Fund asset's current value."""
    im = conn.execute(
        "SELECT AssetID FROM InvestMapping WHERE AssetClass='Liquidity' AND AssetCategory='Emergency Fund' AND AssetType='Savings Account'"
    ).fetchone()
    if not im:
        return
    am = conn.execute("SELECT MappingID FROM AssetMapping WHERE AssetId=?", (im['AssetID'],)).fetchone()
    if not am:
        return
    existing = conn.execute(
        "SELECT AssetEntryID, currentvalue, investedvalue FROM assets WHERE MappingID=?", (am['MappingID'],)
    ).fetchone()
    today = datetime.now().strftime('%Y-%m-%d')
    if existing:
        new_cv = existing['currentvalue'] + savings_amount
        new_iv = existing['investedvalue'] + savings_amount
        conn.execute(
            "UPDATE assets SET currentvalue=?, investedvalue=?, updatedat=? WHERE AssetEntryID=?",
            (new_cv, new_iv, today, existing['AssetEntryID'])
        )
    else:
        conn.execute(
            "INSERT INTO assets (MappingID,purpose,qty,avgprice,ltp,investedvalue,currentvalue,pnl,pnlpct,updatedat) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (am['MappingID'], 'Emergency Fund', 1, savings_amount, savings_amount,
             savings_amount, savings_amount, 0, 0, today)
        )


def _sync_month_savings_to_ef(conn, month):
    """Recompute month's net savings and set it as the Emergency Fund asset delta."""
    inc = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='income' AND strftime('%Y-%m',date)=? AND auto_created=0",
        (month,)
    ).fetchone()[0]
    exp = conn.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='expense' AND strftime('%Y-%m',date)=? AND auto_created=0",
        (month,)
    ).fetchone()[0]
    savings = inc - exp
    if savings <= 0:
        return
    im = conn.execute(
        "SELECT AssetID FROM InvestMapping WHERE AssetClass='Liquidity' AND AssetCategory='Emergency Fund' AND AssetType='Savings Account'"
    ).fetchone()
    if not im:
        return
    am = conn.execute("SELECT MappingID FROM AssetMapping WHERE AssetId=?", (im[0],)).fetchone()
    if not am:
        return
    today = datetime.now().strftime('%Y-%m-%d')
    existing = conn.execute(
        "SELECT AssetEntryID, investedvalue FROM assets WHERE MappingID=? AND onboarding_seed=0 AND purpose LIKE ?",
        (am[0], f'Monthly Savings {month}%')
    ).fetchone()
    if existing:
        conn.execute(
            "UPDATE assets SET investedvalue=?, currentvalue=?, updatedat=? WHERE AssetEntryID=?",
            (savings, savings, today, existing[0])
        )
    else:
        conn.execute(
            "INSERT INTO assets (MappingID,purpose,qty,avgprice,ltp,investedvalue,currentvalue,pnl,pnlpct,onboarding_seed,updatedat) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (am[0], f'Monthly Savings {month}', 1, savings, savings, savings, savings, 0, 0, 0, today)
        )


# ─────────────────────────────────────────────────────────────────────────────
# ENHANCED TRANSACTION SAVE  — also fires cross-links
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/transactions/add_v2', methods=['POST'])
@login_required
def api_add_transaction_v2():
    """
    Enhanced transaction endpoint that auto-mirrors investment entries to
    invest_transactions and monthly_investment_calc, and adds loan EMIs to
    the loans table.

    Accepts same payload as /api/transactions (POST) plus optional fields:
      stock_name, asset_type, quantity, action  (for type='investment')
      loan_master_id                             (for type='expense' + category='Loan EMI')
    """
    data    = request.get_json(force=True) or {}
    txn_type = data.get('type', '').strip()
    category = data.get('category', '').strip()
    amount   = float(data.get('amount', 0))
    date     = data.get('date', datetime.now().strftime('%Y-%m-%d'))
    note     = data.get('note', '')
    sub_cat  = data.get('sub_category', '')

    conn = get_db()
    cur = conn.execute(
        "INSERT INTO transactions (type,category,sub_category,amount,date,note,source_ref,auto_created) VALUES (?,?,?,?,?,?,?,?)",
        (txn_type, category, sub_cat, amount, date, note, data.get('source_ref', ''), 0)
    )
    txn_id = cur.lastrowid

    # Mirror investment entry
    if txn_type == 'investment':
        stock_name = data.get('stock_name', sub_cat or category)
        asset_type = data.get('asset_type', '')
        qty        = float(data.get('quantity', 1))
        price      = float(data.get('price', amount))
        action     = data.get('action', 'BUY')
        _crosslink_investment(conn, txn_id, stock_name, asset_type, qty, price, action, note, date)

    # Mirror loan EMI
    if txn_type == 'expense' and category.lower() in ('loan emi', 'loan_emi', 'emi'):
        loan_master_id = data.get('loan_master_id')
        if loan_master_id:
            conn.execute("INSERT INTO loans (month,loan_type,amount,txn_id) VALUES (?,?,?,?)",
                         (date[:7], category, amount, txn_id))

    # Compute and apply month-end savings → Emergency Fund
    # Trigger only when a savings-marker field is passed
    if data.get('apply_savings_to_ef') and txn_type == 'income':
        month = date[:7]
        inc = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as s FROM transactions WHERE type='income' AND strftime('%Y-%m',date)=?", (month,)
        ).fetchone()['s']
        exp = conn.execute(
            "SELECT COALESCE(SUM(amount),0) as s FROM transactions WHERE type='expense' AND strftime('%Y-%m',date)=?", (month,)
        ).fetchone()['s']
        savings = inc - exp
        if savings > 0:
            _crosslink_savings_to_emergency(conn, month, savings)

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': txn_id})


# ─────────────────────────────────────────────────────────────────────────────
# ONBOARDING SAMPLE CSV FILES
# ─────────────────────────────────────────────────────────────────────────────

_LOANS_SAMPLE_CSV = (
    "loan_name,loan_type,loan_amount,monthly_emi,total_repayment,start_date,target_close_date\n"
    "Home Loan - SBI,Home Loan,3000000,28000,5500000,2021-04-01,2041-03-31\n"
    "Car Loan - HDFC,Car Loan,600000,11500,690000,2023-06-01,2028-05-31\n"
    "Personal Loan - Axis,Personal Loan,200000,7200,259200,2024-01-15,2027-01-14\n"
)

_INVESTS_SAMPLE_CSV = (
    "name,asset_type,quantity,buy_price,buy_date\n"
    "RELIANCE,Stocks,50,2450.00,2023-01-10\n"
    "INFY,Stocks,100,1520.50,2022-08-15\n"
    "SBI Bluechip Fund,Mutual Fund,500,62.30,2023-03-01\n"
    "NIFTYBEES,ETF,200,235.00,2023-05-20\n"
    "Digital Gold,Digital Gold,10,6200.00,2024-02-14\n"
    "PPF Account,PPF,1,150000.00,2020-04-01\n"
)

@app.route('/api/onboarding/sample/<filename>')
@login_required
def api_onboarding_sample(filename):
    if filename == 'loans.csv':
        content = _LOANS_SAMPLE_CSV
    elif filename == 'investments.csv':
        content = _INVESTS_SAMPLE_CSV
    else:
        return 'Not found', 404
    from flask import Response
    return Response(
        content,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


# ─────────────────────────────────────────────────────────────────────────────
# BACKUP & RESTORE
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/backup/download')
@login_required
def api_backup_download():
    """Stream the SQLite DB file as a timestamped download."""
    import shutil, io
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    name = 'moneytracker_backup_{}.db'.format(ts)
    # Copy DB to a temp file so we don't stream while SQLite has it open
    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp.close()
    try:
        shutil.copy2(DB_PATH, tmp.name)
        return send_file(
            tmp.name,
            as_attachment=True,
            download_name=name,
            mimetype='application/octet-stream'
        )
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/backup/restore_selective', methods=['POST'])
@login_required
def api_backup_restore_selective():
    """Restore only selected scopes from a backup file.
    Form fields: db_file (file), scopes (comma-separated scope names).
    Scope → tables mapping mirrors the Import Wizard JS config.
    Always preserves app_settings and InvestMapping.
    """
    SCOPE_TABLES = {
        'wizard_money':  ['transactions', 'loans', 'loan_master', 'alerts'],
        'wizard_wealth': ['assets', 'portfolio', 'AssetMapping', 'invest_transactions',
                          'monthly_investment_calc', 'demat', 'nse_master'],
        'wizard_vision': ['um_vision_cards', 'magnet_status'],
        'wizard_docs':   ['doc_wallet', 'doc_attachments', 'doc_wallet_types'],
    }
    ALWAYS_PRESERVE = {'app_settings', 'InvestMapping'}

    f = request.files.get('db_file')
    if not f:
        return jsonify({'ok': False, 'error': 'No file uploaded'}), 400
    header = f.read(16)
    if not header.startswith(b'SQLite format 3'):
        return jsonify({'ok': False, 'error': 'Not a valid SQLite database file'}), 400
    f.seek(0)

    raw_scopes = request.form.get('scopes', '')
    selected_scopes = [s.strip() for s in raw_scopes.split(',') if s.strip() in SCOPE_TABLES]
    if not selected_scopes:
        return jsonify({'ok': False, 'error': 'No valid scopes selected'}), 400

    tables_to_restore = []
    for scope in selected_scopes:
        tables_to_restore.extend(SCOPE_TABLES[scope])

    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        f.save(tmp_path)
        import sqlite3 as _sq3
        src = _sq3.connect(tmp_path)
        src.row_factory = _sq3.Row
        ok = src.execute('PRAGMA integrity_check').fetchone()
        if ok and ok[0] != 'ok':
            src.close()
            raise ValueError(f'Integrity check failed: {ok[0]}')

        dst = _sq3.connect(DB_PATH)
        dst.row_factory = _sq3.Row
        dc = dst.cursor()

        results = {}
        for tbl in tables_to_restore:
            if tbl in ALWAYS_PRESERVE:
                results[tbl] = 'skipped (protected)'
                continue
            try:
                # Check table exists in backup
                exists = src.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (tbl,)
                ).fetchone()
                if not exists:
                    results[tbl] = 'not found in backup'
                    continue
                rows = src.execute(f'SELECT * FROM {tbl}').fetchall()
                dc.execute(f'DELETE FROM {tbl}')
                if rows:
                    cols = rows[0].keys()
                    col_list = ','.join(cols)
                    placeholders = ','.join('?' * len(cols))
                    for row in rows:
                        dc.execute(f'INSERT OR IGNORE INTO {tbl} ({col_list}) VALUES ({placeholders})', list(row))
                results[tbl] = f'restored {len(rows)} rows'
            except Exception as e:
                results[tbl] = f'error: {e}'

        dst.commit()
        dst.close()
        src.close()
        init_schema()
        return jsonify({'ok': True, 'results': results})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        try: os.unlink(tmp_path)
        except: pass


@app.route('/api/backup/restore_onboarding', methods=['POST'])
def api_backup_restore_onboarding():
    """Restore DB from backup during onboarding (no login required).
    Preserves app_settings (credentials set during onboarding step 1)."""
    f = request.files.get('db_file')
    if not f:
        return jsonify({'ok': False, 'error': 'No file uploaded'}), 400
    header = f.read(16)
    if not header.startswith(b'SQLite format 3'):
        return jsonify({'ok': False, 'error': 'Not a valid SQLite database file'}), 400
    f.seek(0)
    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        f.save(tmp_path)
        import sqlite3 as _sq3
        src = _sq3.connect(tmp_path)
        ok  = src.execute('PRAGMA integrity_check').fetchone()
        if ok and ok[0] != 'ok':
            src.close()
            raise ValueError(f'Integrity check failed: {ok[0]}')
        # Snapshot app_settings (holds the just-set credentials) + InvestMapping
        live = _sq3.connect(DB_PATH)
        live.row_factory = _sq3.Row
        lc = live.cursor()
        preserved = {}
        for tbl in ('app_settings', 'InvestMapping'):
            try:
                preserved[tbl] = lc.execute(f'SELECT * FROM {tbl}').fetchall()
            except Exception:
                preserved[tbl] = []
        live.close()
        dst = _sq3.connect(DB_PATH)
        src.backup(dst, pages=100, progress=None)
        dst.close()
        src.close()
        init_schema()
        # Re-apply preserved tables
        post = _sq3.connect(DB_PATH)
        pc = post.cursor()
        for tbl, rows in preserved.items():
            if not rows:
                continue
            try:
                pc.execute(f'DELETE FROM {tbl}')
                cols = rows[0].keys()
                placeholders = ','.join('?' * len(cols))
                col_list = ','.join(cols)
                for row in rows:
                    pc.execute(f'INSERT OR REPLACE INTO {tbl} ({col_list}) VALUES ({placeholders})', list(row))
            except Exception:
                pass
        post.commit()
        post.close()
        return jsonify({'ok': True, 'msg': 'Database restored successfully!'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        try: os.unlink(tmp_path)
        except: pass


@app.route('/api/backup/restore', methods=['POST'])
@login_required
def api_backup_restore():
    """Replace the current DB with an uploaded backup file."""
    f = request.files.get('db_file')
    if not f:
        return jsonify({'ok': False, 'error': 'No file uploaded'}), 400

    # Validate it's a SQLite file (first 16 bytes = SQLite magic)
    header = f.read(16)
    if not header.startswith(b'SQLite format 3'):
        return jsonify({'ok': False, 'error': 'Not a valid SQLite database file'}), 400
    f.seek(0)

    # Save upload to a temp file (close handle first — Windows requires it)
    tmp = tempfile.NamedTemporaryFile(suffix='.db', delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        f.save(tmp_path)

        import sqlite3 as _sq3

        # Integrity check on the uploaded file
        src = _sq3.connect(tmp_path)
        ok  = src.execute('PRAGMA integrity_check').fetchone()
        if ok and ok[0] != 'ok':
            src.close()
            raise ValueError(f'Integrity check failed: {ok[0]}')

        # Snapshot tables that must NOT be overwritten from the backup
        live = _sq3.connect(DB_PATH)
        live.row_factory = _sq3.Row
        lc = live.cursor()
        preserved = {}
        for tbl in ('app_settings', 'InvestMapping'):
            try:
                preserved[tbl] = lc.execute(f'SELECT * FROM {tbl}').fetchall()
            except Exception:
                preserved[tbl] = []
        live.close()

        # Use SQLite online backup API — copies directly into the live DB file
        # while the server is running, no file lock conflict on Windows
        dst = _sq3.connect(DB_PATH)
        src.backup(dst, pages=100, progress=None)
        dst.close()
        src.close()

        # Re-run schema migrations so any new columns/tables are added
        init_schema()

        # Restore the preserved tables back (overwrite whatever came from backup)
        post = _sq3.connect(DB_PATH)
        pc = post.cursor()
        for tbl, rows in preserved.items():
            if not rows:
                continue
            try:
                pc.execute(f'DELETE FROM {tbl}')
                cols = rows[0].keys()
                placeholders = ','.join('?' * len(cols))
                col_list = ','.join(cols)
                for row in rows:
                    pc.execute(f'INSERT OR REPLACE INTO {tbl} ({col_list}) VALUES ({placeholders})', list(row))
            except Exception:
                pass
        post.commit()
        post.close()

        return jsonify({'ok': True, 'msg': 'Database restored successfully!'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        try: os.unlink(tmp_path)
        except: pass


@app.route('/api/reset/fresh', methods=['POST'])
@login_required
def api_reset_fresh():
    """Wipe all user data and reset onboarding so the app behaves like a new install."""
    data     = request.get_json(silent=True) or {}
    confirm  = data.get('confirm', '')
    if confirm != 'RESET':
        return jsonify({'ok': False, 'error': 'Send {"confirm":"RESET"} to proceed'}), 400

    _ALL_USER_TABLES = [
        # Money Tracker
        'transactions', 'loans', 'loan_master', 'alerts',
        # Wealth Engine
        'invest_transactions', 'monthly_investment_calc',
        'assets', 'portfolio', 'AssetMapping',
        'raw_upload_data', 'raw_upload_meta',
        'nse_master',
        'demat_wallet', 'stock_holdings', 'stock_transactions',
        'stock_dividends', 'stock_pnl',
        'wealth',
        # Vision & Magnet
        'um_vision_cards', 'magnet_status',
        # Docs Wallet
        'doc_wallet', 'doc_attachments', 'doc_wallet_types',
    ]
    conn = get_db()
    try:
        conn.execute("PRAGMA foreign_keys = OFF")
        for t in _ALL_USER_TABLES:
            try:
                conn.execute('DELETE FROM [{}]'.format(t))
            except Exception:
                pass
        # Clear user app_settings but keep password, theme, currency, username
        conn.execute("""DELETE FROM app_settings
                        WHERE key NOT IN ('app_password','theme','currency','user_name')""")
        conn.execute("PRAGMA foreign_keys = ON")
        _restore_seed_data(conn)
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True, 'msg': 'All user data cleared. Onboarding will restart on next page load.'})


if __name__ == '__main__':
    # Auto-install yfinance if missing (YF_AVAILABLE is module-level, no global needed here)
    if not YF_AVAILABLE:
        try:
            import subprocess, sys as _sys
            print("📦 Installing yfinance (required for NSE sync)…")
            subprocess.check_call(
                [_sys.executable, '-m', 'pip', 'install', 'yfinance', '-q'],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
            )
            import yfinance as yf
            YF_AVAILABLE = True
            print("✅ yfinance installed successfully")
        except Exception as _e:
            print(f"⚠️  Could not auto-install yfinance: {_e}")
            print("   Run manually: pip install yfinance")
    init_schema()
    # Bind to '' so Flask listens on both IPv4 (0.0.0.0) and IPv6 (::)
    # Fixes macOS/Safari "localhost denied" when OS resolves localhost → ::1
    app.run(host='', port=9876, debug=False)
